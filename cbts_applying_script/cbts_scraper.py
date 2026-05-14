#!/usr/bin/env python3
"""Standalone CBTS job scraper (Salesforce-based job board)."""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional
from urllib.parse import urljoin

import groq
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


JOBBOARD_BASE = "https://cbts.my.salesforce-sites.com"
JOBBOARD_SEARCH = f"{JOBBOARD_BASE}/jobboard/Jobsearch"

DEFAULT_FUNCTION = "Technical"
DEFAULT_SEARCH_TERMS = [
    "python",
    "ai",
    "machine learning",
    "software engineer",
    "backend",
    "full stack",
    "cloud",
    "data scientist",
]

EMAIL_RE = re.compile(r"(?<![A-Za-z0-9._%+-])([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})(?![A-Za-z0-9._%+-])")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?1[\s.-]?)?(?:\(?\d{3}\)?[\s.-]?)\d{3}[\s.-]?\d{4}(?!\d)")

TITLE_RANKING_WEIGHTS = [
    ("machine learning engineer", 56),
    ("ml engineer", 54),
    ("ai engineer", 54),
    ("ai/ml", 54),
    ("aiml", 54),
    ("generative ai", 52),
    ("gen ai", 52),
    ("llm", 50),
    ("rag", 48),
    ("senior data scientist", 52),
    ("data scientist", 48),
    ("python developer", 46),
    ("python engineer", 46),
    ("senior python", 50),
    ("django", 44),
    ("fastapi", 44),
    ("backend", 38),
    ("back end", 38),
    ("api developer", 36),
    ("software engineer python", 46),
    ("full stack software engineer", 42),
    ("full stack engineer", 40),
    ("full stack developer", 40),
    ("cloud engineer", 30),
    ("aws", 28),
    ("azure", 28),
    ("senior software engineer", 32),
    ("software engineer", 24),
    ("software developer", 24),
    ("developer", 12),
    ("engineer", 10),
]
MIN_TITLE_RANK = 20

TITLE_EXCLUSION_WEIGHTS = [
    ("frontend", -30),
    ("front end", -30),
    ("ui developer", -20),
    ("ux/ui", -12),
    ("java", -12),
    (".net", -35),
    ("qa", -30),
    ("quality assurance", -30),
    ("desktop support", -40),
    ("help desk", -40),
    ("project manager", -35),
    ("business analyst", -30),
    ("data engineer", -100),
    ("network engineer", -50),
    ("civil engineer", -100),
    ("mechanical engineer", -100),
    ("electrical engineer", -100),
    ("structural engineer", -100),
    ("environmental", -80),
    ("protection", -60),
    ("infrastructure engineer", -50),
    ("systems administrator", -50),
    ("sysadmin", -50),
    ("devops engineer", -20),
    ("site reliability", -20),
    ("scrum master", -60),
    ("product manager", -60),
    ("product owner", -60),
    ("it support", -60),
    ("network administrator", -80),
    ("security engineer", -30),
    ("cybersecurity", -30),
    ("sap", -50),
    ("salesforce developer", -30),
    ("salesforce engineer", -30),
    ("ios developer", -50),
    ("android developer", -50),
    ("mobile developer", -50),
    ("embedded", -40),
]
JAVA_TITLE_PATTERNS = [
    (re.compile(r"\bjava\b.*\bfull\s*stack\b|\bfull\s*stack\b.*\bjava\b|\bjava\b.*\bfullstack\b|\bfullstack\b.*\bjava\b", re.I), "Java full stack title"),
    (re.compile(r"\bjava\b.*\b(?:developer|engineer|architect|backend|software)\b|\b(?:developer|engineer|architect|backend|software)\b.*\bjava\b", re.I), "Java developer/engineer title"),
]
DISALLOWED_WORK_PATTERNS = [
    (re.compile(r"\bno\s+c2c\b", re.I), "No C2C"),
    (re.compile(r"\bno\s+corp(?:oration)?\s*[- ]?\s*to\s*[- ]?\s*corp(?:oration)?\b", re.I), "No corp-to-corp"),
    (re.compile(r"\bf\s*2\s*f\b", re.I), "F2F"),
    (re.compile(r"\bface\s*[- ]?\s*to\s*[- ]?\s*face\b", re.I), "Face-to-face"),
    (re.compile(r"\bin\s*[- ]?\s*person\s+interview\b", re.I), "In-person interview"),
    (re.compile(r"\bon\s*[- ]?\s*site\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\bonsite\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\bfull\s*[- ]?\s*time\b", re.I), "Full-time"),
    (re.compile(r"\bpermanent\b", re.I), "Permanent"),
    (re.compile(r"\bdirect\s*[- ]?\s*hire\b", re.I), "Direct hire"),
]


@dataclass
class CBTSJob:
    source_company: str
    search_term: str
    title_rank: int
    title_rank_reasons: str
    title: str
    state: str
    regional_area: str
    location: str
    primary_background: str
    posted_date: str
    job_id: str
    job_number: str
    job_url: str
    contact_info: str
    description_snippet: str
    raw_text: str


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip()).strip()


def normalized_title(text: str) -> str:
    return re.sub(r"[^a-z0-9+#./-]+", " ", (text or "").lower()).strip()


def title_contains(title: str, phrase: str) -> bool:
    phrase = normalized_title(phrase)
    if not phrase:
        return False
    if len(phrase) <= 3:
        return re.search(rf"(?<![a-z0-9]){re.escape(phrase)}(?![a-z0-9])", title) is not None
    return phrase in title


def score_title(title: str) -> tuple[int, str]:
    normalized = normalized_title(title)
    score = 0
    reasons: list[str] = []
    for phrase, weight in TITLE_RANKING_WEIGHTS:
        if title_contains(normalized, phrase):
            score += weight
            reasons.append(f"{phrase}+{weight}")
    for phrase, weight in TITLE_EXCLUSION_WEIGHTS:
        if title_contains(normalized, phrase):
            score += weight
            reasons.append(f"{phrase}{weight}")
    return max(score, 0), "; ".join(reasons)


def disallowed_work_reasons(text: str) -> list[str]:
    cleaned = clean_text(text)
    reasons: list[str] = []
    for pattern, reason in DISALLOWED_WORK_PATTERNS:
        if pattern.search(cleaned) and reason not in reasons:
            reasons.append(reason)
    return reasons


def java_title_reasons(title: str) -> list[str]:
    return [reason for pattern, reason in JAVA_TITLE_PATTERNS if pattern.search(title or "")]


def contact_info_from_text(text: str) -> str:
    emails: list[str] = []
    for email in EMAIL_RE.findall(clean_text(text)):
        lowered = email.lower()
        if lowered not in emails:
            emails.append(lowered)
    phones: list[str] = []
    for phone in PHONE_RE.findall(text or ""):
        digits = re.sub(r"\D", "", phone)
        if len(digits) == 11 and digits.startswith("1"):
            digits = digits[1:]
        if len(digits) == 10:
            display = f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
            if display not in phones:
                phones.append(display)
    return ", ".join(emails + phones)


def parse_cbts_date(value: str) -> Optional[datetime]:
    value = clean_text(value)
    if not value:
        return None
    for fmt in ("%m/%d/%Y", "%B %d, %Y", "%b %d, %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    return None


def job_posted_day(job: CBTSJob) -> str:
    parsed = parse_cbts_date(job.posted_date)
    return parsed.date().isoformat() if parsed else "unknown-date"


def is_within_posted_days(value: str, days: Optional[int], now: Optional[datetime] = None) -> bool:
    if not days or days <= 0:
        return True
    parsed = parse_cbts_date(value)
    if not parsed:
        return False
    now = now or datetime.now(timezone.utc)
    return 0 <= (now - parsed).total_seconds() <= days * 24 * 60 * 60


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 VenkataDoraCBTSAutomation/0.1 (+local personal job search)",
        "Accept": "text/html,application/xhtml+xml,*/*",
        "Accept-Language": "en-US,en;q=0.9",
    })
    return session


def _extract_viewstate(soup: BeautifulSoup) -> dict[str, str]:
    state: dict[str, str] = {}
    for name in (
        "com.salesforce.visualforce.ViewState",
        "com.salesforce.visualforce.ViewStateVersion",
        "com.salesforce.visualforce.ViewStateMAC",
    ):
        tag = soup.find("input", {"name": name})
        if tag and tag.get("value"):
            state[name] = tag["value"]
    return state


def _extract_form_prefix(soup: BeautifulSoup) -> str:
    tag = soup.find("input", {"id": re.compile(r"j_id\d+:j_id\d+:j_id\d+$")})
    if tag and tag.get("name"):
        return tag["name"]
    return "j_id0:j_id1:j_id27"


def fetch_search_page(session: requests.Session, timeout: int) -> tuple[BeautifulSoup, dict[str, str], str]:
    response = session.get(JOBBOARD_SEARCH, timeout=timeout)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    viewstate = _extract_viewstate(soup)
    prefix = _extract_form_prefix(soup)
    return soup, viewstate, prefix


def post_search(
    session: requests.Session,
    viewstate: dict[str, str],
    prefix: str,
    keyword: str,
    function: str,
    timeout: int,
) -> BeautifulSoup:
    data: dict[str, str] = {
        prefix: prefix,
        f"{prefix}:keywordtxt": keyword,
        f"{prefix}:statara": "",
        f"{prefix}:functionpick": function,
        f"{prefix}:regareapick": "",
        f"{prefix}:prmrybackpick": "",
        f"{prefix}:srchbutton": "Search",
    }
    data.update(viewstate)
    response = session.post(
        JOBBOARD_SEARCH,
        data=data,
        headers={"Referer": JOBBOARD_SEARCH, "Content-Type": "application/x-www-form-urlencoded"},
        timeout=timeout,
    )
    response.raise_for_status()
    return BeautifulSoup(response.text, "html.parser")


def parse_search_rows(soup: BeautifulSoup) -> list[dict[str, str]]:
    table = soup.select_one("table.atsSearchResultTable")
    if not table:
        return []
    rows: list[dict[str, str]] = []
    for tr in table.select("tbody tr"):
        cells = tr.find_all("td")
        if len(cells) < 5:
            continue
        title_link = cells[1].select_one("a[href]")
        if not title_link:
            continue
        href = title_link.get("href", "")
        job_id_match = re.search(r"JobId=([^&]+)", href)
        job_id = job_id_match.group(1) if job_id_match else ""
        job_url = urljoin(JOBBOARD_BASE, href) if href else ""
        rows.append({
            "posted_date": clean_text(cells[0].get_text(" ", strip=True)),
            "title": clean_text(title_link.get_text(" ", strip=True)),
            "state": clean_text(cells[2].get_text(" ", strip=True)),
            "regional_area": clean_text(cells[3].get_text(" ", strip=True)),
            "primary_background": clean_text(cells[4].get_text(" ", strip=True)),
            "job_id": job_id,
            "job_url": job_url,
        })
    return rows


def fetch_job_detail(session: requests.Session, job_id: str, timeout: int) -> dict[str, str]:
    url = f"{JOBBOARD_SEARCH}?JobId={job_id}"
    response = session.get(url, timeout=timeout)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    detail_table = soup.select_one("table.atsJobDetailsTable")
    if not detail_table:
        return {}
    text = clean_text(detail_table.get_text(" ", strip=True))
    job_number = ""
    m = re.search(r"Job Number[:\s]+([^\s]+)", text)
    if m:
        job_number = clean_text(m.group(1))
    desc_start = text.find("Job Description")
    description = clean_text(text[desc_start + len("Job Description"):]) if desc_start != -1 else text
    return {"job_number": job_number, "description": description, "job_url": url}


def normalize_job(row: dict[str, str], detail: dict[str, str], search_term: str) -> CBTSJob:
    title = row["title"]
    description = detail.get("description") or ""
    title_rank, title_rank_reasons = score_title(title)
    location = ", ".join(part for part in (row["state"], row["regional_area"]) if part)
    return CBTSJob(
        source_company="CBTS",
        search_term=search_term,
        title_rank=title_rank,
        title_rank_reasons=title_rank_reasons,
        title=title,
        state=row["state"],
        regional_area=row["regional_area"],
        location=location,
        primary_background=row["primary_background"],
        posted_date=row["posted_date"],
        job_id=row["job_id"],
        job_number=detail.get("job_number") or "",
        job_url=row["job_url"],
        contact_info=contact_info_from_text(description),
        description_snippet=description[:900],
        raw_text=description,
    )


def sort_jobs(jobs: Iterable[CBTSJob]) -> list[CBTSJob]:
    return sorted(
        jobs,
        key=lambda job: (job_posted_day(job), bool(job.contact_info), job.title_rank, job.title.lower()),
        reverse=True,
    )


GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")
GROQ_MODEL = "llama-3.1-8b-instant"

_AI_FILTER_SYSTEM = (
    "You filter job listings for a candidate. Profile: Senior Python Developer, "
    "AI/ML Engineer, Data Scientist, Machine Learning Engineer, Software Engineer. "
    "They use Python, AI/ML frameworks, data science tools. "
    "They will NOT do: pure data engineering/ETL pipelines, frontend/UI, Java/.NET, "
    "QA/testing, project management, network/civil/mechanical/electrical engineering, "
    "DevOps/SRE, cybersecurity, mobile, embedded, SAP, Salesforce admin, IT support."
)

_AI_FILTER_PROMPT = """Below is a JSON list of jobs. For each, decide if it is RELEVANT or IRRELEVANT for the candidate.

Reply with ONLY a JSON array in this exact format:
[{{"id": "JOB_ID", "verdict": "RELEVANT", "reason": "one short phrase"}}, ...]

Jobs:
{jobs_json}"""


def ai_filter_jobs(jobs: list[CBTSJob], batch_size: int = 20) -> list[CBTSJob]:
    if not GROQ_API_KEY:
        print("Skipping AI filter: GROQ_API_KEY is not set.")
        return jobs
    client = groq.Groq(api_key=GROQ_API_KEY)
    kept: list[CBTSJob] = []

    for batch_start in range(0, len(jobs), batch_size):
        batch = jobs[batch_start: batch_start + batch_size]
        jobs_payload = [
            {"id": j.job_id, "title": j.title, "snippet": j.description_snippet[:300]}
            for j in batch
        ]
        prompt = _AI_FILTER_PROMPT.format(jobs_json=json.dumps(jobs_payload, indent=2))
        try:
            response = client.chat.completions.create(
                model=GROQ_MODEL,
                messages=[
                    {"role": "system", "content": _AI_FILTER_SYSTEM},
                    {"role": "user", "content": prompt},
                ],
                max_tokens=1024,
                temperature=0,
            )
            raw = response.choices[0].message.content.strip()
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
            verdicts: list[dict[str, str]] = json.loads(raw)
            verdict_map = {v["id"]: v["verdict"].upper() for v in verdicts}
        except Exception as exc:
            print(f"AI filter batch failed: {exc} — keeping all {len(batch)} jobs in this batch")
            kept.extend(batch)
            continue

        for job in batch:
            verdict = verdict_map.get(job.job_id, "RELEVANT")
            if verdict == "RELEVANT":
                kept.append(job)
            else:
                reason = next((v.get("reason", "") for v in verdicts if v.get("id") == job.job_id), "")
                print(f"AI excluded {job.job_number or job.job_id}: {job.title} — {reason}")

    return kept


def scrape_cbts(
    search_terms: Iterable[str],
    function_filter: str,
    posted_within_days: Optional[int],
    exclude_disallowed_work: bool,
    use_ai_filter: bool,
    timeout: int,
    sleep_seconds: float,
) -> list[CBTSJob]:
    session = make_session()
    seen_ids: set[str] = set()
    jobs: list[CBTSJob] = []

    print("Fetching CBTS job board search page...")
    try:
        _, viewstate, prefix = fetch_search_page(session, timeout)
    except requests.RequestException as exc:
        print(f"Failed to load CBTS search page: {exc}")
        return []

    all_rows: list[tuple[str, dict[str, str]]] = []

    for term in search_terms:
        term = term.strip()
        if not term:
            continue
        print(f"Searching CBTS: {term!r} (Function={function_filter})")
        try:
            soup = post_search(session, viewstate, prefix, term, function_filter, timeout)
        except requests.RequestException as exc:
            print(f"Search failed for {term!r}: {exc}")
            time.sleep(sleep_seconds)
            continue

        rows = parse_search_rows(soup)
        for row in rows:
            if not row["job_id"] or row["job_id"] in seen_ids:
                continue
            if not is_within_posted_days(row["posted_date"], posted_within_days):
                continue
            seen_ids.add(row["job_id"])
            all_rows.append((term, row))
        time.sleep(sleep_seconds)

    for term, row in all_rows:
        try:
            detail = fetch_job_detail(session, row["job_id"], timeout)
        except requests.RequestException as exc:
            print(f"Detail failed for {row['job_id']}: {exc}")
            detail = {}
        job = normalize_job(row, detail, term)
        title_reasons = java_title_reasons(job.title)
        if title_reasons:
            print(f"Excluded {job.job_number or job.job_id}: {', '.join(title_reasons)} — {job.title}")
            continue
        if job.title_rank < MIN_TITLE_RANK:
            print(f"Excluded {job.job_number or job.job_id}: low title rank ({job.title_rank}) — {job.title}")
            continue
        if exclude_disallowed_work and disallowed_work_reasons(job.raw_text):
            reasons = ", ".join(disallowed_work_reasons(job.raw_text))
            print(f"Excluded {job.job_number or job.job_id}: {reasons}")
            continue
        jobs.append(job)
        time.sleep(sleep_seconds)

    print(f"\n--- {len(jobs)} jobs passed rule filter ---")
    for j in jobs:
        print(f"  [{j.title_rank:3d}] {j.title} | {j.location}")
    print()

    if use_ai_filter and jobs:
        print(f"AI filtering {len(jobs)} jobs with groq/llama...")
        jobs = ai_filter_jobs(jobs)
        print(f"AI filter kept {len(jobs)} jobs.")

    return sort_jobs(jobs)


def write_csv(jobs: list[CBTSJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(CBTSJob.__dataclass_fields__.keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for job in jobs:
            writer.writerow(asdict(job))


def write_json(jobs: list[CBTSJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump([asdict(job) for job in jobs], handle, indent=2)


EXCEL_COLUMNS = [
    ("posted_day", "Posted Day"),
    ("title_rank", "Title Rank"),
    ("title", "Title"),
    ("location", "Location"),
    ("primary_background", "Background"),
    ("contact_info", "Contact Info"),
    ("search_term", "Search Term"),
    ("job_number", "Job Number"),
    ("job_url", "Job URL"),
    ("title_rank_reasons", "Rank Reasons"),
]


def excel_row(job: CBTSJob) -> dict[str, Any]:
    row = asdict(job)
    row["posted_day"] = job_posted_day(job)
    return row


def style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="003366")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [13, 11, 34, 28, 22, 30, 20, 16, 14, 45]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        rank = row[1].value or 0
        if isinstance(rank, int) and rank >= 45:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="E8F5E9")
        if row[5].value:
            row[5].fill = PatternFill("solid", fgColor="FFF3CD")


def append_jobs_sheet(workbook: Workbook, sheet_name: str, jobs: list[CBTSJob]) -> None:
    safe_name = re.sub(r"[\[\]:*?/\\]", "-", sheet_name)[:31] or "Sheet"
    sheet = workbook.create_sheet(safe_name)
    sheet.append([label for _, label in EXCEL_COLUMNS])
    for job in jobs:
        source = excel_row(job)
        sheet.append([source.get(key, "") for key, _ in EXCEL_COLUMNS])
        url_col = next((i + 1 for i, (k, _) in enumerate(EXCEL_COLUMNS) if k == "job_url"), None)
        if url_col and job.job_url:
            cell = sheet.cell(row=sheet.max_row, column=url_col)
            cell.value = "Open Job"
            cell.hyperlink = job.job_url
            cell.style = "Hyperlink"
    style_sheet(sheet)


def write_excel(jobs: list[CBTSJob], path: Path, posted_within_days: Optional[int]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "CBTS Daily Jobs"
    summary["A1"].font = Font(size=18, bold=True, color="003366")
    summary["A3"] = "Generated"
    summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary["A4"] = "Posting Window"
    summary["B4"] = "All dates" if not posted_within_days else f"Last {posted_within_days} days"
    summary["A5"] = "Total Jobs"
    summary["B5"] = len(jobs)
    summary.append([])
    summary.append(["Day", "Jobs", "With Contact", "Top Rank"])
    for cell in summary[7]:
        cell.fill = PatternFill("solid", fgColor="003366")
        cell.font = Font(color="FFFFFF", bold=True)
    grouped: dict[str, list[CBTSJob]] = {}
    for job in jobs:
        grouped.setdefault(job_posted_day(job), []).append(job)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        summary.append([day, len(day_jobs), sum(1 for j in day_jobs if j.contact_info), max((j.title_rank for j in day_jobs), default=0)])
    for col in range(1, 5):
        summary.column_dimensions[get_column_letter(col)].width = 22
    append_jobs_sheet(workbook, "All Jobs", jobs)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        append_jobs_sheet(workbook, day, sorted(day_jobs, key=lambda j: (not bool(j.contact_info), -j.title_rank, j.title.lower())))
    workbook.save(path)


def write_daily_outputs(jobs: list[CBTSJob], out_dir: Path, timestamp: str) -> Path:
    daily_dir = out_dir / f"daily_{timestamp}"
    grouped: dict[str, list[CBTSJob]] = {}
    for job in jobs:
        grouped.setdefault(job_posted_day(job), []).append(job)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        write_csv(day_jobs, daily_dir / f"{day}_jobs.csv")
        write_json(day_jobs, daily_dir / f"{day}_jobs.json")
    return daily_dir


def load_terms(args: argparse.Namespace) -> list[str]:
    terms = list(args.terms or [])
    if args.terms_file:
        terms.extend(line.strip() for line in args.terms_file.read_text(encoding="utf-8").splitlines())
    return [t for t in terms if t.strip()] or DEFAULT_SEARCH_TERMS


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape CBTS jobs into CSV/JSON/Excel.")
    parser.add_argument("--term", action="append", dest="terms", help="Search term. Repeat for multiple terms.")
    parser.add_argument("--terms-file", type=Path)
    parser.add_argument("--function", default=DEFAULT_FUNCTION, help="Job function filter (Technical/Professional/Clerical).")
    parser.add_argument("--posted-within-days", type=int, default=4, help="Filter by posted days (0=disabled).")
    parser.add_argument("--keep-w2-f2f-onsite-interview", action="store_true")
    parser.add_argument("--ai-filter", action="store_true", default=False, help="Use Groq/Llama to filter irrelevant jobs (default: off).")
    parser.add_argument("--timeout", type=int, default=20)
    parser.add_argument("--sleep", type=float, default=0.5)
    parser.add_argument("--out-dir", type=Path, default=Path(__file__).resolve().parent / "output")
    parser.add_argument("--no-excel", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    jobs = scrape_cbts(
        load_terms(args),
        args.function,
        args.posted_within_days or None,
        not args.keep_w2_f2f_onsite_interview,
        args.ai_filter,
        args.timeout,
        args.sleep,
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = args.out_dir / f"cbts_jobs_{timestamp}.csv"
    json_path = args.out_dir / f"cbts_jobs_{timestamp}.json"
    excel_path = args.out_dir / f"cbts_jobs_{timestamp}.xlsx"
    write_csv(jobs, csv_path)
    write_json(jobs, json_path)
    if not args.no_excel:
        write_excel(jobs, excel_path, args.posted_within_days or None)
    daily_dir = write_daily_outputs(jobs, args.out_dir, timestamp)
    print(f"Saved {len(jobs)} jobs")
    print(f"CSV:   {csv_path}")
    print(f"JSON:  {json_path}")
    if not args.no_excel:
        print(f"Excel: {excel_path}")
    print(f"Daily: {daily_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
