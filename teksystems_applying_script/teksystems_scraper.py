#!/usr/bin/env python3
"""Standalone TEKsystems job scraper using the Phenom careers page data."""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional
from urllib.parse import quote_plus

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = "https://careers.teksystems.com"
SEARCH_URL = f"{BASE_URL}/us/en/search-results?keywords={{query}}&from=0&s=1"

DEFAULT_SEARCH_TERMS = [
    "python developer",
    "python engineer",
    "senior python developer",
    "backend python engineer",
    "backend software engineer",
    "software engineer python",
    "django developer",
    "fastapi developer",
    "api developer",
    "full stack developer",
    "full stack engineer",
    "software engineer",
    "software developer",
    "data engineer",
    "etl developer",
    "data pipeline",
    "ai engineer",
    "machine learning engineer",
    "ml engineer",
    "generative ai engineer",
    "llm engineer",
    "rag engineer",
    "data scientist",
    "cloud engineer",
    "aws python developer",
    "azure python developer",
]

TITLE_RANKING_WEIGHTS = [
    ("principal ai architect", 60),
    ("ai lead engineer", 58),
    ("machine learning engineer", 56),
    ("ml engineer", 54),
    ("ai engineer", 54),
    ("ai/ml", 54),
    ("aiml", 54),
    ("generative ai", 52),
    ("gen ai", 52),
    ("agentic ai", 50),
    ("llm", 50),
    ("rag", 48),
    ("senior data scientist", 52),
    ("data scientist", 48),
    ("applied data scientist", 50),
    ("senior python", 50),
    ("python developer", 46),
    ("python engineer", 46),
    ("software engineer python", 46),
    ("django", 44),
    ("fastapi", 44),
    ("full stack software engineer", 42),
    ("full stack engineer", 40),
    ("full stack developer", 40),
    ("data engineer", 40),
    ("data engineering", 40),
    ("etl", 36),
    ("data pipeline", 36),
    ("backend", 38),
    ("back end", 38),
    ("api developer", 36),
    ("api engineer", 36),
    ("cloud engineer", 30),
    ("aws", 28),
    ("azure", 28),
    ("senior software engineer", 32),
    ("software engineer", 24),
    ("software developer", 24),
    ("developer", 12),
    ("engineer", 10),
]

TITLE_EXCLUSION_WEIGHTS = [
    ("frontend", -30),
    ("front end", -30),
    ("ui developer", -20),
    ("ux/ui", -12),
    ("java", -12),
    (".net", -35),
    ("qa", -30),
    ("quality assurance", -30),
    ("desktop support", -40),
    ("help desk", -40),
    ("project manager", -35),
    ("business analyst", -30),
    ("network engineer", -50),
    ("civil engineer", -100),
    ("mechanical engineer", -100),
    ("electrical engineer", -100),
    ("systems administrator", -50),
    ("sysadmin", -50),
    ("scrum master", -60),
    ("product manager", -60),
    ("product owner", -60),
    ("it support", -60),
    ("security engineer", -30),
    ("cybersecurity", -30),
    ("sap", -50),
    ("salesforce developer", -30),
    ("ios developer", -50),
    ("android developer", -50),
    ("mobile developer", -50),
    ("embedded", -40),
    ("junior", -100),
    ("jr ", -100),
    ("entry level", -100),
]

TITLE_EXCLUSION_PATTERNS = [
    (re.compile(r"\bjunior\b|\bjr\.?\s", re.I), "Junior title"),
    (re.compile(r"\bentry[\s-]level\b", re.I), "Entry-level title"),
    (re.compile(r"\bintern(ship)?\b", re.I), "Intern title"),
]

DISALLOWED_WORK_PATTERNS = [
    (re.compile(r"\bno\s+c2c\b", re.I), "No C2C"),
    (re.compile(r"\bno\s+corp(?:oration)?\s*[- ]?\s*to\s*[- ]?\s*corp(?:oration)?\b", re.I), "No corp-to-corp"),
    (re.compile(r"\bnot\s+(?:open\s+to\s+)?c2c\b", re.I), "Not open to C2C"),
    (re.compile(r"\bcannot\s+(?:do|support|accept)\s+c2c\b", re.I), "Cannot support C2C"),
    (re.compile(r"\bunable\s+to\s+(?:do|support|accept)\s+c2c\b", re.I), "Unable to support C2C"),
    (re.compile(r"\bw\s*[- ]?\s*2\s+only\b", re.I), "W2 only"),
    (re.compile(r"\bw\s*[- ]?\s*2\s+role\b", re.I), "W2 role"),
    (re.compile(r"\bw\s*[- ]?\s*2\s+contract\b", re.I), "W2 contract"),
    (re.compile(r"\bf\s*2\s*f\b", re.I), "F2F"),
    (re.compile(r"\bface\s*[- ]?\s*to\s*[- ]?\s*face\b", re.I), "Face-to-face"),
    (re.compile(r"\bin\s*[- ]?\s*person\s+interview\b", re.I), "In-person interview"),
    (re.compile(r"\bon\s*[- ]?\s*site\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\bonsite\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\bmust\s+interview\s+on\s*[- ]?\s*site\b", re.I), "Must interview onsite"),
    (re.compile(r"\blocal\s+candidates?\s+only\b", re.I), "Local candidates only"),
    (re.compile(r"\bfull\s*[- ]?\s*time\b", re.I), "Full-time"),
    (re.compile(r"\bpermanent\b", re.I), "Permanent"),
    (re.compile(r"\bdirect\s*[- ]?\s*hire\b", re.I), "Direct hire"),
]

EMAIL_RE = re.compile(r"(?<![A-Za-z0-9._%+-])([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})(?![A-Za-z0-9._%+-])")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?1[\s.-]?)?(?:\(?\d{3}\)?[\s.-]?)\d{3}[\s.-]?\d{4}(?!\d)")
MIN_TITLE_RANK = 20


@dataclass
class TEKsystemsJob:
    source_company: str
    search_term: str
    title_rank: int
    title_rank_reasons: str
    title: str
    category: str
    city: str
    state: str
    location: str
    employment_type: str
    remote_status: str
    posted_date: str
    job_id: str
    job_seq_no: str
    job_url: str
    apply_url: str
    contact_info: str
    skills: str
    description_snippet: str
    raw_text: str


def clean_text(text: Any) -> str:
    return re.sub(r"\s+", " ", html.unescape(str(text or ""))).strip()


def normalized_title(text: str) -> str:
    return re.sub(r"[^a-z0-9+#./-]+", " ", (text or "").lower()).strip()


def title_contains(title: str, phrase: str) -> bool:
    phrase = normalized_title(phrase)
    if not phrase:
        return False
    if len(phrase) <= 3:
        return re.search(rf"(?<![a-z0-9]){re.escape(phrase)}(?![a-z0-9])", title) is not None
    return phrase in title


def score_title(title: str) -> tuple[int, str]:
    normalized = normalized_title(title)
    score = 0
    reasons: list[str] = []
    for phrase, weight in TITLE_RANKING_WEIGHTS:
        if title_contains(normalized, phrase):
            score += weight
            reasons.append(f"{phrase}+{weight}")
    for phrase, weight in TITLE_EXCLUSION_WEIGHTS:
        if title_contains(normalized, phrase):
            score += weight
            reasons.append(f"{phrase}{weight}")
    return max(score, 0), "; ".join(reasons)


def title_exclusion_reasons(title: str) -> list[str]:
    return [reason for pattern, reason in TITLE_EXCLUSION_PATTERNS if pattern.search(title or "")]


def disallowed_work_reasons(text: str) -> list[str]:
    cleaned = clean_text(text)
    reasons: list[str] = []
    for pattern, reason in DISALLOWED_WORK_PATTERNS:
        if pattern.search(cleaned) and reason not in reasons:
            reasons.append(reason)
    return reasons


def extract_contact_info(text: str) -> str:
    emails: list[str] = []
    for email in EMAIL_RE.findall(clean_text(text)):
        lowered = email.lower()
        if lowered not in emails:
            emails.append(lowered)
    phones: list[str] = []
    for phone in PHONE_RE.findall(text or ""):
        digits = re.sub(r"\D", "", phone)
        if len(digits) == 11 and digits.startswith("1"):
            digits = digits[1:]
        if len(digits) == 10:
            display = f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
            if display not in phones:
                phones.append(display)
    return ", ".join(emails + phones)


def parse_teksystems_date(value: str) -> Optional[datetime]:
    value = clean_text(value)
    if not value:
        return None
    if value.endswith("Z"):
        value = value[:-1] + "+00:00"
    if re.search(r"[+-]\d{4}$", value):
        value = value[:-5] + value[-5:-2] + ":" + value[-2:]
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def job_posted_day(job: TEKsystemsJob) -> str:
    parsed = parse_teksystems_date(job.posted_date)
    return parsed.date().isoformat() if parsed else "unknown-date"


def is_within_posted_days(posted_date: str, days: Optional[int]) -> bool:
    if not days or days <= 0:
        return True
    parsed = parse_teksystems_date(posted_date)
    if not parsed:
        return True
    now = datetime.now(timezone.utc)
    return 0 <= (now - parsed).total_seconds() <= days * 86400


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
    })
    return session


def extract_json_object(text: str, start_index: int) -> Optional[dict[str, Any]]:
    depth = 0
    in_string = False
    escape_next = False
    for index in range(start_index, len(text)):
        char = text[index]
        if escape_next:
            escape_next = False
        elif char == "\\" and in_string:
            escape_next = True
        elif char == '"':
            in_string = not in_string
        elif not in_string:
            if char == "{":
                depth += 1
            elif char == "}":
                depth -= 1
                if depth == 0:
                    try:
                        return json.loads(text[start_index:index + 1])
                    except json.JSONDecodeError:
                        return None
    return None


def extract_eager_search(html_text: str) -> dict[str, Any]:
    marker = '"eagerLoadRefineSearch":'
    marker_index = html_text.find(marker)
    if marker_index == -1:
        return {}
    start = html_text.find("{", marker_index + len(marker))
    if start == -1:
        return {}
    return extract_json_object(html_text, start) or {}


def search_url(term: str) -> str:
    return SEARCH_URL.format(query=quote_plus(term))


def job_detail_url(row: dict[str, Any]) -> str:
    seq = clean_text(row.get("jobSeqNo"))
    title = re.sub(r"[^a-z0-9]+", "-", clean_text(row.get("title")).lower()).strip("-")
    if seq:
        return f"{BASE_URL}/us/en/job/{seq}/{title or 'job'}"
    job_id = clean_text(row.get("jobId"))
    return f"{BASE_URL}/us/en/search-results?keywords={quote_plus(job_id)}" if job_id else ""


def row_description(row: dict[str, Any]) -> str:
    parser = row.get("ml_job_parser") or {}
    parts = [
        row.get("descriptionTeaser"),
        parser.get("descriptionTeaser_keyword"),
        parser.get("descriptionTeaser_ats"),
        parser.get("descriptionTeaser_first200"),
        " ".join(str(skill) for skill in row.get("ml_skills") or []),
    ]
    return clean_text(" ".join(str(part or "") for part in parts))


def normalize_job(row: dict[str, Any], search_term: str) -> TEKsystemsJob:
    title = clean_text(row.get("title") or "Untitled job")
    title_rank, title_rank_reasons = score_title(title)
    city = clean_text(row.get("city"))
    state = clean_text(row.get("state"))
    location = clean_text(row.get("location")) or ", ".join(part for part in (city, state) if part)
    remote = clean_text(row.get("remoteOnsite") or row.get("badge"))
    raw_text = row_description(row)
    skills = ", ".join(clean_text(skill) for skill in row.get("ml_skills") or [] if clean_text(skill))
    return TEKsystemsJob(
        source_company="TEKsystems",
        search_term=search_term,
        title_rank=title_rank,
        title_rank_reasons=title_rank_reasons,
        title=title,
        category=clean_text(row.get("category") or row.get("subCategory")),
        city=city,
        state=state,
        location=location,
        employment_type=clean_text(row.get("type")),
        remote_status=remote,
        posted_date=clean_text(row.get("postedDate") or row.get("dateCreated")),
        job_id=clean_text(row.get("jobId") or row.get("reqId")),
        job_seq_no=clean_text(row.get("jobSeqNo")),
        job_url=job_detail_url(row),
        apply_url=clean_text(row.get("applyUrl")),
        contact_info=extract_contact_info(raw_text),
        skills=skills,
        description_snippet=raw_text[:900],
        raw_text=raw_text,
    )


def sort_jobs(jobs: Iterable[TEKsystemsJob]) -> list[TEKsystemsJob]:
    return sorted(
        jobs,
        key=lambda job: (job_posted_day(job), bool(job.contact_info), job.title_rank, job.title.lower()),
        reverse=True,
    )


def scrape_teksystems(
    search_terms: Iterable[str],
    posted_within_days: Optional[int],
    exclude_disallowed_work: bool,
    timeout: int,
    sleep_seconds: float,
) -> list[TEKsystemsJob]:
    session = make_session()
    seen: set[str] = set()
    jobs: list[TEKsystemsJob] = []

    for term in search_terms:
        term = term.strip()
        if not term:
            continue
        url = search_url(term)
        print(f"Searching TEKsystems: {term} -> {url}")
        try:
            response = session.get(url, timeout=timeout)
            response.raise_for_status()
        except requests.RequestException as exc:
            print(f"  Request failed: {exc}")
            time.sleep(sleep_seconds)
            continue

        payload = extract_eager_search(response.text)
        data = payload.get("data") or {}
        rows = [row for row in data.get("jobs") or [] if isinstance(row, dict)]
        print(f"  Found {data.get('totalHits', len(rows))} total, {len(rows)} on page")

        for row in rows:
            key = clean_text(row.get("jobSeqNo") or row.get("jobId") or row.get("reqId"))
            if not key or key in seen:
                continue
            seen.add(key)

            job = normalize_job(row, term)
            if not is_within_posted_days(job.posted_date, posted_within_days):
                continue

            if job.title_rank < MIN_TITLE_RANK:
                print(f"  Skipped low-rank ({job.title_rank}): {job.title}")
                continue

            title_reasons = title_exclusion_reasons(job.title)
            if title_reasons:
                print(f"  Excluded ({', '.join(title_reasons)}): {job.title}")
                continue

            if exclude_disallowed_work:
                reasons = disallowed_work_reasons(" ".join([job.title, job.employment_type, job.raw_text]))
                if reasons:
                    print(f"  Excluded ({', '.join(reasons)}): {job.title}")
                    continue

            jobs.append(job)
        time.sleep(sleep_seconds)

    return sort_jobs(jobs)


def write_csv(jobs: list[TEKsystemsJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(TEKsystemsJob.__dataclass_fields__.keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for job in jobs:
            writer.writerow(asdict(job))


def write_json(jobs: list[TEKsystemsJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump([asdict(job) for job in jobs], handle, indent=2)


EXCEL_COLUMNS = [
    ("posted_day", "Posted Day"),
    ("title_rank", "Title Rank"),
    ("title", "Title"),
    ("location", "Location"),
    ("employment_type", "Employment"),
    ("remote_status", "Remote"),
    ("category", "Category"),
    ("contact_info", "Contact Info"),
    ("search_term", "Search Term"),
    ("job_id", "Job ID"),
    ("job_url", "Job URL"),
    ("apply_url", "Apply URL"),
    ("title_rank_reasons", "Rank Reasons"),
    ("skills", "Skills"),
]


def excel_row(job: TEKsystemsJob) -> dict[str, Any]:
    row = asdict(job)
    row["posted_day"] = job_posted_day(job)
    return row


def style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F3864")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [13, 11, 38, 28, 16, 12, 18, 30, 24, 16, 14, 14, 45, 55]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        rank = row[1].value or 0
        if isinstance(rank, int) and rank >= 45:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="E8F5E9")
        if row[7].value:
            row[7].fill = PatternFill("solid", fgColor="FFF3CD")


def append_jobs_sheet(workbook: Workbook, sheet_name: str, jobs: list[TEKsystemsJob]) -> None:
    safe_name = re.sub(r"[\[\]:*?/\\]", "-", sheet_name)[:31] or "Sheet"
    sheet = workbook.create_sheet(safe_name)
    sheet.append([label for _, label in EXCEL_COLUMNS])
    for job in jobs:
        source = excel_row(job)
        sheet.append([source.get(key, "") for key, _ in EXCEL_COLUMNS])
        for col_index, (key, _) in enumerate(EXCEL_COLUMNS, start=1):
            cell = sheet.cell(row=sheet.max_row, column=col_index)
            if key == "job_url" and job.job_url:
                cell.value = "Open Job"
                cell.hyperlink = job.job_url
                cell.style = "Hyperlink"
            elif key == "apply_url" and job.apply_url:
                cell.value = "Apply"
                cell.hyperlink = job.apply_url
                cell.style = "Hyperlink"
    style_sheet(sheet)


def write_excel(jobs: list[TEKsystemsJob], path: Path, posted_within_days: Optional[int]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "TEKsystems Daily Jobs"
    summary["A1"].font = Font(size=18, bold=True, color="1F3864")
    summary["A3"] = "Generated"
    summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary["A4"] = "Posting Window"
    summary["B4"] = "All dates" if not posted_within_days else f"Last {posted_within_days} days"
    summary["A5"] = "Total Jobs"
    summary["B5"] = len(jobs)
    summary.append([])
    summary.append(["Day", "Jobs", "With Contact", "Top Rank"])
    for cell in summary[7]:
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.font = Font(color="FFFFFF", bold=True)
    grouped: dict[str, list[TEKsystemsJob]] = {}
    for job in jobs:
        grouped.setdefault(job_posted_day(job), []).append(job)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        summary.append([day, len(day_jobs), sum(1 for job in day_jobs if job.contact_info), max((job.title_rank for job in day_jobs), default=0)])
    for col in range(1, 5):
        summary.column_dimensions[get_column_letter(col)].width = 22
    append_jobs_sheet(workbook, "All Jobs", jobs)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        append_jobs_sheet(workbook, day, sorted(day_jobs, key=lambda job: (not bool(job.contact_info), -job.title_rank, job.title.lower())))
    workbook.save(path)


def write_daily_outputs(jobs: list[TEKsystemsJob], out_dir: Path, timestamp: str) -> Path:
    daily_dir = out_dir / f"daily_{timestamp}"
    grouped: dict[str, list[TEKsystemsJob]] = {}
    for job in jobs:
        grouped.setdefault(job_posted_day(job), []).append(job)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        write_csv(day_jobs, daily_dir / f"{day}_jobs.csv")
        write_json(day_jobs, daily_dir / f"{day}_jobs.json")
    return daily_dir


def load_terms(args: argparse.Namespace) -> list[str]:
    terms = list(args.terms or [])
    if args.terms_file:
        terms.extend(line.strip() for line in args.terms_file.read_text(encoding="utf-8").splitlines())
    return [term for term in terms if term.strip()] or DEFAULT_SEARCH_TERMS


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape TEKsystems jobs into CSV/JSON/Excel.")
    parser.add_argument("--term", action="append", dest="terms", help="Search term. Repeat for multiple terms.")
    parser.add_argument("--terms-file", type=Path)
    parser.add_argument("--posted-within-days", type=int, default=4)
    parser.add_argument("--keep-w2-f2f-onsite-interview", action="store_true")
    parser.add_argument("--timeout", type=int, default=25)
    parser.add_argument("--sleep", type=float, default=0.5)
    parser.add_argument("--out-dir", type=Path, default=Path(__file__).resolve().parent / "output")
    parser.add_argument("--no-excel", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    jobs = scrape_teksystems(
        load_terms(args),
        args.posted_within_days,
        not args.keep_w2_f2f_onsite_interview,
        args.timeout,
        args.sleep,
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = args.out_dir / f"teksystems_jobs_{timestamp}.csv"
    json_path = args.out_dir / f"teksystems_jobs_{timestamp}.json"
    excel_path = args.out_dir / f"teksystems_jobs_{timestamp}.xlsx"
    write_csv(jobs, csv_path)
    write_json(jobs, json_path)
    if not args.no_excel:
        write_excel(jobs, excel_path, args.posted_within_days)
    daily_dir = write_daily_outputs(jobs, args.out_dir, timestamp)
    print(f"\nSaved {len(jobs)} jobs")
    print(f"CSV:   {csv_path}")
    print(f"JSON:  {json_path}")
    if not args.no_excel:
        print(f"Excel: {excel_path}")
    print(f"Daily: {daily_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
