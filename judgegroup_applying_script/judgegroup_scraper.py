#!/usr/bin/env python3
"""Standalone Judge Group job scraper using the public Judge AJAX endpoint."""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = "https://www.judge.com"
AJAX_URL = f"{BASE_URL}/wp-admin/admin-ajax.php?action=jdg_get_jobs"

DEFAULT_SEARCH_TERMS = [
    "python developer",
    "python engineer",
    "senior python developer",
    "backend python engineer",
    "backend software engineer",
    "software engineer python",
    "django developer",
    "fastapi developer",
    "api developer",
    "full stack developer",
    "full stack engineer",
    "software engineer",
    "software developer",
    "data engineer",
    "data engineering",
    "etl developer",
    "data pipeline",
    "ai engineer",
    "machine learning engineer",
    "ml engineer",
    "generative ai engineer",
    "llm engineer",
    "rag engineer",
    "data scientist",
    "cloud engineer",
    "aws python developer",
    "azure python developer",
]

TITLE_RANKING_WEIGHTS = [
    ("principal ai architect", 60),
    ("ai lead engineer", 58),
    ("machine learning engineer", 56),
    ("ml engineer", 54),
    ("ai engineer", 54),
    ("ai/ml", 54),
    ("generative ai", 52),
    ("gen ai", 52),
    ("agentic ai", 50),
    ("llm", 50),
    ("rag", 48),
    ("senior data scientist", 52),
    ("data scientist", 48),
    ("senior python", 50),
    ("python developer", 46),
    ("python engineer", 46),
    ("software engineer python", 46),
    ("django", 44),
    ("fastapi", 44),
    ("full stack software engineer", 42),
    ("full stack engineer", 40),
    ("full stack developer", 40),
    ("data engineer", 42),
    ("data engineering", 42),
    ("etl", 36),
    ("data pipeline", 36),
    ("backend", 38),
    ("back-end", 38),
    ("back end", 38),
    ("api developer", 36),
    ("api engineer", 36),
    ("cloud engineer", 30),
    ("aws", 28),
    ("azure", 28),
    ("senior software engineer", 34),
    ("specialty software engineer", 32),
    ("software engineer", 24),
    ("software developer", 24),
    ("developer", 12),
    ("engineer", 10),
]

CONTENT_BONUS_WEIGHTS = [
    ("python", 12),
    ("django", 10),
    ("fastapi", 10),
    ("machine learning", 12),
    ("generative ai", 12),
    ("agentic ai", 12),
    ("llm", 10),
    ("rag", 10),
    ("langchain", 10),
    ("langgraph", 10),
    ("data engineer", 12),
    ("data engineering", 12),
    ("etl", 8),
    ("data pipeline", 8),
    ("react", 6),
    ("node.js", 6),
    ("aws", 6),
    ("azure", 6),
    ("gcp", 6),
]

TITLE_EXCLUSION_WEIGHTS = [
    ("frontend", -30),
    ("front end", -30),
    ("ui developer", -20),
    (".net", -35),
    ("qa", -30),
    ("quality assurance", -30),
    ("desktop support", -40),
    ("help desk", -40),
    ("project manager", -35),
    ("business analyst", -30),
    ("network engineer", -45),
    ("civil engineer", -100),
    ("mechanical engineer", -100),
    ("electrical engineer", -100),
    ("systems administrator", -50),
    ("scrum master", -60),
    ("product manager", -60),
    ("product owner", -60),
    ("it support", -60),
    ("sap", -50),
    ("salesforce developer", -30),
    ("ios developer", -50),
    ("android developer", -50),
    ("mobile developer", -50),
    ("junior", -100),
    ("jr ", -100),
    ("entry level", -100),
]

TITLE_EXCLUSION_PATTERNS = [
    (re.compile(r"\bjunior\b|\bjr\.?\s", re.I), "Junior title"),
    (re.compile(r"\bentry[\s-]level\b", re.I), "Entry-level title"),
    (re.compile(r"\bintern(ship)?\b", re.I), "Intern title"),
    (re.compile(r"\bembedded\b", re.I), "Embedded title"),
]

ROLE_EXCLUSION_PATTERNS = [
    (re.compile(r"\bembedded\s+software\b", re.I), "Embedded software"),
    (re.compile(r"\bmissile\s+systems?\b", re.I), "Missile systems"),
    (re.compile(r"\bros\s*/?\s*ros2\b|\bros2?\b", re.I), "ROS/hardware role"),
    (re.compile(r"\belectronics\s+hardware\b", re.I), "Electronics hardware"),
    (re.compile(r"\bwiring\s+a\s+motor\b", re.I), "Hardware wiring"),
]

DISALLOWED_WORK_PATTERNS = [
    (re.compile(r"\bno\s+c2c\b", re.I), "No C2C"),
    (re.compile(r"\bno\s+corp(?:oration)?\s*[- ]?\s*to\s*[- ]?\s*corp(?:oration)?\b", re.I), "No corp-to-corp"),
    (re.compile(r"\bnot\s+(?:open\s+to\s+)?c2c\b", re.I), "Not open to C2C"),
    (re.compile(r"\bcannot\s+(?:do|support|accept)\s+c2c\b", re.I), "Cannot support C2C"),
    (re.compile(r"\bunable\s+to\s+(?:do|support|accept)\s+c2c\b", re.I), "Unable to support C2C"),
    (re.compile(r"\bw\s*[- ]?\s*2\s+only\b", re.I), "W2 only"),
    (re.compile(r"\bw\s*[- ]?\s*2\s+contract\b", re.I), "W2 contract"),
    (re.compile(r"\bw\s*[- ]?\s*2\b", re.I), "W2"),
    (re.compile(r"\bf\s*2\s*f\b", re.I), "F2F"),
    (re.compile(r"\bface\s*[- ]?\s*to\s*[- ]?\s*face\b", re.I), "Face-to-face"),
    (re.compile(r"\bin\s*[- ]?\s*person\s+interview\b", re.I), "In-person interview"),
    (re.compile(r"\bon\s*[- ]?\s*site\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\blocal\s+candidates?\s+only\b", re.I), "Local candidates only"),
    (re.compile(r"\bpermanent\b", re.I), "Permanent"),
    (re.compile(r"\bdirect\s*[- ]?\s*hire\b", re.I), "Direct hire"),
]

EMAIL_RE = re.compile(r"(?<![A-Za-z0-9._%+-])([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})(?![A-Za-z0-9._%+-])")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?1[\s.-]?)?(?:\(?\d{3}\)?[\s.-]?)\d{3}[\s.-]?\d{4}(?!\d)")
MIN_TITLE_RANK = 20


@dataclass
class JudgeGroupJob:
    source_company: str
    search_term: str
    title_rank: int
    title_rank_reasons: str
    title: str
    category: str
    location: str
    employment_type: str
    salary: str
    posted_date: str
    job_id: str
    job_url: str
    apply_url: str
    contact_info: str
    description_snippet: str
    raw_text: str


def clean_text(text: Any) -> str:
    text = re.sub(r"<[^>]+>", " ", str(text or ""))
    return re.sub(r"\s+", " ", html.unescape(text)).strip()


def normalized_text(text: str) -> str:
    return re.sub(r"[^a-z0-9+#./-]+", " ", (text or "").lower()).strip()


def text_contains(haystack: str, phrase: str) -> bool:
    phrase = normalized_text(phrase)
    if not phrase:
        return False
    if len(phrase) <= 3:
        return re.search(rf"(?<![a-z0-9]){re.escape(phrase)}(?![a-z0-9])", haystack) is not None
    return phrase in haystack


def score_title(title: str, raw_text: str) -> tuple[int, str]:
    normalized_title = normalized_text(title)
    normalized_raw = normalized_text(raw_text)
    score = 0
    reasons: list[str] = []
    for phrase, weight in TITLE_RANKING_WEIGHTS:
        if text_contains(normalized_title, phrase):
            score += weight
            reasons.append(f"{phrase}+{weight}")
    for phrase, weight in CONTENT_BONUS_WEIGHTS:
        if text_contains(normalized_raw, phrase):
            score += weight
            reasons.append(f"{phrase}+{weight} desc")
    for phrase, weight in TITLE_EXCLUSION_WEIGHTS:
        if text_contains(normalized_title, phrase):
            score += weight
            reasons.append(f"{phrase}{weight}")
    return max(score, 0), "; ".join(reasons)


def title_exclusion_reasons(title: str) -> list[str]:
    return [reason for pattern, reason in TITLE_EXCLUSION_PATTERNS if pattern.search(title or "")]


def disallowed_work_reasons(text: str) -> list[str]:
    cleaned = clean_text(text)
    reasons: list[str] = []
    for pattern, reason in DISALLOWED_WORK_PATTERNS:
        if pattern.search(cleaned) and reason not in reasons:
            reasons.append(reason)
    return reasons


def role_exclusion_reasons(text: str) -> list[str]:
    return [reason for pattern, reason in ROLE_EXCLUSION_PATTERNS if pattern.search(text or "")]


def extract_contact_info(text: str) -> str:
    emails: list[str] = []
    for email in EMAIL_RE.findall(clean_text(text)):
        lowered = email.lower()
        if lowered not in emails:
            emails.append(lowered)
    phones: list[str] = []
    for phone in PHONE_RE.findall(text or ""):
        digits = re.sub(r"\D", "", phone)
        if len(digits) == 11 and digits.startswith("1"):
            digits = digits[1:]
        if len(digits) == 10:
            display = f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
            if display not in phones:
                phones.append(display)
    return ", ".join(emails + phones)


def parse_judge_date(value: Any) -> Optional[datetime]:
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number > 10_000_000_000:
        number /= 1000
    return datetime.fromtimestamp(number, tz=timezone.utc)


def posted_day(job: JudgeGroupJob) -> str:
    parsed = parse_judge_date(job.posted_date)
    return parsed.date().isoformat() if parsed else "unknown-date"


def is_within_posted_days(posted_date: str, days: Optional[int]) -> bool:
    if not days or days <= 0:
        return True
    parsed = parse_judge_date(posted_date)
    if not parsed:
        return True
    return 0 <= (datetime.now(timezone.utc) - parsed).total_seconds() <= days * 86400


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "application/json,text/plain,*/*",
        "Accept-Language": "en-US,en;q=0.9",
        "Content-Type": "application/json",
        "Referer": f"{BASE_URL}/jobs/",
    })
    return session


def payload_for(term: str, page: int, employment_type: str) -> dict[str, Any]:
    return {
        "payload": {
            "categories": ["InformationTechnology"],
            "countries": "USA",
            "geo": [{"distance": "", "latLong": [0], "location": ""}],
            "query": term,
            "states": "",
            "type": [employment_type],
            "page": page,
            "remote": False,
        }
    }


def fetch_jobs(session: requests.Session, term: str, page: int, employment_type: str, timeout: int) -> dict[str, Any]:
    response = session.post(AJAX_URL, json=payload_for(term, page, employment_type), timeout=timeout)
    response.raise_for_status()
    data = response.json()
    if isinstance(data, str):
        data = json.loads(data)
    return data if isinstance(data, dict) else {}


def normalize_job(row: dict[str, Any], search_term: str) -> JudgeGroupJob:
    title = clean_text(row.get("title") or "Untitled job")
    raw_text = clean_text(row.get("description"))
    title_rank, title_rank_reasons = score_title(title, raw_text)
    job_id = str(row.get("jobOrderId") or "").strip()
    job_url = f"{BASE_URL}/jobs/details/{job_id}/" if job_id else f"{BASE_URL}/jobs/"
    return JudgeGroupJob(
        source_company="The Judge Group",
        search_term=search_term,
        title_rank=title_rank,
        title_rank_reasons=title_rank_reasons,
        title=title,
        category=clean_text((row.get("category") or {}).get("description") if isinstance(row.get("category"), dict) else row.get("category")),
        location=clean_text(row.get("location")),
        employment_type=clean_text(row.get("type")),
        salary=clean_text(row.get("salary")),
        posted_date=str(row.get("opened") or ""),
        job_id=job_id,
        job_url=job_url,
        apply_url=job_url,
        contact_info=extract_contact_info(raw_text),
        description_snippet=raw_text[:900],
        raw_text=raw_text,
    )


def sort_jobs(jobs: Iterable[JudgeGroupJob]) -> list[JudgeGroupJob]:
    return sorted(jobs, key=lambda job: (posted_day(job), bool(job.contact_info), job.title_rank, job.title.lower()), reverse=True)


def scrape_judgegroup(
    search_terms: Iterable[str],
    posted_within_days: Optional[int],
    exclude_disallowed_work: bool,
    timeout: int,
    sleep_seconds: float,
    max_pages: int,
) -> list[JudgeGroupJob]:
    session = make_session()
    seen: set[str] = set()
    jobs: list[JudgeGroupJob] = []

    for term in search_terms:
        term = term.strip()
        if not term:
            continue
        page = 0
        while page < max_pages:
            print(f"Searching Judge Group: {term} page {page + 1}")
            try:
                data = fetch_jobs(session, term, page, "Contract", timeout)
            except (requests.RequestException, json.JSONDecodeError) as exc:
                print(f"  Request failed: {exc}")
                break

            rows = [row for row in data.get("hits") or [] if isinstance(row, dict)]
            total = int(data.get("total") or len(rows))
            size = int(data.get("size") or len(rows) or 20)
            print(f"  Found {total} total, {len(rows)} on page")
            if not rows:
                break

            for row in rows:
                key = str(row.get("jobOrderId") or "").strip()
                if not key or key in seen:
                    continue
                seen.add(key)
                job = normalize_job(row, term)

                if not is_within_posted_days(job.posted_date, posted_within_days):
                    continue
                if job.employment_type.lower() != "contract":
                    print(f"  Excluded non-contract ({job.employment_type}): {job.title}")
                    continue
                if job.title_rank < MIN_TITLE_RANK:
                    print(f"  Skipped low-rank ({job.title_rank}): {job.title}")
                    continue

                title_reasons = title_exclusion_reasons(job.title)
                if title_reasons:
                    print(f"  Excluded ({', '.join(title_reasons)}): {job.title}")
                    continue

                role_reasons = role_exclusion_reasons(" ".join([job.title, job.raw_text]))
                if role_reasons:
                    print(f"  Excluded ({', '.join(role_reasons)}): {job.title}")
                    continue

                if exclude_disallowed_work:
                    reasons = disallowed_work_reasons(" ".join([job.title, job.employment_type, job.raw_text]))
                    if reasons:
                        print(f"  Excluded ({', '.join(reasons)}): {job.title}")
                        continue

                jobs.append(job)

            page += 1
            if page * size >= total:
                break
            time.sleep(sleep_seconds)
        time.sleep(sleep_seconds)

    return sort_jobs(jobs)


def write_csv(jobs: list[JudgeGroupJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(JudgeGroupJob.__dataclass_fields__.keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for job in jobs:
            writer.writerow(asdict(job))


def write_json(jobs: list[JudgeGroupJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump([asdict(job) for job in jobs], handle, indent=2)


EXCEL_COLUMNS = [
    ("posted_day", "Posted Day"),
    ("title_rank", "Title Rank"),
    ("title", "Title"),
    ("location", "Location"),
    ("employment_type", "Employment"),
    ("salary", "Salary"),
    ("category", "Category"),
    ("contact_info", "Contact Info"),
    ("search_term", "Search Term"),
    ("job_id", "Job ID"),
    ("job_url", "Job URL"),
    ("title_rank_reasons", "Rank Reasons"),
]


def excel_row(job: JudgeGroupJob) -> dict[str, Any]:
    row = asdict(job)
    row["posted_day"] = posted_day(job)
    return row


def style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F3864")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [13, 11, 38, 28, 16, 24, 24, 30, 24, 16, 14, 55]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        rank = row[1].value or 0
        if isinstance(rank, int) and rank >= 45:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="E8F5E9")


def append_jobs_sheet(workbook: Workbook, sheet_name: str, jobs: list[JudgeGroupJob]) -> None:
    safe_name = re.sub(r"[\[\]:*?/\\]", "-", sheet_name)[:31] or "Sheet"
    sheet = workbook.create_sheet(safe_name)
    sheet.append([label for _, label in EXCEL_COLUMNS])
    for job in jobs:
        source = excel_row(job)
        sheet.append([source.get(key, "") for key, _ in EXCEL_COLUMNS])
        url_cell = sheet.cell(row=sheet.max_row, column=11)
        if job.job_url:
            url_cell.value = "Open Job"
            url_cell.hyperlink = job.job_url
            url_cell.style = "Hyperlink"
    style_sheet(sheet)


def write_excel(jobs: list[JudgeGroupJob], path: Path, posted_within_days: Optional[int]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "Judge Group Daily Jobs"
    summary["A1"].font = Font(size=18, bold=True, color="1F3864")
    summary["A3"] = "Generated"
    summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary["A4"] = "Posting Window"
    summary["B4"] = "All dates" if not posted_within_days else f"Last {posted_within_days} days"
    summary["A5"] = "Total Jobs"
    summary["B5"] = len(jobs)
    summary.append([])
    summary.append(["Day", "Jobs", "With Contact", "Top Rank"])
    for cell in summary[7]:
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.font = Font(color="FFFFFF", bold=True)
    grouped: dict[str, list[JudgeGroupJob]] = {}
    for job in jobs:
        grouped.setdefault(posted_day(job), []).append(job)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        summary.append([day, len(day_jobs), sum(1 for job in day_jobs if job.contact_info), max((job.title_rank for job in day_jobs), default=0)])
    for col in range(1, 5):
        summary.column_dimensions[get_column_letter(col)].width = 22
    append_jobs_sheet(workbook, "All Jobs", jobs)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        append_jobs_sheet(workbook, day, sorted(day_jobs, key=lambda job: (not bool(job.contact_info), -job.title_rank, job.title.lower())))
    workbook.save(path)


def write_daily_outputs(jobs: list[JudgeGroupJob], out_dir: Path, timestamp: str) -> Path:
    daily_dir = out_dir / f"daily_{timestamp}"
    grouped: dict[str, list[JudgeGroupJob]] = {}
    for job in jobs:
        grouped.setdefault(posted_day(job), []).append(job)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        write_csv(day_jobs, daily_dir / f"{day}_jobs.csv")
        write_json(day_jobs, daily_dir / f"{day}_jobs.json")
    return daily_dir


def load_terms(args: argparse.Namespace) -> list[str]:
    terms = list(args.terms or [])
    if args.terms_file:
        terms.extend(line.strip() for line in args.terms_file.read_text(encoding="utf-8").splitlines())
    return [term for term in terms if term.strip()] or DEFAULT_SEARCH_TERMS


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape Judge Group jobs into CSV/JSON/Excel.")
    parser.add_argument("--term", action="append", dest="terms", help="Search term. Repeat for multiple terms.")
    parser.add_argument("--terms-file", type=Path)
    parser.add_argument("--posted-within-days", type=int, default=4)
    parser.add_argument("--keep-w2-f2f-onsite-interview", action="store_true")
    parser.add_argument("--timeout", type=int, default=25)
    parser.add_argument("--sleep", type=float, default=0.5)
    parser.add_argument("--max-pages", type=int, default=3)
    parser.add_argument("--out-dir", type=Path, default=Path(__file__).resolve().parent / "output")
    parser.add_argument("--no-excel", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    jobs = scrape_judgegroup(
        load_terms(args),
        args.posted_within_days,
        not args.keep_w2_f2f_onsite_interview,
        args.timeout,
        args.sleep,
        args.max_pages,
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = args.out_dir / f"judgegroup_jobs_{timestamp}.csv"
    json_path = args.out_dir / f"judgegroup_jobs_{timestamp}.json"
    excel_path = args.out_dir / f"judgegroup_jobs_{timestamp}.xlsx"
    write_csv(jobs, csv_path)
    write_json(jobs, json_path)
    if not args.no_excel:
        write_excel(jobs, excel_path, args.posted_within_days)
    daily_dir = write_daily_outputs(jobs, args.out_dir, timestamp)
    print(f"\nSaved {len(jobs)} jobs")
    print(f"CSV:   {csv_path}")
    print(f"JSON:  {json_path}")
    if not args.no_excel:
        print(f"Excel: {excel_path}")
    print(f"Daily: {daily_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
