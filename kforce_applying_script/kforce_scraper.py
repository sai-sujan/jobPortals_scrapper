#!/usr/bin/env python3
"""Standalone Kforce job scraper using Azure Cognitive Search API."""

from __future__ import annotations

import argparse
import base64
import csv
import json
import os
import re
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional
from urllib.parse import parse_qs, urlparse

import groq
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


AZURE_SEARCH_URL = (
    "https://kforcewebeast.search.windows.net"
    "/indexes/kforcewebjobentity/docs/search"
    "?api-version=2020-06-30"
)
AZURE_SEARCH_KEY = os.environ.get("KFORCE_AZURE_SEARCH_KEY", "")
KFORCE_JOB_BASE = "https://www.kforce.com/Jobs"

DEFAULT_SEARCH_TERMS = [
    "ai engineer",
    "ai ml engineer",
    "machine learning engineer",
    "ml engineer",
    "generative ai engineer",
    "llm engineer",
    "rag engineer",
    "data scientist",
    "senior data scientist",
    "python developer",
    "python engineer",
    "senior python developer",
    "senior python engineer",
    "backend python engineer",
    "backend software engineer",
    "software engineer python",
    "django developer",
    "fastapi developer",
    "api developer",
    "full stack software engineer",
    "full stack developer",
    "senior software engineer",
    "software engineer",
    "cloud engineer",
    "aws python developer",
    "azure python developer",
]

EMAIL_RE = re.compile(r"(?<![A-Za-z0-9._%+-])([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})(?![A-Za-z0-9._%+-])")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?1[\s.-]?)?(?:\(?\d{3}\)?[\s.-]?)\d{3}[\s.-]?\d{4}(?!\d)")

TITLE_RANKING_WEIGHTS = [
    ("principal ai architect", 60),
    ("ai lead engineer", 58),
    ("machine learning engineer", 56),
    ("ml engineer", 54),
    ("ai engineer", 54),
    ("ai/ml", 54),
    ("aiml", 54),
    ("generative ai", 52),
    ("gen ai", 52),
    ("agentic ai", 50),
    ("llm", 50),
    ("rag", 48),
    ("senior data scientist", 52),
    ("data scientist", 48),
    ("applied data scientist", 50),
    ("python developer", 46),
    ("python engineer", 46),
    ("senior python", 50),
    ("django", 44),
    ("fastapi", 44),
    ("backend", 38),
    ("back end", 38),
    ("api developer", 36),
    ("software engineer python", 46),
    ("full stack software engineer", 42),
    ("full stack engineer", 40),
    ("full stack developer", 40),
    ("cloud engineer", 30),
    ("aws", 28),
    ("azure", 28),
    ("senior software engineer", 32),
    ("software engineer", 24),
    ("software developer", 24),
    ("developer", 12),
    ("engineer", 10),
]
MIN_TITLE_RANK = 20

TITLE_EXCLUSION_WEIGHTS = [
    ("frontend", -30),
    ("front end", -30),
    ("ui developer", -20),
    ("ux/ui", -12),
    ("java", -12),
    (".net", -35),
    ("qa", -30),
    ("quality assurance", -30),
    ("desktop support", -40),
    ("help desk", -40),
    ("project manager", -35),
    ("business analyst", -30),
    ("data engineer", -100),
    ("network engineer", -50),
    ("civil engineer", -100),
    ("mechanical engineer", -100),
    ("electrical engineer", -100),
    ("structural engineer", -100),
    ("environmental", -80),
    ("protection", -60),
    ("infrastructure engineer", -50),
    ("systems administrator", -50),
    ("sysadmin", -50),
    ("devops engineer", -20),
    ("site reliability", -20),
    ("scrum master", -60),
    ("product manager", -60),
    ("product owner", -60),
    ("it support", -60),
    ("network administrator", -80),
    ("security engineer", -30),
    ("cybersecurity", -30),
    ("sap", -50),
    ("salesforce developer", -30),
    ("salesforce engineer", -30),
    ("ios developer", -50),
    ("android developer", -50),
    ("mobile developer", -50),
    ("embedded", -40),
    ("junior", -100),
    ("jr ", -100),
    ("entry level", -100),
]
TITLE_EXCLUSION_PATTERNS = [
    (re.compile(r"\bjunior\b|\bjr\.?\s", re.I), "Junior title"),
    (re.compile(r"\bentry[\s-]level\b", re.I), "Entry-level title"),
]
DISALLOWED_WORK_PATTERNS = [
    (re.compile(r"\bno\s+c2c\b", re.I), "No C2C"),
    (re.compile(r"\bno\s+corp(?:oration)?\s*[- ]?\s*to\s*[- ]?\s*corp(?:oration)?\b", re.I), "No corp-to-corp"),
    (re.compile(r"\bf\s*2\s*f\b", re.I), "F2F"),
    (re.compile(r"\bface\s*[- ]?\s*to\s*[- ]?\s*face\b", re.I), "Face-to-face"),
    (re.compile(r"\bin\s*[- ]?\s*person\s+interview\b", re.I), "In-person interview"),
    (re.compile(r"\bon\s*[- ]?\s*site\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\bonsite\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\bmust\s+interview\s+on\s*[- ]?\s*site\b", re.I), "Must interview onsite"),
    (re.compile(r"\blocal\s+candidates?\s+only\b", re.I), "Local candidates only"),
    (re.compile(r"\bfull\s*[- ]?\s*time\b", re.I), "Full-time"),
    (re.compile(r"\bpermanent\b", re.I), "Permanent"),
    (re.compile(r"\bdirect\s*[- ]?\s*hire\b", re.I), "Direct hire"),
    (re.compile(r"\bcontract\s*[- ]?\s*to\s*[- ]?\s*hire\b", re.I), "Contract-to-Hire"),
    (re.compile(r"\bcth\b", re.I), "Contract-to-Hire (CTH)"),
]


@dataclass
class KforceJob:
    source_company: str
    search_term: str
    title_rank: int
    title_rank_reasons: str
    title: str
    city: str
    state: str
    location: str
    employment_type: str
    min_pay_rate: str
    max_pay_rate: str
    pay_rate_unit: str
    posted_date: str
    job_id: str
    reference_code: str
    job_url: str
    apply_url: str
    contact_info: str
    description_snippet: str
    raw_text: str


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").strip()).strip()


def normalized_title(text: str) -> str:
    return re.sub(r"[^a-z0-9+#./-]+", " ", (text or "").lower()).strip()


def title_contains(title: str, phrase: str) -> bool:
    phrase = normalized_title(phrase)
    if not phrase:
        return False
    if len(phrase) <= 3:
        return re.search(rf"(?<![a-z0-9]){re.escape(phrase)}(?![a-z0-9])", title) is not None
    return phrase in title


def score_title(title: str) -> tuple[int, str]:
    normalized = normalized_title(title)
    score = 0
    reasons: list[str] = []
    for phrase, weight in TITLE_RANKING_WEIGHTS:
        if title_contains(normalized, phrase):
            score += weight
            reasons.append(f"{phrase}+{weight}")
    for phrase, weight in TITLE_EXCLUSION_WEIGHTS:
        if title_contains(normalized, phrase):
            score += weight
            reasons.append(f"{phrase}{weight}")
    return max(score, 0), "; ".join(reasons)


def disallowed_work_reasons(text: str) -> list[str]:
    cleaned = clean_text(text)
    reasons: list[str] = []
    for pattern, reason in DISALLOWED_WORK_PATTERNS:
        if pattern.search(cleaned) and reason not in reasons:
            reasons.append(reason)
    return reasons


def decode_kforce_id(job_id: str) -> str:
    if not job_id:
        return ""
    padded = job_id + "=" * ((4 - len(job_id) % 4) % 4)
    try:
        return base64.b64decode(padded.replace("-", "+").replace("_", "/")).decode("utf-8")
    except Exception:
        return ""


def kforce_job_url(job_id: str) -> str:
    path = decode_kforce_id(job_id)
    return f"{KFORCE_JOB_BASE}/{path}/" if path else ""


def recruiter_email_from_apply_url(apply_url: str) -> str:
    if not apply_url:
        return ""
    qs = parse_qs(urlparse(apply_url).query)
    email = (qs.get("Industry") or [""])[0]
    return email.lower() if "@" in email else ""


def contact_info_from_text(text: str) -> str:
    emails: list[str] = []
    for email in EMAIL_RE.findall(clean_text(text)):
        lowered = email.lower()
        if lowered not in emails:
            emails.append(lowered)
    phones: list[str] = []
    for phone in PHONE_RE.findall(text or ""):
        digits = re.sub(r"\D", "", phone)
        if len(digits) == 11 and digits.startswith("1"):
            digits = digits[1:]
        if len(digits) == 10:
            display = f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
            if display not in phones:
                phones.append(display)
    return ", ".join(emails + phones)


def combine_contact_info(*parts: str) -> str:
    seen: list[str] = []
    for part in parts:
        for item in (part or "").split(", "):
            item = item.strip()
            if item and item not in seen:
                seen.append(item)
    return ", ".join(seen)


def parse_kforce_date(value: str) -> Optional[datetime]:
    value = clean_text(value)
    if not value:
        return None
    if value.endswith("Z"):
        value = value[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def job_posted_day(job: KforceJob) -> str:
    parsed = parse_kforce_date(job.posted_date)
    return parsed.date().isoformat() if parsed else "unknown-date"


def is_within_posted_days(post_date: str, days: Optional[int], now: Optional[datetime] = None) -> bool:
    if not days or days <= 0:
        return True
    parsed = parse_kforce_date(post_date)
    if not parsed:
        return False
    now = now or datetime.now(timezone.utc)
    return 0 <= (now - parsed).total_seconds() <= days * 24 * 60 * 60


def is_hourly_unit(unit: str) -> bool:
    return normalized_title(unit) in {"hour", "hourly", "hours", "hr"}


def below_min_hourly(job: KforceJob, min_hourly_rate: Optional[float]) -> bool:
    if not min_hourly_rate or min_hourly_rate <= 0 or not is_hourly_unit(job.pay_rate_unit):
        return False
    rates = [float(r) for r in (job.min_pay_rate, job.max_pay_rate) if r]
    return bool(rates) and max(rates) < min_hourly_rate


def normalize_job(row: dict[str, Any], search_term: str) -> KforceJob:
    title = clean_text(str(row.get("Title") or "Untitled"))
    city = clean_text(str(row.get("City") or ""))
    state = clean_text(str(row.get("State") or ""))
    location = ", ".join(part for part in (city, state) if part)
    responsibilities = clean_text(str(row.get("Responsibilities") or ""))
    skills = clean_text(str(row.get("Skills") or ""))
    raw_text = " ".join(filter(None, [responsibilities, skills]))
    apply_url = clean_text(str(row.get("ApplyUrl") or ""))
    recruiter_email = recruiter_email_from_apply_url(apply_url)
    text_contact = contact_info_from_text(raw_text)
    all_contact = combine_contact_info(recruiter_email, text_contact)
    title_rank, title_rank_reasons = score_title(title)
    job_id = clean_text(str(row.get("Id") or ""))
    salary_min = clean_text(str(row.get("SalaryMin") or ""))
    salary_max = clean_text(str(row.get("SalaryMax") or ""))
    salary_text = clean_text(str(row.get("SalaryText") or ""))
    return KforceJob(
        source_company="Kforce",
        search_term=search_term,
        title_rank=title_rank,
        title_rank_reasons=title_rank_reasons,
        title=title,
        city=city,
        state=state,
        location=location,
        employment_type=clean_text(str(row.get("TypeCode") or "")),
        min_pay_rate=salary_min,
        max_pay_rate=salary_max,
        pay_rate_unit=salary_text,
        posted_date=clean_text(str(row.get("PostDate") or "")),
        job_id=job_id,
        reference_code=clean_text(str(row.get("ReferenceCode") or "")),
        job_url=kforce_job_url(job_id),
        apply_url=apply_url,
        contact_info=all_contact,
        description_snippet=raw_text[:900],
        raw_text=raw_text,
    )


def make_session() -> requests.Session:
    if not AZURE_SEARCH_KEY:
        raise RuntimeError("KFORCE_AZURE_SEARCH_KEY must be set to scrape Kforce.")
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 VenkataDoraKforceAutomation/0.1 (+local personal job search)",
        "Accept": "application/json",
        "api-key": AZURE_SEARCH_KEY,
        "Content-Type": "application/json",
    })
    return session


def build_filter(contract_only: bool, posted_within_days: Optional[int]) -> str:
    parts: list[str] = []
    if contract_only:
        parts.append("TypeCode eq 'Contract'")
    if posted_within_days and posted_within_days > 0:
        cutoff = datetime.now(timezone.utc)
        from datetime import timedelta
        cutoff = cutoff - timedelta(days=posted_within_days)
        parts.append(f"PostDate ge {cutoff.strftime('%Y-%m-%dT%H:%M:%SZ')}")
    return " and ".join(parts)


def search_jobs(
    session: requests.Session,
    term: str,
    odata_filter: str,
    timeout: int,
    top: int = 100,
) -> list[dict[str, Any]]:
    body = {
        "search": term,
        "searchFields": "Industry, Title, Responsibilities, Skills, City, State, Zip",
        "select": "Industry, Title, Id, PostDate, Responsibilities, Skills, City, State, Zip, SalaryMin, SalaryMax, SalaryText, ReferenceCode, TypeCode, VisaSponsorshipJob, ApplyUrl",
        "filter": odata_filter,
        "count": True,
        "queryType": "simple",
        "searchMode": "any",
        "top": top,
        "skip": 0,
    }
    response = session.post(AZURE_SEARCH_URL, json=body, timeout=timeout)
    response.raise_for_status()
    payload = response.json()
    return [r for r in payload.get("value", []) if isinstance(r, dict)]


def sort_jobs(jobs: Iterable[KforceJob]) -> list[KforceJob]:
    return sorted(
        jobs,
        key=lambda job: (job_posted_day(job), bool(job.contact_info), job.title_rank, job.title.lower()),
        reverse=True,
    )


GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")
GROQ_MODEL = "llama-3.1-8b-instant"

_AI_FILTER_SYSTEM = (
    "You filter job listings for a candidate. Profile: Senior Python Developer, "
    "AI/ML Engineer, Data Scientist, Machine Learning Engineer, Software Engineer. "
    "They use Python, AI/ML frameworks, data science tools. "
    "They will NOT do: pure data engineering/ETL pipelines, frontend/UI, Java/.NET, "
    "QA/testing, project management, network/civil/mechanical/electrical engineering, "
    "DevOps/SRE, cybersecurity, mobile, embedded, SAP, Salesforce admin, IT support."
)

_AI_FILTER_PROMPT = """Below is a JSON list of jobs. For each, decide if it is RELEVANT or IRRELEVANT for the candidate.

Reply with ONLY a JSON array in this exact format:
[{{"id": "REF_CODE", "verdict": "RELEVANT", "reason": "one short phrase"}}, ...]

Jobs:
{jobs_json}"""


def ai_filter_jobs(jobs: list[KforceJob], batch_size: int = 20) -> list[KforceJob]:
    if not GROQ_API_KEY:
        print("Skipping AI filter: GROQ_API_KEY is not set.")
        return jobs
    client = groq.Groq(api_key=GROQ_API_KEY)
    kept: list[KforceJob] = []

    for batch_start in range(0, len(jobs), batch_size):
        batch = jobs[batch_start: batch_start + batch_size]
        jobs_payload = [
            {"id": j.reference_code, "title": j.title, "snippet": j.description_snippet[:300]}
            for j in batch
        ]
        prompt = _AI_FILTER_PROMPT.format(jobs_json=json.dumps(jobs_payload, indent=2))
        try:
            response = client.chat.completions.create(
                model=GROQ_MODEL,
                messages=[
                    {"role": "system", "content": _AI_FILTER_SYSTEM},
                    {"role": "user", "content": prompt},
                ],
                max_tokens=1024,
                temperature=0,
            )
            raw = response.choices[0].message.content.strip()
            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)
            verdicts: list[dict[str, str]] = json.loads(raw)
            verdict_map = {v["id"]: v["verdict"].upper() for v in verdicts}
        except Exception as exc:
            print(f"AI filter batch failed: {exc} — keeping all {len(batch)} jobs in this batch")
            kept.extend(batch)
            continue

        for job in batch:
            verdict = verdict_map.get(job.reference_code, "RELEVANT")
            if verdict == "RELEVANT":
                kept.append(job)
            else:
                reason = next((v.get("reason", "") for v in verdicts if v.get("id") == job.reference_code), "")
                print(f"AI excluded {job.reference_code}: {job.title} — {reason}")

    return kept


def scrape_kforce(
    search_terms: Iterable[str],
    posted_within_days: Optional[int],
    contract_only: bool,
    exclude_disallowed_work: bool,
    min_hourly_rate: Optional[float],
    use_ai_filter: bool,
    timeout: int,
    sleep_seconds: float,
) -> list[KforceJob]:
    session = make_session()
    odata_filter = build_filter(contract_only, posted_within_days)
    seen_ids: set[str] = set()
    jobs: list[KforceJob] = []

    for term in search_terms:
        term = term.strip()
        if not term:
            continue
        print(f"Searching Kforce: {term}")
        try:
            rows = search_jobs(session, term, odata_filter, timeout)
        except requests.RequestException as exc:
            print(f"Search failed for {term!r}: {exc}")
            time.sleep(sleep_seconds)
            continue

        for row in rows:
            job_id = clean_text(str(row.get("Id") or ""))
            if not job_id or job_id in seen_ids:
                continue
            seen_ids.add(job_id)
            job = normalize_job(row, term)
            if job.title_rank < MIN_TITLE_RANK:
                print(f"Excluded {job.reference_code}: low title rank ({job.title_rank}) — {job.title}")
                continue
            if exclude_disallowed_work and disallowed_work_reasons(job.raw_text):
                reasons = ", ".join(disallowed_work_reasons(job.raw_text))
                print(f"Excluded {job.reference_code}: {reasons}")
                continue
            if below_min_hourly(job, min_hourly_rate):
                print(f"Excluded {job.reference_code}: below ${min_hourly_rate:g}/hour pay threshold")
                continue
            if job.pay_rate_unit == "year":
                print(f"  Excluded (annual salary = full-time signal): {job.title}")
                continue
            jobs.append(job)
        time.sleep(sleep_seconds)

    print(f"\n--- {len(jobs)} jobs passed rule filter ---")
    for j in jobs:
        print(f"  [{j.title_rank:3d}] {j.title} | {j.location} | ${j.min_pay_rate}-${j.max_pay_rate} {j.pay_rate_unit}")
    print()

    if use_ai_filter and jobs:
        print(f"AI filtering {len(jobs)} jobs with groq/llama...")
        jobs = ai_filter_jobs(jobs)
        print(f"AI filter kept {len(jobs)} jobs.")

    return sort_jobs(jobs)


def write_csv(jobs: list[KforceJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(KforceJob.__dataclass_fields__.keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for job in jobs:
            writer.writerow(asdict(job))


def write_json(jobs: list[KforceJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump([asdict(job) for job in jobs], handle, indent=2)


EXCEL_COLUMNS = [
    ("posted_day", "Posted Day"),
    ("title_rank", "Title Rank"),
    ("title", "Title"),
    ("location", "Location"),
    ("employment_type", "Employment"),
    ("pay_range", "Pay"),
    ("contact_info", "Contact Info"),
    ("search_term", "Search Term"),
    ("reference_code", "Ref Code"),
    ("job_url", "Job URL"),
    ("apply_url", "Apply URL"),
    ("title_rank_reasons", "Rank Reasons"),
]


def excel_row(job: KforceJob) -> dict[str, Any]:
    row = asdict(job)
    row["posted_day"] = job_posted_day(job)
    row["pay_range"] = ""
    if job.min_pay_rate or job.max_pay_rate:
        lo = job.min_pay_rate or job.max_pay_rate
        hi = job.max_pay_rate or job.min_pay_rate
        row["pay_range"] = f"${lo}-${hi} / {job.pay_rate_unit}".strip(" /") if lo != hi else f"${hi} / {job.pay_rate_unit}".strip(" /")
    return row


def style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="1A3A5C")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [13, 11, 34, 26, 16, 18, 30, 24, 14, 14, 14, 45]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        rank = row[1].value or 0
        if isinstance(rank, int) and rank >= 45:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="E8F5E9")
        if row[6].value:
            row[6].fill = PatternFill("solid", fgColor="FFF3CD")


def append_jobs_sheet(workbook: Workbook, sheet_name: str, jobs: list[KforceJob]) -> None:
    safe_name = re.sub(r"[\[\]:*?/\\]", "-", sheet_name)[:31] or "Sheet"
    sheet = workbook.create_sheet(safe_name)
    sheet.append([label for _, label in EXCEL_COLUMNS])
    for job in jobs:
        source = excel_row(job)
        sheet.append([source.get(key, "") for key, _ in EXCEL_COLUMNS])
        for col_idx, (key, _) in enumerate(EXCEL_COLUMNS, start=1):
            cell = sheet.cell(row=sheet.max_row, column=col_idx)
            if key == "job_url" and job.job_url:
                cell.value = "Open Job"
                cell.hyperlink = job.job_url
                cell.style = "Hyperlink"
            elif key == "apply_url" and job.apply_url:
                cell.value = "Apply"
                cell.hyperlink = job.apply_url
                cell.style = "Hyperlink"
    style_sheet(sheet)


def write_excel(jobs: list[KforceJob], path: Path, posted_within_days: Optional[int], min_hourly_rate: Optional[float]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "Kforce Daily Jobs"
    summary["A1"].font = Font(size=18, bold=True, color="1A3A5C")
    summary["A3"] = "Generated"
    summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary["A4"] = "Posting Window"
    summary["B4"] = "All dates" if not posted_within_days else f"Last {posted_within_days} days"
    summary["A5"] = "Minimum Hourly Pay"
    summary["B5"] = "Disabled" if not min_hourly_rate else f"${min_hourly_rate:g}/hour"
    summary["A6"] = "Total Jobs"
    summary["B6"] = len(jobs)
    summary.append([])
    summary.append(["Day", "Jobs", "With Contact", "Top Rank"])
    for cell in summary[8]:
        cell.fill = PatternFill("solid", fgColor="1A3A5C")
        cell.font = Font(color="FFFFFF", bold=True)
    grouped: dict[str, list[KforceJob]] = {}
    for job in jobs:
        grouped.setdefault(job_posted_day(job), []).append(job)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        summary.append([day, len(day_jobs), sum(1 for j in day_jobs if j.contact_info), max((j.title_rank for j in day_jobs), default=0)])
    for col in range(1, 5):
        summary.column_dimensions[get_column_letter(col)].width = 22
    append_jobs_sheet(workbook, "All Jobs", jobs)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        append_jobs_sheet(workbook, day, sorted(day_jobs, key=lambda j: (not bool(j.contact_info), -j.title_rank, j.title.lower())))
    workbook.save(path)


def write_daily_outputs(jobs: list[KforceJob], out_dir: Path, timestamp: str) -> Path:
    daily_dir = out_dir / f"daily_{timestamp}"
    grouped: dict[str, list[KforceJob]] = {}
    for job in jobs:
        grouped.setdefault(job_posted_day(job), []).append(job)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        write_csv(day_jobs, daily_dir / f"{day}_jobs.csv")
        write_json(day_jobs, daily_dir / f"{day}_jobs.json")
    return daily_dir


def load_terms(args: argparse.Namespace) -> list[str]:
    terms = list(args.terms or [])
    if args.terms_file:
        terms.extend(line.strip() for line in args.terms_file.read_text(encoding="utf-8").splitlines())
    return [t for t in terms if t.strip()] or DEFAULT_SEARCH_TERMS


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape Kforce jobs into CSV/JSON/Excel.")
    parser.add_argument("--term", action="append", dest="terms", help="Search term. Repeat for multiple terms.")
    parser.add_argument("--terms-file", type=Path)
    parser.add_argument("--posted-within-days", type=int, default=4)
    parser.add_argument("--all-job-types", action="store_true", help="Include Direct Hire jobs (default: Contract only).")
    parser.add_argument("--keep-w2-f2f-onsite-interview", action="store_true")
    parser.add_argument("--min-hourly-rate", type=float, default=55)
    parser.add_argument("--ai-filter", action="store_true", default=False, help="Use Groq/Llama to filter irrelevant jobs (default: off).")
    parser.add_argument("--no-ai-filter", action="store_true", help="Skip AI filter (default).")
    parser.add_argument("--timeout", type=int, default=20)
    parser.add_argument("--sleep", type=float, default=0.3)
    parser.add_argument("--out-dir", type=Path, default=Path(__file__).resolve().parent / "output")
    parser.add_argument("--no-excel", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    jobs = scrape_kforce(
        load_terms(args),
        args.posted_within_days,
        not args.all_job_types,
        not args.keep_w2_f2f_onsite_interview,
        args.min_hourly_rate,
        args.ai_filter,
        args.timeout,
        args.sleep,
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = args.out_dir / f"kforce_jobs_{timestamp}.csv"
    json_path = args.out_dir / f"kforce_jobs_{timestamp}.json"
    excel_path = args.out_dir / f"kforce_jobs_{timestamp}.xlsx"
    write_csv(jobs, csv_path)
    write_json(jobs, json_path)
    if not args.no_excel:
        write_excel(jobs, excel_path, args.posted_within_days, args.min_hourly_rate)
    daily_dir = write_daily_outputs(jobs, args.out_dir, timestamp)
    print(f"Saved {len(jobs)} jobs")
    print(f"CSV:   {csv_path}")
    print(f"JSON:  {json_path}")
    if not args.no_excel:
        print(f"Excel: {excel_path}")
    print(f"Daily: {daily_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
