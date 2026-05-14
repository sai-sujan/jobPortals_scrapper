#!/usr/bin/env python3
"""Standalone Apex Systems job scraper."""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional
from urllib.parse import quote_plus, urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = "https://www.apexsystems.com"
SEARCH_URL = f"{BASE_URL}/search-results-usa"
DEFAULT_SEARCH_TERMS = [
    "ai engineer",
    "ai ml engineer",
    "aiml engineer",
    "machine learning engineer",
    "ml engineer",
    "generative ai engineer",
    "gen ai engineer",
    "llm engineer",
    "rag engineer",
    "data scientist",
    "senior data scientist",
    "senior python developer",
    "python developer",
    "python engineer",
    "backend python engineer",
    "backend software engineer",
    "software engineer python",
    "django developer",
    "fastapi developer",
    "api developer",
    "full stack software engineer",
    "full stack developer",
    "senior software engineer",
    "software engineer",
    "cloud engineer",
    "aws python developer",
    "azure python developer",
]

TITLE_RANKING_WEIGHTS = [
    ("machine learning engineer", 56),
    ("ml engineer", 54),
    ("ai engineer", 54),
    ("ai/ml", 54),
    ("aiml", 54),
    ("generative ai", 52),
    ("gen ai", 52),
    ("llm", 50),
    ("rag", 48),
    ("senior data scientist", 52),
    ("data scientist", 48),
    ("python developer", 46),
    ("python engineer", 46),
    ("senior python", 50),
    ("django", 44),
    ("fastapi", 44),
    ("backend", 38),
    ("back end", 38),
    ("api developer", 36),
    ("software engineer python", 46),
    ("full stack software engineer", 42),
    ("full stack engineer", 40),
    ("full stack developer", 40),
    ("data engineer", 38),
    ("etl", 32),
    ("cloud engineer", 30),
    ("aws", 28),
    ("azure", 28),
    ("senior software engineer", 32),
    ("software engineer", 24),
    ("software developer", 24),
    ("developer", 12),
    ("engineer", 10),
]
TITLE_EXCLUSION_WEIGHTS = [
    ("data engineer", -100),
    ("frontend", -30),
    ("front end", -30),
    ("ui developer", -20),
    ("ux/ui", -12),
    ("java", -12),
    (".net", -35),
    ("qa", -30),
    ("quality assurance", -30),
    ("desktop support", -40),
    ("help desk", -40),
    ("project manager", -35),
    ("business analyst", -30),
    ("junior", -100),
    ("jr ", -100),
    ("entry level", -100),
]
TITLE_EXCLUSION_PATTERNS = [
    (re.compile(r"\bjava\b.*\bfull\s*stack\b|\bfull\s*stack\b.*\bjava\b|\bjava\b.*\bfullstack\b|\bfullstack\b.*\bjava\b", re.I), "Java full stack title"),
    (re.compile(r"\bjava\b.*\b(?:developer|engineer|architect|backend|software)\b|\b(?:developer|engineer|architect|backend|software)\b.*\bjava\b", re.I), "Java developer/engineer title"),
    (re.compile(r"\bdata\s+engineer\b", re.I), "Data Engineer title"),
    (re.compile(r"\bjunior\b|\bjr\.?\s", re.I), "Junior title"),
    (re.compile(r"\bentry[\s-]level\b", re.I), "Entry-level title"),
]
DISALLOWED_WORK_PATTERNS = [
    (re.compile(r"\bno\s+c2c\b", re.I), "No C2C"),
    (re.compile(r"\bno\s+corp(?:oration)?\s*[- ]?\s*to\s*[- ]?\s*corp(?:oration)?\b", re.I), "No corp-to-corp"),
    (re.compile(r"\bf\s*2\s*f\b", re.I), "F2F"),
    (re.compile(r"\bface\s*[- ]?\s*to\s*[- ]?\s*face\b", re.I), "Face-to-face"),
    (re.compile(r"\bin\s*[- ]?\s*person\s+interview\b", re.I), "In-person interview"),
    (re.compile(r"\bon\s*[- ]?\s*site\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\bonsite\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\bfull\s*[- ]?\s*time\b", re.I), "Full-time"),
    (re.compile(r"\bpermanent\b", re.I), "Permanent"),
    (re.compile(r"\bdirect\s*[- ]?\s*hire\b", re.I), "Direct hire"),
    (re.compile(r"\bcontract\s*[- ]?\s*to\s*[- ]?\s*hire\b", re.I), "Contract-to-Hire"),
    (re.compile(r"\bcth\b", re.I), "Contract-to-Hire (CTH)"),
]
EMAIL_RE = re.compile(r"(?<![A-Za-z0-9._%+-])([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})(?![A-Za-z0-9._%+-])")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?1[\s.-]?)?(?:\(?\d{3}\)?[\s.-]?)\d{3}[\s.-]?\d{4}(?!\d)")


@dataclass
class ApexJob:
    source_company: str
    search_term: str
    title_rank: int
    title_rank_reasons: str
    title: str
    city: str
    state: str
    location: str
    employment_type: str
    min_pay_rate: str
    max_pay_rate: str
    pay_rate_unit: str
    posted_date: str
    job_id: str
    job_url: str
    apply_url: str
    contact_info: str
    description_snippet: str
    raw_text: str


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", html.unescape(text or "")).strip()


def normalized_title(text: str) -> str:
    return re.sub(r"[^a-z0-9+#./-]+", " ", (text or "").lower()).strip()


def title_contains(title: str, phrase: str) -> bool:
    phrase = normalized_title(phrase)
    if len(phrase) <= 3:
        return re.search(rf"(?<![a-z0-9]){re.escape(phrase)}(?![a-z0-9])", title) is not None
    return phrase in title


def score_title(title: str) -> tuple[int, str]:
    normalized = normalized_title(title)
    score = 0
    reasons: list[str] = []
    for phrase, weight in TITLE_RANKING_WEIGHTS:
        if title_contains(normalized, phrase):
            score += weight
            reasons.append(f"{phrase}+{weight}")
    for phrase, weight in TITLE_EXCLUSION_WEIGHTS:
        if title_contains(normalized, phrase):
            score += weight
            reasons.append(f"{phrase}{weight}")
    return max(score, 0), "; ".join(reasons)


def parse_apex_date(value: str) -> Optional[datetime]:
    value = clean_text(value)
    if not value:
        return None
    for fmt in ("%m/%d/%Y", "%B %d, %Y", "%b %d, %Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    return None


def is_within_posted_days(value: str, days: Optional[int], now: Optional[datetime] = None) -> bool:
    if not days or days <= 0:
        return True
    posted = parse_apex_date(value)
    if not posted:
        return False
    now = now or datetime.now(timezone.utc)
    return 0 <= (now - posted).total_seconds() <= days * 24 * 60 * 60


def job_posted_day(job: ApexJob) -> str:
    parsed = parse_apex_date(job.posted_date)
    return parsed.date().isoformat() if parsed else "unknown-date"


def contact_info(text: str) -> str:
    emails = []
    for email in EMAIL_RE.findall(clean_text(text)):
        lowered = email.lower()
        if lowered not in emails:
            emails.append(lowered)
    phones = []
    for phone in PHONE_RE.findall(text or ""):
        digits = re.sub(r"\D", "", phone)
        if len(digits) == 11 and digits.startswith("1"):
            digits = digits[1:]
        if len(digits) == 10:
            display = f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
            if display not in phones:
                phones.append(display)
    return ", ".join(emails + phones)


def disallowed_work_reasons(text: str) -> list[str]:
    reasons: list[str] = []
    cleaned = clean_text(text)
    for pattern, reason in DISALLOWED_WORK_PATTERNS:
        if pattern.search(cleaned) and reason not in reasons:
            reasons.append(reason)
    return reasons


def title_exclusion_reasons(title: str) -> list[str]:
    return [reason for pattern, reason in TITLE_EXCLUSION_PATTERNS if pattern.search(title or "")]


def parse_pay_range(text: str) -> tuple[Optional[float], Optional[float], str]:
    text = clean_text(text)
    match = re.search(r"\$?\s*([0-9]+(?:\.[0-9]+)?)\s*(?:-|to)\s*\$?\s*([0-9]+(?:\.[0-9]+)?)\s*(?:per\s+)?([A-Za-z]+)?", text, re.I)
    if match:
        return float(match.group(1)), float(match.group(2)), clean_text(match.group(3) or "")
    match = re.search(r"\$?\s*([0-9]+(?:\.[0-9]+)?)\s*(?:per\s+)?([A-Za-z]+)?", text, re.I)
    if match:
        rate = float(match.group(1))
        return rate, rate, clean_text(match.group(2) or "")
    return None, None, ""


def is_hourly_unit(unit: str) -> bool:
    return normalized_title(unit) in {"hour", "hourly", "hr"}


def below_min_hourly(job: ApexJob, min_hourly_rate: Optional[float]) -> bool:
    if not min_hourly_rate or min_hourly_rate <= 0 or not is_hourly_unit(job.pay_rate_unit):
        return False
    rates = [float(rate) for rate in (job.min_pay_rate, job.max_pay_rate) if rate]
    return bool(rates) and max(rates) < min_hourly_rate


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": "Mozilla/5.0 VenkataDoraApexAutomation/0.1", "Accept": "text/html,*/*"})
    return session


def search_url(term: str, page: int, rows: int, remote: str = "") -> str:
    return (
        f"{SEARCH_URL}?catalogcode=USA&address=&radius=50&page={page}&rows={rows}"
        f"&query={quote_plus(term)}&remote={quote_plus(remote)}"
    )


def parse_search_results(html_text: str, term: str) -> list[dict[str, str]]:
    soup = BeautifulSoup(html_text, "html.parser")
    table = soup.select_one("table.job-table")
    if not table:
        return []
    jobs: list[dict[str, str]] = []
    for tr in table.select("tbody tr"):
        cells = tr.find_all("td")
        if len(cells) < 4:
            continue
        title_link = cells[0].select_one("a[href]")
        if not title_link:
            continue
        url = urljoin(BASE_URL, title_link["href"])
        job_id_match = re.search(r"/job/([^/]+)/", title_link["href"])
        jobs.append(
            {
                "search_term": term,
                "title": clean_text(title_link.get_text(" ", strip=True)),
                "city": clean_text(cells[1].get_text(" ", strip=True)),
                "state": clean_text(cells[2].get_text(" ", strip=True)),
                "posted_date": clean_text(cells[3].get_text(" ", strip=True)),
                "job_id": job_id_match.group(1) if job_id_match else "",
                "job_url": url,
            }
        )
    return jobs


def fetch_detail(session: requests.Session, url: str, timeout: int) -> dict[str, Any]:
    response = session.get(url, timeout=timeout)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")
    description_node = soup.select_one(".job-desc-wrapper")
    sidebar_node = soup.select_one(".job-snapshot-wrapper")
    apply_link = soup.select_one("a.apply-btn[href]")
    description = clean_text(description_node.get_text(" ", strip=True) if description_node else "")
    sidebar_text = clean_text(sidebar_node.get_text(" ", strip=True) if sidebar_node else "")
    employment = ""
    location = ""
    posted_date = ""
    pay_text = ""
    for label, field in [("Employee Type:", "employment"), ("Location:", "location"), ("Date Posted:", "posted"), ("Pay Range:", "pay")]:
        match = re.search(re.escape(label) + r"\s*(.*?)(?=Employee Type:|Location:|Job Type:|Date Posted:|Pay Range:|Similar Jobs|$)", sidebar_text)
        if match:
            value = clean_text(match.group(1))
            if field == "employment":
                employment = value
            elif field == "location":
                location = value
            elif field == "posted":
                posted_date = value
            elif field == "pay":
                pay_text = value
    min_pay, max_pay, unit = parse_pay_range(pay_text)
    return {
        "description": description,
        "employment_type": employment,
        "location": location,
        "posted_date": posted_date,
        "min_pay": min_pay,
        "max_pay": max_pay,
        "pay_unit": unit,
        "apply_url": apply_link["href"] if apply_link else "",
    }


def normalize_job(row: dict[str, str], detail: dict[str, Any]) -> ApexJob:
    title_rank, title_rank_reasons = score_title(row["title"])
    description = detail.get("description") or ""
    return ApexJob(
        source_company="Apex Systems",
        search_term=row["search_term"],
        title_rank=title_rank,
        title_rank_reasons=title_rank_reasons,
        title=row["title"],
        city=row["city"],
        state=row["state"],
        location=detail.get("location") or ", ".join(part for part in (row["city"], row["state"]) if part),
        employment_type=detail.get("employment_type") or "",
        min_pay_rate="" if detail.get("min_pay") is None else f"{detail['min_pay']:g}",
        max_pay_rate="" if detail.get("max_pay") is None else f"{detail['max_pay']:g}",
        pay_rate_unit=detail.get("pay_unit") or "",
        posted_date=detail.get("posted_date") or row["posted_date"],
        job_id=row["job_id"],
        job_url=row["job_url"],
        apply_url=detail.get("apply_url") or "",
        contact_info=contact_info(description),
        description_snippet=description[:900],
        raw_text=description,
    )


def scrape_apex(
    search_terms: Iterable[str],
    posted_within_days: Optional[int],
    exclude_disallowed_work: bool,
    min_hourly_rate: Optional[float],
    rows_per_search: int,
    timeout: int,
    sleep_seconds: float,
) -> list[ApexJob]:
    session = make_session()
    seen: set[str] = set()
    candidates: list[dict[str, str]] = []
    for term in search_terms:
        term = term.strip()
        if not term:
            continue
        print(f"Searching Apex Systems: {term}")
        response = session.get(search_url(term, 1, rows_per_search), timeout=timeout)
        response.raise_for_status()
        for row in parse_search_results(response.text, term):
            if not is_within_posted_days(row["posted_date"], posted_within_days):
                continue
            key = row["job_id"] or row["job_url"]
            if key in seen:
                continue
            seen.add(key)
            candidates.append(row)
        time.sleep(sleep_seconds)

    jobs: list[ApexJob] = []
    for row in candidates:
        try:
            detail = fetch_detail(session, row["job_url"], timeout)
        except requests.RequestException as exc:
            print(f"Detail failed for {row['job_url']}: {exc}")
            detail = {}
        job = normalize_job(row, detail)
        excluded_title_reasons = title_exclusion_reasons(job.title)
        if excluded_title_reasons:
            print(f"Excluded {job.job_id}: {', '.join(excluded_title_reasons)}")
            continue
        if exclude_disallowed_work and disallowed_work_reasons(" ".join([job.title, job.employment_type, job.raw_text])):
            print(f"Excluded {job.job_id}: {', '.join(disallowed_work_reasons(job.raw_text)) or 'disallowed work signal'}")
            continue
        if below_min_hourly(job, min_hourly_rate):
            print(f"Excluded {job.job_id}: below ${min_hourly_rate:g}/hour pay threshold (max {job.max_pay_rate} {job.pay_rate_unit})")
            continue
        if job.pay_rate_unit == "year":
            print(f"  Excluded (annual salary = full-time signal): {job.title}")
            continue
        jobs.append(job)
        time.sleep(sleep_seconds)
    return sort_jobs(jobs)


def sort_jobs(jobs: Iterable[ApexJob]) -> list[ApexJob]:
    return sorted(jobs, key=lambda job: (job_posted_day(job), bool(job.contact_info), job.title_rank, job.title.lower()), reverse=True)


def write_csv(jobs: list[ApexJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(ApexJob.__dataclass_fields__.keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for job in jobs:
            writer.writerow(asdict(job))


def write_json(jobs: list[ApexJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump([asdict(job) for job in jobs], handle, indent=2)


EXCEL_COLUMNS = [
    ("posted_day", "Posted Day"),
    ("title_rank", "Title Rank"),
    ("title", "Title"),
    ("location", "Location"),
    ("employment_type", "Employment"),
    ("pay_range", "Pay"),
    ("contact_info", "Contact Info"),
    ("search_term", "Search Term"),
    ("job_id", "Job ID"),
    ("job_url", "Job URL"),
    ("apply_url", "Apply URL"),
    ("title_rank_reasons", "Rank Reasons"),
]


def excel_row(job: ApexJob) -> dict[str, Any]:
    row = asdict(job)
    row["posted_day"] = job_posted_day(job)
    row["pay_range"] = ""
    if job.min_pay_rate or job.max_pay_rate:
        row["pay_range"] = f"${job.min_pay_rate or job.max_pay_rate}-${job.max_pay_rate or job.min_pay_rate} / {job.pay_rate_unit}".strip()
    return row


def style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="0B2D3D")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [13, 11, 34, 26, 16, 18, 30, 24, 18, 14, 14, 45]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        rank = row[1].value or 0
        if isinstance(rank, int) and rank >= 45:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="E8F5E9")
        if row[6].value:
            row[6].fill = PatternFill("solid", fgColor="FFF3CD")


def append_jobs_sheet(workbook: Workbook, sheet_name: str, jobs: list[ApexJob]) -> None:
    safe_name = re.sub(r"[\[\]:*?/\\]", "-", sheet_name)[:31] or "Sheet"
    sheet = workbook.create_sheet(safe_name)
    sheet.append([label for _, label in EXCEL_COLUMNS])
    for job in jobs:
        source = excel_row(job)
        sheet.append([source.get(key, "") for key, _ in EXCEL_COLUMNS])
        for col in (10, 11):
            cell = sheet.cell(row=sheet.max_row, column=col)
            url = job.job_url if col == 10 else job.apply_url
            if url:
                cell.value = "Open Job" if col == 10 else "Apply"
                cell.hyperlink = url
                cell.style = "Hyperlink"
    style_sheet(sheet)


def write_excel(jobs: list[ApexJob], path: Path, posted_within_days: Optional[int], min_hourly_rate: Optional[float]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "Apex Systems Daily Jobs"
    summary["A1"].font = Font(size=18, bold=True, color="0B2D3D")
    summary["A3"] = "Generated"
    summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary["A4"] = "Posting Window"
    summary["B4"] = "All dates" if not posted_within_days else f"Last {posted_within_days} days"
    summary["A5"] = "Minimum Hourly Pay"
    summary["B5"] = "Disabled" if not min_hourly_rate else f"${min_hourly_rate:g}/hour"
    summary["A6"] = "Total Jobs"
    summary["B6"] = len(jobs)
    summary.append([])
    summary.append(["Day", "Jobs", "With Contact", "Top Rank"])
    for cell in summary[8]:
        cell.fill = PatternFill("solid", fgColor="0B2D3D")
        cell.font = Font(color="FFFFFF", bold=True)
    grouped: dict[str, list[ApexJob]] = {}
    for job in jobs:
        grouped.setdefault(job_posted_day(job), []).append(job)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        summary.append([day, len(day_jobs), sum(1 for job in day_jobs if job.contact_info), max((job.title_rank for job in day_jobs), default=0)])
    for column in range(1, 5):
        summary.column_dimensions[get_column_letter(column)].width = 22
    append_jobs_sheet(workbook, "All Jobs", jobs)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        append_jobs_sheet(workbook, day, sorted(day_jobs, key=lambda job: (not bool(job.contact_info), -job.title_rank, job.title.lower())))
    workbook.save(path)


def write_daily_outputs(jobs: list[ApexJob], out_dir: Path, timestamp: str) -> Path:
    daily_dir = out_dir / f"daily_{timestamp}"
    grouped: dict[str, list[ApexJob]] = {}
    for job in jobs:
        grouped.setdefault(job_posted_day(job), []).append(job)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        write_csv(day_jobs, daily_dir / f"{day}_jobs.csv")
        write_json(day_jobs, daily_dir / f"{day}_jobs.json")
    return daily_dir


def load_terms(args: argparse.Namespace) -> list[str]:
    terms = list(args.terms or [])
    if args.terms_file:
        terms.extend(line.strip() for line in args.terms_file.read_text(encoding="utf-8").splitlines())
    return [term for term in terms if term.strip()] or DEFAULT_SEARCH_TERMS


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape Apex Systems jobs into CSV/JSON/Excel.")
    parser.add_argument("--term", action="append", dest="terms", help="Search term. Repeat for multiple terms.")
    parser.add_argument("--terms-file", type=Path)
    parser.add_argument("--posted-within-days", type=int, default=4)
    parser.add_argument("--keep-w2-f2f-onsite-interview", action="store_true")
    parser.add_argument("--min-hourly-rate", type=float, default=55)
    parser.add_argument("--rows-per-search", type=int, default=100)
    parser.add_argument("--timeout", type=int, default=20)
    parser.add_argument("--sleep", type=float, default=0.2)
    parser.add_argument("--out-dir", type=Path, default=Path(__file__).resolve().parent / "output")
    parser.add_argument("--no-excel", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    jobs = scrape_apex(
        load_terms(args),
        args.posted_within_days,
        not args.keep_w2_f2f_onsite_interview,
        args.min_hourly_rate,
        args.rows_per_search,
        args.timeout,
        args.sleep,
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = args.out_dir / f"apex_jobs_{timestamp}.csv"
    json_path = args.out_dir / f"apex_jobs_{timestamp}.json"
    excel_path = args.out_dir / f"apex_jobs_{timestamp}.xlsx"
    write_csv(jobs, csv_path)
    write_json(jobs, json_path)
    if not args.no_excel:
        write_excel(jobs, excel_path, args.posted_within_days, args.min_hourly_rate)
    daily_dir = write_daily_outputs(jobs, args.out_dir, timestamp)
    print(f"Saved {len(jobs)} jobs")
    print(f"CSV: {csv_path}")
    print(f"JSON: {json_path}")
    if not args.no_excel:
        print(f"Excel: {excel_path}")
    print(f"Daily ranked files: {daily_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
