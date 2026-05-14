#!/usr/bin/env python3
"""Shared filtering/output helpers for the separate vendor scraper folders."""

from __future__ import annotations

import csv
import html
import json
import re
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


TITLE_RANKING_WEIGHTS = [
    ("principal ai architect", 60), ("lead ai engineer", 58), ("machine learning engineer", 56),
    ("ml engineer", 54), ("ai engineer", 54), ("ai/ml", 54), ("ai consultant", 52),
    ("generative ai", 52), ("gen ai", 52), ("agentic ai", 50), ("llm", 50), ("rag", 48),
    ("senior data scientist", 52), ("data scientist", 48), ("senior python", 50),
    ("python developer", 46), ("python engineer", 46), ("software engineer python", 46),
    ("django", 44), ("fastapi", 44), ("full stack software engineer", 42),
    ("full stack engineer", 40), ("full stack developer", 40), ("data engineer", 42),
    ("data engineering", 42), ("big data engineer", 42), ("data architect", 38),
    ("etl", 36), ("data pipeline", 36), ("backend", 38), ("back-end", 38), ("back end", 38),
    ("api developer", 36), ("api engineer", 36), ("cloud engineer", 30), ("aws", 28),
    ("azure", 28), ("gcp", 28), ("senior software engineer", 34),
    ("software development engineer", 32), ("software engineer", 24), ("software developer", 24),
    ("developer", 12), ("engineer", 10),
]

CONTENT_BONUS_WEIGHTS = [
    ("python", 12), ("django", 10), ("fastapi", 10), ("machine learning", 12),
    ("generative ai", 12), ("agentic ai", 12), ("llm", 10), ("rag", 10), ("langchain", 10),
    ("langgraph", 10), ("data engineer", 12), ("data engineering", 12), ("etl", 8),
    ("data pipeline", 8), ("bigquery", 8), ("spark", 8), ("snowflake", 8), ("react", 6),
    ("node.js", 6), ("aws", 6), ("azure", 6), ("gcp", 6),
]

TITLE_EXCLUSION_WEIGHTS = [
    ("qa", -30), ("quality assurance", -30), ("desktop support", -40), ("help desk", -40),
    ("project manager", -35), ("program manager", -35), ("business analyst", -30),
    ("product manager", -60), ("product owner", -60), ("network engineer", -45),
    ("data center", -35), ("civil engineer", -100), ("mechanical engineer", -100),
    ("electrical engineer", -100), ("systems administrator", -50), ("scrum master", -60),
    ("agile coach", -60), ("it support", -60), ("sap", -50), ("salesforce administrator", -45),
    ("salesforce developer", -30), ("ios developer", -50), ("android developer", -50),
    ("mobile developer", -40), ("junior", -100), ("jr ", -100), ("entry level", -100),
]

TITLE_EXCLUSION_PATTERNS = [
    (re.compile(r"\bjunior\b|\bjr\.?\s", re.I), "Junior title"),
    (re.compile(r"\bentry[\s-]level\b", re.I), "Entry-level title"),
    (re.compile(r"\bintern(ship)?\b", re.I), "Intern title"),
    (re.compile(r"\bembedded\b", re.I), "Embedded title"),
]

ROLE_EXCLUSION_PATTERNS = [
    (re.compile(r"\bembedded\s+software\b", re.I), "Embedded software"),
    (re.compile(r"\bmissile\s+systems?\b", re.I), "Missile systems"),
    (re.compile(r"\bros\s*/?\s*ros2\b|\bros2?\b", re.I), "ROS/hardware role"),
    (re.compile(r"\belectronics\s+hardware\b", re.I), "Electronics hardware"),
    (re.compile(r"\bwiring\s+a\s+motor\b", re.I), "Hardware wiring"),
]

DISALLOWED_WORK_PATTERNS = [
    (re.compile(r"\bno\s+c2c\b", re.I), "No C2C"),
    (re.compile(r"\bno\s+corp(?:oration)?\s*[- ]?\s*to\s*[- ]?\s*corp(?:oration)?\b", re.I), "No corp-to-corp"),
    (re.compile(r"\bnot\s+(?:open\s+to\s+)?c2c\b", re.I), "Not open to C2C"),
    (re.compile(r"\bcannot\s+(?:do|support|accept)\s+c2c\b", re.I), "Cannot support C2C"),
    (re.compile(r"\bunable\s+to\s+(?:do|support|accept)\s+c2c\b", re.I), "Unable to support C2C"),
    (re.compile(r"\bmust\s+be\s+willing\s+and\s+able\s+to\s+work\s+on\s+a\s+w\s*[- ]?\s*2\s+basis\b", re.I), "W2 basis required"),
    (re.compile(r"\bw\s*[- ]?\s*2\s+only\b", re.I), "W2 only"),
    (re.compile(r"\bw\s*[- ]?\s*2\s+contract\b", re.I), "W2 contract"),
    (re.compile(r"\bw\s*[- ]?\s*2\s+hourly\s+rate\b", re.I), "W2 hourly rate"),
    (re.compile(r"\bw\s*[- ]?\s*2\b", re.I), "W2"),
    (re.compile(r"\bf\s*2\s*f\b", re.I), "F2F"),
    (re.compile(r"\bface\s*[- ]?\s*to\s*[- ]?\s*face\b", re.I), "Face-to-face"),
    (re.compile(r"\bin\s*[- ]?\s*person\s+interview\b", re.I), "In-person interview"),
    (re.compile(r"\bon\s*[- ]?\s*site\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\blocal\s+candidates?\s+only\b", re.I), "Local candidates only"),
    (re.compile(r"\bpermanent\b", re.I), "Permanent"),
    (re.compile(r"\bdirect\s*[- ]?\s*hire\b", re.I), "Direct hire"),
]

EMAIL_RE = re.compile(r"(?<![A-Za-z0-9._%+-])([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})(?![A-Za-z0-9._%+-])")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?1[\s.-]?)?(?:\(?\d{3}\)?[\s.-]?)\d{3}[\s.-]?\d{4}(?!\d)")
MIN_TITLE_RANK = 20


@dataclass
class VendorJob:
    source_company: str
    search_term: str
    title_rank: int
    title_rank_reasons: str
    title: str
    category: str
    location: str
    employment_type: str
    salary: str
    posted_date: str
    job_id: str
    job_url: str
    apply_url: str
    contact_info: str
    description_snippet: str
    raw_text: str


def clean_text(text: Any) -> str:
    text = re.sub(r"<[^>]+>", " ", str(text or ""))
    return re.sub(r"\s+", " ", html.unescape(text)).strip()


def normalized_text(text: str) -> str:
    return re.sub(r"[^a-z0-9+#./-]+", " ", (text or "").lower()).strip()


def text_contains(haystack: str, phrase: str) -> bool:
    phrase = normalized_text(phrase)
    if not phrase:
        return False
    if len(phrase) <= 3:
        return re.search(rf"(?<![a-z0-9]){re.escape(phrase)}(?![a-z0-9])", haystack) is not None
    return phrase in haystack


def score_title(title: str, raw_text: str) -> tuple[int, str]:
    normalized_title = normalized_text(title)
    normalized_raw = normalized_text(raw_text)
    score = 0
    reasons: list[str] = []
    for phrase, weight in TITLE_RANKING_WEIGHTS:
        if text_contains(normalized_title, phrase):
            score += weight
            reasons.append(f"{phrase}+{weight}")
    for phrase, weight in CONTENT_BONUS_WEIGHTS:
        if text_contains(normalized_raw, phrase):
            score += weight
            reasons.append(f"{phrase}+{weight} desc")
    for phrase, weight in TITLE_EXCLUSION_WEIGHTS:
        if text_contains(normalized_title, phrase):
            score += weight
            reasons.append(f"{phrase}{weight}")
    return max(score, 0), "; ".join(reasons)


def parse_posted_date(value: Any) -> Optional[datetime]:
    if not value:
        return None
    text = str(value).strip()
    if re.fullmatch(r"\d{10,13}", text):
        stamp = int(text)
        if stamp > 9_999_999_999:
            stamp = stamp / 1000
        return datetime.fromtimestamp(stamp, tz=timezone.utc)
    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    if re.search(r"[+-]\d{4}$", text):
        text = text[:-5] + text[-5:-2] + ":" + text[-2:]
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def posted_day(job: VendorJob) -> str:
    parsed = parse_posted_date(job.posted_date)
    return parsed.date().isoformat() if parsed else "unknown-date"


def is_within_posted_days(posted_date: str, days: Optional[int]) -> bool:
    if not days or days <= 0:
        return True
    parsed = parse_posted_date(posted_date)
    if not parsed:
        return True
    return 0 <= (datetime.now(timezone.utc) - parsed).total_seconds() <= days * 86400


def extract_contact_info(text: str) -> str:
    emails = []
    for email in EMAIL_RE.findall(clean_text(text)):
        lowered = email.lower()
        if lowered not in emails:
            emails.append(lowered)
    phones = []
    for phone in PHONE_RE.findall(text or ""):
        digits = re.sub(r"\D", "", phone)
        if len(digits) == 11 and digits.startswith("1"):
            digits = digits[1:]
        if len(digits) == 10:
            display = f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
            if display not in phones:
                phones.append(display)
    return ", ".join(emails + phones)


def disallowed_work_reasons(text: str) -> list[str]:
    cleaned = clean_text(text)
    reasons = []
    for pattern, reason in DISALLOWED_WORK_PATTERNS:
        if pattern.search(cleaned) and reason not in reasons:
            reasons.append(reason)
    return reasons


def filter_and_sort_jobs(jobs: Iterable[VendorJob], posted_within_days: Optional[int], exclude_disallowed_work: bool) -> list[VendorJob]:
    kept = []
    for job in jobs:
        if not is_within_posted_days(job.posted_date, posted_within_days):
            continue
        if job.title_rank < MIN_TITLE_RANK:
            print(f"  Skipped low-rank ({job.title_rank}): {job.title}")
            continue
        title_reasons = [reason for pattern, reason in TITLE_EXCLUSION_PATTERNS if pattern.search(job.title or "")]
        if title_reasons:
            print(f"  Excluded ({', '.join(title_reasons)}): {job.title}")
            continue
        role_reasons = [reason for pattern, reason in ROLE_EXCLUSION_PATTERNS if pattern.search(" ".join([job.title, job.raw_text]))]
        if role_reasons:
            print(f"  Excluded ({', '.join(role_reasons)}): {job.title}")
            continue
        if exclude_disallowed_work:
            reasons = disallowed_work_reasons(" ".join([job.title, job.employment_type, job.raw_text]))
            if reasons:
                print(f"  Excluded ({', '.join(reasons)}): {job.title}")
                continue
        kept.append(job)
    return sorted(kept, key=lambda job: (posted_day(job), bool(job.contact_info), job.title_rank, job.title.lower()), reverse=True)


def write_outputs(company_slug: str, jobs: list[VendorJob], out_dir: Path, posted_within_days: Optional[int], no_excel: bool) -> None:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir.mkdir(parents=True, exist_ok=True)
    csv_path = out_dir / f"{company_slug}_jobs_{timestamp}.csv"
    json_path = out_dir / f"{company_slug}_jobs_{timestamp}.json"
    xlsx_path = out_dir / f"{company_slug}_jobs_{timestamp}.xlsx"
    fields = list(VendorJob.__dataclass_fields__.keys())
    with csv_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for job in jobs:
            writer.writerow(asdict(job))
    with json_path.open("w", encoding="utf-8") as handle:
        json.dump([asdict(job) for job in jobs], handle, indent=2)
    if not no_excel:
        write_excel(jobs, xlsx_path, posted_within_days, company_slug)
    daily_dir = out_dir / f"daily_{timestamp}"
    grouped: dict[str, list[VendorJob]] = {}
    for job in jobs:
        grouped.setdefault(posted_day(job), []).append(job)
    daily_dir.mkdir(parents=True, exist_ok=True)
    for day, day_jobs in grouped.items():
        with (daily_dir / f"{day}_jobs.json").open("w", encoding="utf-8") as handle:
            json.dump([asdict(job) for job in day_jobs], handle, indent=2)
    print(f"\nSaved {len(jobs)} jobs")
    print(f"CSV:   {csv_path}")
    print(f"JSON:  {json_path}")
    if not no_excel:
        print(f"Excel: {xlsx_path}")
    print(f"Daily: {daily_dir}")


def write_excel(jobs: list[VendorJob], path: Path, posted_within_days: Optional[int], company_slug: str) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = f"{company_slug} Daily Jobs"
    summary["A1"].font = Font(size=18, bold=True, color="1F3864")
    summary["A3"] = "Generated"
    summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary["A4"] = "Posting Window"
    summary["B4"] = "All dates" if not posted_within_days else f"Last {posted_within_days} days"
    summary["A5"] = "Total Jobs"
    summary["B5"] = len(jobs)
    append_jobs_sheet(workbook, "All Jobs", jobs)
    workbook.save(path)


def append_jobs_sheet(workbook: Workbook, sheet_name: str, jobs: list[VendorJob]) -> None:
    columns = [
        ("posted_day", "Posted Day"), ("title_rank", "Title Rank"), ("title", "Title"),
        ("location", "Location"), ("employment_type", "Employment"), ("salary", "Salary"),
        ("contact_info", "Contact Info"), ("job_id", "Job ID"), ("job_url", "Job URL"),
        ("apply_url", "Apply URL"), ("title_rank_reasons", "Rank Reasons"),
    ]
    sheet = workbook.create_sheet(re.sub(r"[\[\]:*?/\\]", "-", sheet_name)[:31] or "Sheet")
    sheet.append([label for _, label in columns])
    for job in jobs:
        source = asdict(job)
        source["posted_day"] = posted_day(job)
        sheet.append([source.get(key, "") for key, _ in columns])
        for column, url in ((9, job.job_url), (10, job.apply_url)):
            if url:
                cell = sheet.cell(row=sheet.max_row, column=column)
                cell.value = "Open Job" if column == 9 else "Apply"
                cell.hyperlink = url
                cell.style = "Hyperlink"
    fill = PatternFill("solid", fgColor="1F3864")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, width in enumerate([13, 11, 42, 28, 22, 22, 32, 18, 14, 14, 55], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        if isinstance(row[1].value, int) and row[1].value >= 45:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="E8F5E9")
