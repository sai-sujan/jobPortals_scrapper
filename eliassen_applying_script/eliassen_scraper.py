#!/usr/bin/env python3
"""Standalone Eliassen Group scraper using the public Atom jobs feed."""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import time
import xml.etree.ElementTree as ET
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional
from urllib.parse import urlparse

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = "https://careers.eliassen.com"
FEED_URL = f"{BASE_URL}/feeds/jobs.atom"
ATOM_NS = {"a": "http://www.w3.org/2005/Atom"}

TITLE_RANKING_WEIGHTS = [
    ("principal ai architect", 60),
    ("lead ai engineer", 58),
    ("machine learning engineer", 56),
    ("ml engineer", 54),
    ("ai engineer", 54),
    ("ai consultant", 52),
    ("generative ai", 52),
    ("gen ai", 52),
    ("agentic ai", 50),
    ("llm", 50),
    ("rag", 48),
    ("senior data scientist", 52),
    ("data scientist", 48),
    ("senior python", 50),
    ("python developer", 46),
    ("python engineer", 46),
    ("software engineer python", 46),
    ("django", 44),
    ("fastapi", 44),
    ("full stack software engineer", 42),
    ("full stack engineer", 40),
    ("full stack developer", 40),
    ("data engineer", 42),
    ("data engineering", 42),
    ("big data engineer", 42),
    ("data architect", 38),
    ("etl", 36),
    ("data pipeline", 36),
    ("backend", 38),
    ("back-end", 38),
    ("back end", 38),
    ("api developer", 36),
    ("api engineer", 36),
    ("cloud engineer", 30),
    ("aws", 28),
    ("azure", 28),
    ("gcp", 28),
    ("senior software engineer", 34),
    ("software development engineer", 32),
    ("software engineer", 24),
    ("software developer", 24),
    ("developer", 12),
    ("engineer", 10),
]

CONTENT_BONUS_WEIGHTS = [
    ("python", 12),
    ("django", 10),
    ("fastapi", 10),
    ("machine learning", 12),
    ("generative ai", 12),
    ("agentic ai", 12),
    ("llm", 10),
    ("rag", 10),
    ("langchain", 10),
    ("langgraph", 10),
    ("data engineer", 12),
    ("data engineering", 12),
    ("etl", 8),
    ("data pipeline", 8),
    ("bigquery", 8),
    ("spark", 8),
    ("snowflake", 8),
    ("react", 6),
    ("node.js", 6),
    ("aws", 6),
    ("azure", 6),
    ("gcp", 6),
]

TITLE_EXCLUSION_WEIGHTS = [
    ("qa", -30),
    ("quality assurance", -30),
    ("desktop support", -40),
    ("help desk", -40),
    ("project manager", -35),
    ("program manager", -35),
    ("business analyst", -30),
    ("product analyst", -35),
    ("product manager", -60),
    ("product owner", -60),
    ("network engineer", -45),
    ("tactical network engineer", -80),
    ("data center", -35),
    ("civil engineer", -100),
    ("mechanical engineer", -100),
    ("electrical engineer", -100),
    ("systems administrator", -50),
    ("scrum master", -60),
    ("agile coach", -60),
    ("it support", -60),
    ("sap", -50),
    ("salesforce administrator", -45),
    ("salesforce developer", -30),
    ("ios developer", -50),
    ("android developer", -50),
    ("mobile developer", -40),
    ("junior", -100),
    ("jr ", -100),
    ("entry level", -100),
]

TITLE_EXCLUSION_PATTERNS = [
    (re.compile(r"\bjunior\b|\bjr\.?\s", re.I), "Junior title"),
    (re.compile(r"\bentry[\s-]level\b", re.I), "Entry-level title"),
    (re.compile(r"\bintern(ship)?\b", re.I), "Intern title"),
    (re.compile(r"\bembedded\b", re.I), "Embedded title"),
]

ROLE_EXCLUSION_PATTERNS = [
    (re.compile(r"\bembedded\s+software\b", re.I), "Embedded software"),
    (re.compile(r"\bmissile\s+systems?\b", re.I), "Missile systems"),
    (re.compile(r"\bros\s*/?\s*ros2\b|\bros2?\b", re.I), "ROS/hardware role"),
    (re.compile(r"\belectronics\s+hardware\b", re.I), "Electronics hardware"),
    (re.compile(r"\bwiring\s+a\s+motor\b", re.I), "Hardware wiring"),
]

DISALLOWED_WORK_PATTERNS = [
    (re.compile(r"\bno\s+c2c\b", re.I), "No C2C"),
    (re.compile(r"\bno\s+corp(?:oration)?\s*[- ]?\s*to\s*[- ]?\s*corp(?:oration)?\b", re.I), "No corp-to-corp"),
    (re.compile(r"\bnot\s+(?:open\s+to\s+)?c2c\b", re.I), "Not open to C2C"),
    (re.compile(r"\bcannot\s+(?:do|support|accept)\s+c2c\b", re.I), "Cannot support C2C"),
    (re.compile(r"\bunable\s+to\s+(?:do|support|accept)\s+c2c\b", re.I), "Unable to support C2C"),
    (re.compile(r"\bmust\s+be\s+willing\s+and\s+able\s+to\s+work\s+on\s+a\s+w\s*[- ]?\s*2\s+basis\b", re.I), "W2 basis required"),
    (re.compile(r"\bw\s*[- ]?\s*2\s+only\b", re.I), "W2 only"),
    (re.compile(r"\bw\s*[- ]?\s*2\s+contract\b", re.I), "W2 contract"),
    (re.compile(r"\bw\s*[- ]?\s*2\s+hourly\s+rate\b", re.I), "W2 hourly rate"),
    (re.compile(r"\bw\s*[- ]?\s*2\b", re.I), "W2"),
    (re.compile(r"\bf\s*2\s*f\b", re.I), "F2F"),
    (re.compile(r"\bface\s*[- ]?\s*to\s*[- ]?\s*face\b", re.I), "Face-to-face"),
    (re.compile(r"\bin\s*[- ]?\s*person\s+interview\b", re.I), "In-person interview"),
    (re.compile(r"\bon\s*[- ]?\s*site\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\blocal\s+candidates?\s+only\b", re.I), "Local candidates only"),
    (re.compile(r"\bpermanent\b", re.I), "Permanent"),
    (re.compile(r"\bdirect\s*[- ]?\s*hire\b", re.I), "Direct hire"),
]

EMAIL_RE = re.compile(r"(?<![A-Za-z0-9._%+-])([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})(?![A-Za-z0-9._%+-])")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?1[\s.-]?)?(?:\(?\d{3}\)?[\s.-]?)\d{3}[\s.-]?\d{4}(?!\d)")
MIN_TITLE_RANK = 20


@dataclass
class EliassenJob:
    source_company: str
    search_term: str
    title_rank: int
    title_rank_reasons: str
    title: str
    category: str
    location: str
    employment_type: str
    salary: str
    posted_date: str
    job_id: str
    job_url: str
    apply_url: str
    contact_info: str
    description_snippet: str
    raw_text: str


def clean_text(text: Any) -> str:
    text = re.sub(r"<[^>]+>", " ", str(text or ""))
    return re.sub(r"\s+", " ", html.unescape(text)).strip()


def normalized_text(text: str) -> str:
    return re.sub(r"[^a-z0-9+#./-]+", " ", (text or "").lower()).strip()


def text_contains(haystack: str, phrase: str) -> bool:
    phrase = normalized_text(phrase)
    if not phrase:
        return False
    if len(phrase) <= 3:
        return re.search(rf"(?<![a-z0-9]){re.escape(phrase)}(?![a-z0-9])", haystack) is not None
    return phrase in haystack


def score_title(title: str, raw_text: str) -> tuple[int, str]:
    normalized_title = normalized_text(title)
    normalized_raw = normalized_text(raw_text)
    score = 0
    reasons: list[str] = []
    for phrase, weight in TITLE_RANKING_WEIGHTS:
        if text_contains(normalized_title, phrase):
            score += weight
            reasons.append(f"{phrase}+{weight}")
    for phrase, weight in CONTENT_BONUS_WEIGHTS:
        if text_contains(normalized_raw, phrase):
            score += weight
            reasons.append(f"{phrase}+{weight} desc")
    for phrase, weight in TITLE_EXCLUSION_WEIGHTS:
        if text_contains(normalized_title, phrase):
            score += weight
            reasons.append(f"{phrase}{weight}")
    return max(score, 0), "; ".join(reasons)


def parse_posted_date(value: Any) -> Optional[datetime]:
    if not value:
        return None
    text = str(value).strip()
    if text.endswith("Z"):
        text = text[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(text)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def posted_day(job: EliassenJob) -> str:
    parsed = parse_posted_date(job.posted_date)
    return parsed.date().isoformat() if parsed else "unknown-date"


def is_within_posted_days(posted_date: str, days: Optional[int]) -> bool:
    if not days or days <= 0:
        return True
    parsed = parse_posted_date(posted_date)
    if not parsed:
        return True
    return 0 <= (datetime.now(timezone.utc) - parsed).total_seconds() <= days * 86400


def title_exclusion_reasons(title: str) -> list[str]:
    return [reason for pattern, reason in TITLE_EXCLUSION_PATTERNS if pattern.search(title or "")]


def role_exclusion_reasons(text: str) -> list[str]:
    return [reason for pattern, reason in ROLE_EXCLUSION_PATTERNS if pattern.search(text or "")]


def disallowed_work_reasons(text: str) -> list[str]:
    cleaned = clean_text(text)
    reasons: list[str] = []
    for pattern, reason in DISALLOWED_WORK_PATTERNS:
        if pattern.search(cleaned) and reason not in reasons:
            reasons.append(reason)
    return reasons


def extract_contact_info(text: str) -> str:
    emails: list[str] = []
    for email in EMAIL_RE.findall(clean_text(text)):
        lowered = email.lower()
        if lowered not in emails:
            emails.append(lowered)
    phones: list[str] = []
    for phone in PHONE_RE.findall(text or ""):
        digits = re.sub(r"\D", "", phone)
        if len(digits) == 11 and digits.startswith("1"):
            digits = digits[1:]
        if len(digits) == 10:
            display = f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
            if display not in phones:
                phones.append(display)
    return ", ".join(emails + phones)


def extract_location(summary_text: str, job_url: str) -> str:
    first = re.search(r"\b(Remote|Hybrid|Onsite)\b", summary_text or "", re.I)
    parsed = urlparse(job_url)
    slug = parsed.path.strip("/").split("/")[-1] if parsed.path else ""
    parts = slug.split("-")
    if len(parts) >= 3:
        tail = " ".join(parts[-4:-1])
        if tail and not tail.lower().startswith("a1"):
            return clean_text(f"{first.group(1)} {tail}" if first else tail)
    return first.group(1) if first else ""


def extract_salary(text: str) -> str:
    match = re.search(r"\bRate:\s*([^.\n\r<]+)", text or "", re.I)
    return clean_text(match.group(1)) if match else ""


def fetch_feed(timeout: int) -> bytes:
    response = requests.get(FEED_URL, headers={"User-Agent": "Mozilla/5.0", "Accept": "application/atom+xml,application/xml,text/xml,*/*"}, timeout=timeout)
    response.raise_for_status()
    return response.content


def normalize_entry(entry: ET.Element) -> EliassenJob:
    title = clean_text(entry.findtext("a:title", default="", namespaces=ATOM_NS))
    summary_html = entry.findtext("a:summary", default="", namespaces=ATOM_NS)
    raw_text = clean_text(summary_html)
    link = entry.find("a:link", ATOM_NS)
    job_url = (link.attrib.get("href") if link is not None else "") or BASE_URL
    published = entry.findtext("a:published", default="", namespaces=ATOM_NS)
    entry_id = entry.findtext("a:id", default="", namespaces=ATOM_NS).rsplit("/", 1)[-1]
    title_rank, title_rank_reasons = score_title(title, raw_text)
    return EliassenJob(
        source_company="Eliassen Group",
        search_term="all-feed",
        title_rank=title_rank,
        title_rank_reasons=title_rank_reasons,
        title=title or "Untitled job",
        category="",
        location=extract_location(raw_text, job_url),
        employment_type="",
        salary=extract_salary(raw_text),
        posted_date=published,
        job_id=entry_id,
        job_url=job_url.replace("http://", "https://"),
        apply_url=job_url.replace("http://", "https://"),
        contact_info=extract_contact_info(raw_text),
        description_snippet=raw_text[:900],
        raw_text=raw_text,
    )


def sort_jobs(jobs: Iterable[EliassenJob]) -> list[EliassenJob]:
    return sorted(jobs, key=lambda job: (posted_day(job), bool(job.contact_info), job.title_rank, job.title.lower()), reverse=True)


def scrape_eliassen(posted_within_days: Optional[int], exclude_disallowed_work: bool, timeout: int) -> list[EliassenJob]:
    root = ET.fromstring(fetch_feed(timeout))
    entries = root.findall("a:entry", ATOM_NS)
    print(f"Extracted {len(entries)} Eliassen feed entries before filtering")
    jobs: list[EliassenJob] = []
    for entry in entries:
        job = normalize_entry(entry)
        if not is_within_posted_days(job.posted_date, posted_within_days):
            continue
        if job.title_rank < MIN_TITLE_RANK:
            print(f"  Skipped low-rank ({job.title_rank}): {job.title}")
            continue
        title_reasons = title_exclusion_reasons(job.title)
        if title_reasons:
            print(f"  Excluded ({', '.join(title_reasons)}): {job.title}")
            continue
        role_reasons = role_exclusion_reasons(" ".join([job.title, job.raw_text]))
        if role_reasons:
            print(f"  Excluded ({', '.join(role_reasons)}): {job.title}")
            continue
        if exclude_disallowed_work:
            reasons = disallowed_work_reasons(" ".join([job.title, job.raw_text]))
            if reasons:
                print(f"  Excluded ({', '.join(reasons)}): {job.title}")
                continue
        jobs.append(job)
    return sort_jobs(jobs)


def write_csv(jobs: list[EliassenJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(EliassenJob.__dataclass_fields__.keys())
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for job in jobs:
            writer.writerow(asdict(job))


def write_json(jobs: list[EliassenJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump([asdict(job) for job in jobs], handle, indent=2)


EXCEL_COLUMNS = [
    ("posted_day", "Posted Day"),
    ("title_rank", "Title Rank"),
    ("title", "Title"),
    ("location", "Location"),
    ("salary", "Salary"),
    ("contact_info", "Contact Info"),
    ("job_id", "Job ID"),
    ("job_url", "Job URL"),
    ("title_rank_reasons", "Rank Reasons"),
]


def excel_row(job: EliassenJob) -> dict[str, Any]:
    row = asdict(job)
    row["posted_day"] = posted_day(job)
    return row


def style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F3864")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [13, 11, 42, 28, 22, 30, 36, 14, 55]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        rank = row[1].value or 0
        if isinstance(rank, int) and rank >= 45:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="E8F5E9")


def append_jobs_sheet(workbook: Workbook, sheet_name: str, jobs: list[EliassenJob]) -> None:
    sheet = workbook.create_sheet(re.sub(r"[\[\]:*?/\\]", "-", sheet_name)[:31] or "Sheet")
    sheet.append([label for _, label in EXCEL_COLUMNS])
    for job in jobs:
        source = excel_row(job)
        sheet.append([source.get(key, "") for key, _ in EXCEL_COLUMNS])
        if job.job_url:
            cell = sheet.cell(row=sheet.max_row, column=8)
            cell.value = "Open Job"
            cell.hyperlink = job.job_url
            cell.style = "Hyperlink"
    style_sheet(sheet)


def write_excel(jobs: list[EliassenJob], path: Path, posted_within_days: Optional[int]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "Eliassen Daily Jobs"
    summary["A1"].font = Font(size=18, bold=True, color="1F3864")
    summary["A3"] = "Generated"
    summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary["A4"] = "Posting Window"
    summary["B4"] = "All dates" if not posted_within_days else f"Last {posted_within_days} days"
    summary["A5"] = "Total Jobs"
    summary["B5"] = len(jobs)
    summary.append([])
    summary.append(["Day", "Jobs", "With Contact", "Top Rank"])
    for cell in summary[7]:
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.font = Font(color="FFFFFF", bold=True)
    grouped: dict[str, list[EliassenJob]] = {}
    for job in jobs:
        grouped.setdefault(posted_day(job), []).append(job)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        summary.append([day, len(day_jobs), sum(1 for job in day_jobs if job.contact_info), max((job.title_rank for job in day_jobs), default=0)])
    for col in range(1, 5):
        summary.column_dimensions[get_column_letter(col)].width = 22
    append_jobs_sheet(workbook, "All Jobs", jobs)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        append_jobs_sheet(workbook, day, day_jobs)
    workbook.save(path)


def write_daily_outputs(jobs: list[EliassenJob], out_dir: Path, timestamp: str) -> Path:
    daily_dir = out_dir / f"daily_{timestamp}"
    grouped: dict[str, list[EliassenJob]] = {}
    for job in jobs:
        grouped.setdefault(posted_day(job), []).append(job)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        write_csv(day_jobs, daily_dir / f"{day}_jobs.csv")
        write_json(day_jobs, daily_dir / f"{day}_jobs.json")
    return daily_dir


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape Eliassen jobs into CSV/JSON/Excel.")
    parser.add_argument("--posted-within-days", type=int, default=4)
    parser.add_argument("--keep-w2-f2f-onsite-interview", action="store_true")
    parser.add_argument("--timeout", type=int, default=25)
    parser.add_argument("--out-dir", type=Path, default=Path(__file__).resolve().parent / "output")
    parser.add_argument("--no-excel", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    jobs = scrape_eliassen(args.posted_within_days, not args.keep_w2_f2f_onsite_interview, args.timeout)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = args.out_dir / f"eliassen_jobs_{timestamp}.csv"
    json_path = args.out_dir / f"eliassen_jobs_{timestamp}.json"
    excel_path = args.out_dir / f"eliassen_jobs_{timestamp}.xlsx"
    write_csv(jobs, csv_path)
    write_json(jobs, json_path)
    if not args.no_excel:
        write_excel(jobs, excel_path, args.posted_within_days)
    daily_dir = write_daily_outputs(jobs, args.out_dir, timestamp)
    print(f"\nSaved {len(jobs)} jobs")
    print(f"CSV:   {csv_path}")
    print(f"JSON:  {json_path}")
    if not args.no_excel:
        print(f"Excel: {excel_path}")
    print(f"Daily: {daily_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
