#!/usr/bin/env python3
"""
Standalone Akkodis job scraper.

This is intentionally separate from the main job_checker package so we can build
the Akkodis auto-apply flow one step at a time.
"""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional
from urllib.parse import quote_plus

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


AKKODIS_SEARCH_URL = "https://www.akkodis.com/api/data/jobs/summarized"
AKKODIS_DETAIL_URL = "https://www.akkodis.com/api/data/jobs/job-description-details/{job_id}/modis/US/en/job-details"
PUBLIC_JOB_URL = "https://www.akkodis.com/en-us/careers/job/{title_location}/{job_id}"
APPLY_JOB_URL = "https://www.akkodis.com/en-us/careers/job-apply?id={job_id}"
DEFAULT_SEARCH_TERMS = [
    "ai engineer",
    "ai ml engineer",
    "aiml engineer",
    "machine learning engineer",
    "ml engineer",
    "generative ai engineer",
    "gen ai engineer",
    "llm engineer",
    "rag engineer",
    "data scientist",
    "senior data scientist",
    "applied data scientist",
    "machine learning scientist",
    "senior python developer",
    "python developer",
    "python engineer",
    "senior python engineer",
    "backend python engineer",
    "backend software engineer",
    "software engineer python",
    "software developer python",
    "django developer",
    "django engineer",
    "fastapi developer",
    "api developer",
    "rest api developer",
    "full stack software engineer",
    "full stack developer",
    "senior software engineer",
    "software engineer",
    "cloud engineer",
    "aws python developer",
    "azure python developer",
]

EMAIL_RE = re.compile(r"(?<![A-Za-z0-9._%+-])([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})(?![A-Za-z0-9._%+-])")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?1[\s.-]?)?(?:\(?\d{3}\)?[\s.-]?)\d{3}[\s.-]?\d{4}(?!\d)")
TITLE_RANKING_WEIGHTS = [
    ("principal ai architect", 60),
    ("ai lead engineer", 58),
    ("machine learning engineer", 56),
    ("ml engineer", 54),
    ("ai engineer", 54),
    ("ai/ml", 54),
    ("aiml", 54),
    ("generative ai", 52),
    ("gen ai", 52),
    ("agentic ai", 50),
    ("llm", 50),
    ("rag", 48),
    ("data scientist", 48),
    ("senior data scientist", 52),
    ("applied data scientist", 50),
    ("python developer", 46),
    ("python engineer", 46),
    ("senior python", 50),
    ("django", 44),
    ("fastapi", 44),
    ("backend", 38),
    ("back end", 38),
    ("api developer", 36),
    ("software engineer python", 46),
    ("full stack software engineer", 42),
    ("full stack engineer", 40),
    ("full stack developer", 40),
    ("data engineer", 38),
    ("etl", 32),
    ("cloud engineer", 30),
    ("aws", 28),
    ("azure", 28),
    ("senior software engineer", 32),
    ("software engineer", 24),
    ("software developer", 24),
    ("developer", 12),
    ("engineer", 10),
]
TITLE_EXCLUSION_WEIGHTS = [
    ("frontend", -30),
    ("front end", -30),
    ("ui developer", -20),
    ("ux/ui", -12),
    ("java", -12),
    (".net", -35),
    ("qa", -30),
    ("quality assurance", -30),
    ("desktop support", -40),
    ("help desk", -40),
    ("project manager", -35),
    ("business analyst", -30),
]
DISALLOWED_WORK_PATTERNS = [
    (re.compile(r"\bno\s+c2c\b", re.I), "No C2C"),
    (re.compile(r"\bno\s+corp(?:oration)?\s*[- ]?\s*to\s*[- ]?\s*corp(?:oration)?\b", re.I), "No corp-to-corp"),
    (re.compile(r"\bf\s*2\s*f\b", re.I), "F2F"),
    (re.compile(r"\bface\s*[- ]?\s*to\s*[- ]?\s*face\b", re.I), "Face-to-face"),
    (re.compile(r"\bin\s*[- ]?\s*person\s+interview\b", re.I), "In-person interview"),
    (re.compile(r"\bon\s*[- ]?\s*site\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\bonsite\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\bmust\s+interview\s+on\s*[- ]?\s*site\b", re.I), "Must interview onsite"),
    (re.compile(r"\blocal\s+candidates?\s+only\b", re.I), "Local candidates only"),
    (re.compile(r"\bfull\s*[- ]?\s*time\b", re.I), "Full-time"),
    (re.compile(r"\bpermanent\b", re.I), "Permanent"),
    (re.compile(r"\bdirect\s*[- ]?\s*hire\b", re.I), "Direct hire"),
]


@dataclass
class AkkodisJob:
    source_company: str
    search_term: str
    title_rank: int
    title_rank_reasons: str
    title: str
    location: str
    employment_type: str
    remote_status: str
    min_pay_rate: str
    max_pay_rate: str
    pay_rate_unit: str
    posted_date: str
    job_id: str
    external_reference: str
    job_url: str
    contact_info: str
    description_snippet: str
    raw_text: str


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", html.unescape(text or "")).strip()


def normalized_title(text: str) -> str:
    return re.sub(r"[^a-z0-9+#./-]+", " ", (text or "").lower()).strip()


def title_contains(title: str, phrase: str) -> bool:
    phrase = normalized_title(phrase)
    if not phrase:
        return False
    if len(phrase) <= 3:
        return re.search(rf"(?<![a-z0-9]){re.escape(phrase)}(?![a-z0-9])", title) is not None
    return phrase in title


def score_title(title: str) -> tuple[int, str]:
    normalized = normalized_title(title)
    score = 0
    reasons: List[str] = []
    for phrase, weight in TITLE_RANKING_WEIGHTS:
        if title_contains(normalized, phrase):
            score += weight
            reasons.append(f"{phrase}+{weight}")
    for phrase, weight in TITLE_EXCLUSION_WEIGHTS:
        if title_contains(normalized, phrase):
            score += weight
            reasons.append(f"{phrase}{weight}")
    return max(score, 0), "; ".join(reasons)


def disallowed_work_reasons(text: str) -> List[str]:
    cleaned = clean_text(text)
    reasons: List[str] = []
    for pattern, reason in DISALLOWED_WORK_PATTERNS:
        if pattern.search(cleaned) and reason not in reasons:
            reasons.append(reason)
    return reasons


def is_disallowed_work_job(job: "AkkodisJob") -> bool:
    text = " ".join([job.title, job.location, job.employment_type, job.description_snippet, job.raw_text])
    return bool(disallowed_work_reasons(text))


def parse_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[^0-9.]+", "", str(value))
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def pay_rate_from_row(row: Dict[str, Any], detail: Optional[Dict[str, Any]] = None) -> tuple[Optional[float], Optional[float], str]:
    detail = detail or {}
    min_rate = parse_number(detail.get("minsalary") or detail.get("minSalary") or row.get("minsalary") or row.get("minSalary"))
    max_rate = parse_number(detail.get("maxSalary") or detail.get("maxsalary") or row.get("maxsalary") or row.get("maxSalary"))
    unit = clean_text(
        str(
            detail.get("salaryTimeScale")
            or detail.get("salaryTimeScaleID")
            or row.get("salaryTimeScale")
            or row.get("salaryTimeScaleID")
            or ""
        )
    )
    return min_rate, max_rate, unit


def is_hourly_pay_unit(unit: str) -> bool:
    normalized = normalized_title(unit)
    return normalized in {"hour", "hourly", "perhour", "per-hour", "per hour", "hr"}


def is_below_min_hourly_pay(row: Dict[str, Any], detail: Optional[Dict[str, Any]], min_hourly_rate: Optional[float]) -> bool:
    if not min_hourly_rate or min_hourly_rate <= 0:
        return False
    min_rate, max_rate, unit = pay_rate_from_row(row, detail)
    if not is_hourly_pay_unit(unit):
        return False
    best_rate = max(rate for rate in (min_rate, max_rate) if rate is not None) if any(rate is not None for rate in (min_rate, max_rate)) else None
    return best_rate is not None and best_rate < min_hourly_rate


def html_to_text(value: str) -> str:
    return clean_text(re.sub("<[^>]+>", " ", value or ""))


def slugify(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9 ]+", "", value or "")
    value = re.sub(r"\s+", "-", value.strip().lower())
    return value.strip("-")


def contact_info(text: str) -> str:
    emails: List[str] = []
    for email in EMAIL_RE.findall(clean_text(text)):
        lowered = email.lower()
        if lowered not in emails:
            emails.append(lowered)

    phones: List[str] = []
    for phone in PHONE_RE.findall(text or ""):
        digits = re.sub(r"\D", "", phone)
        if len(digits) == 11 and digits.startswith("1"):
            digits = digits[1:]
        if len(digits) != 10:
            continue
        display = f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
        if display not in phones:
            phones.append(display)
    return ", ".join(emails + phones)


def parse_akkodis_date(value: str) -> Optional[datetime]:
    value = clean_text(value)
    if not value:
        return None
    if value.endswith("Z"):
        value = value[:-1] + "+00:00"
    try:
        parsed = datetime.fromisoformat(value)
    except ValueError:
        return None
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def row_posted_date(row: Dict[str, Any]) -> Optional[datetime]:
    for field in ("postedDate", "jobCreationDate", "firstPostedDate"):
        parsed = parse_akkodis_date(str(row.get(field) or ""))
        if parsed:
            return parsed
    return None


def is_within_posted_days(row: Dict[str, Any], days: Optional[int], now: Optional[datetime] = None) -> bool:
    if not days or days <= 0:
        return True
    posted = row_posted_date(row)
    if not posted:
        return False
    now = now or datetime.now(timezone.utc)
    return 0 <= (now - posted).total_seconds() <= days * 24 * 60 * 60


def public_job_url(row: Dict[str, Any]) -> str:
    job_id = clean_text(str(row.get("jobId") or ""))
    return APPLY_JOB_URL.format(job_id=job_id)


def search_payload(term: str) -> Dict[str, Any]:
    return {
        "baseSearchQuery": "",
        "filtersToDisplay": "",
        "selectedFilters": "",
        "queryString": f"&q={term}",
        "range": 0,
        "siteName": "akkodis",
        "brand": "modis",
        "brandFromDictionary": "",
        "countryCookie": "US",
        "langCookie": "en",
    }


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update(
        {
            "User-Agent": "Mozilla/5.0 VenkataDoraAkkodisAutomation/0.1 (+local personal job search)",
            "Accept": "application/json,text/plain,*/*",
            "Origin": "https://www.akkodis.com",
        }
    )
    return session


def search_jobs(session: requests.Session, term: str, timeout: int) -> List[Dict[str, Any]]:
    response = session.post(
        AKKODIS_SEARCH_URL,
        json=search_payload(term),
        headers={"Referer": f"https://www.akkodis.com/en-us/careers/job-results?k={quote_plus(term)}"},
        timeout=timeout,
    )
    response.raise_for_status()
    payload = response.json()
    jobs = payload.get("jobs", []) if isinstance(payload, dict) else []
    return [job for job in jobs if isinstance(job, dict)]


def fetch_detail(session: requests.Session, job_id: str, timeout: int) -> Optional[Dict[str, Any]]:
    if not job_id:
        return None
    response = session.get(AKKODIS_DETAIL_URL.format(job_id=quote_plus(job_id)), timeout=timeout)
    response.raise_for_status()
    payload = response.json()
    return payload if isinstance(payload, dict) else None


def normalize_job(row: Dict[str, Any], search_term: str, detail: Optional[Dict[str, Any]]) -> AkkodisJob:
    detail = detail or {}
    description = html_to_text(str(detail.get("jobDescription") or row.get("description") or row.get("clientJobDescription") or ""))
    title = clean_text(str(detail.get("jobName") or row.get("jobTitle") or "Untitled job"))
    location = clean_text(str(detail.get("location") or row.get("jobLocation") or ""))
    employment_type = clean_text(str(detail.get("contract") or row.get("contractTypeTitle") or row.get("jobType") or ""))
    is_remote = bool(detail.get("isRemote") or row.get("isRemote"))
    title_rank, title_rank_reasons = score_title(title)
    min_rate, max_rate, pay_unit = pay_rate_from_row(row, detail)
    return AkkodisJob(
        source_company="Akkodis",
        search_term=search_term,
        title_rank=title_rank,
        title_rank_reasons=title_rank_reasons,
        title=title,
        location=location,
        employment_type=employment_type,
        remote_status="Remote" if is_remote else "",
        min_pay_rate="" if min_rate is None else f"{min_rate:g}",
        max_pay_rate="" if max_rate is None else f"{max_rate:g}",
        pay_rate_unit=pay_unit,
        posted_date=clean_text(str(detail.get("jobCreatedDate") or row.get("postedDate") or row.get("jobCreationDate") or "")),
        job_id=clean_text(str(row.get("jobId") or detail.get("jobId") or "")),
        external_reference=clean_text(str(row.get("externalReference") or detail.get("externalReference") or "")),
        job_url=public_job_url(row),
        contact_info=contact_info(description),
        description_snippet=description[:900],
        raw_text=description,
    )


def job_posted_day(job: AkkodisJob) -> str:
    parsed = parse_akkodis_date(job.posted_date)
    return parsed.date().isoformat() if parsed else "unknown-date"


def sort_jobs(jobs: Iterable[AkkodisJob]) -> List[AkkodisJob]:
    def key(job: AkkodisJob) -> tuple[str, bool, int, str]:
        return (job_posted_day(job), bool(job.contact_info), job.title_rank, job.title.lower())

    return sorted(jobs, key=key, reverse=True)


def scrape_akkodis(
    search_terms: Iterable[str],
    max_detail_pages: int,
    timeout: int,
    sleep_seconds: float,
    posted_within_days: Optional[int],
    exclude_disallowed_work: bool,
    min_hourly_rate: Optional[float],
) -> List[AkkodisJob]:
    session = make_session()
    seen_job_ids = set()
    rows: List[tuple[str, Dict[str, Any]]] = []

    for term in search_terms:
        term = term.strip()
        if not term:
            continue
        print(f"Searching Akkodis: {term}")
        for row in search_jobs(session, term, timeout):
            if not is_within_posted_days(row, posted_within_days):
                continue
            job_id = str(row.get("jobId") or "")
            if not job_id or job_id in seen_job_ids:
                continue
            seen_job_ids.add(job_id)
            rows.append((term, row))
        time.sleep(sleep_seconds)

    jobs: List[AkkodisJob] = []
    for index, (term, row) in enumerate(rows):
        detail = None
        should_fetch_detail = exclude_disallowed_work or index < max_detail_pages
        if should_fetch_detail:
            try:
                detail = fetch_detail(session, str(row.get("jobId") or ""), timeout)
            except requests.RequestException as exc:
                print(f"Detail failed for {row.get('jobId')}: {exc}")
            time.sleep(sleep_seconds)
        job = normalize_job(row, term, detail)
        if exclude_disallowed_work and is_disallowed_work_job(job):
            print(f"Excluded {job.job_id}: {', '.join(disallowed_work_reasons(job.raw_text)) or 'disallowed work signal'}")
            continue
        if is_below_min_hourly_pay(row, detail, min_hourly_rate):
            _, max_rate, unit = pay_rate_from_row(row, detail)
            print(f"Excluded {job.job_id}: below ${min_hourly_rate:g}/hour pay threshold (max {max_rate:g} {unit})")
            continue
        jobs.append(job)
    return jobs


def write_csv(jobs: List[AkkodisJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(asdict(jobs[0]).keys()) if jobs else list(AkkodisJob.__dataclass_fields__.keys()))
        writer.writeheader()
        for job in jobs:
            writer.writerow(asdict(job))


def write_json(jobs: List[AkkodisJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        json.dump([asdict(job) for job in jobs], handle, indent=2)


def write_daily_outputs(jobs: List[AkkodisJob], out_dir: Path, timestamp: str) -> Path:
    daily_dir = out_dir / f"daily_{timestamp}"
    grouped: Dict[str, List[AkkodisJob]] = {}
    for job in jobs:
        grouped.setdefault(job_posted_day(job), []).append(job)

    for day, day_jobs in sorted(grouped.items(), reverse=True):
        ranked = sorted(day_jobs, key=lambda job: (not bool(job.contact_info), -job.title_rank, job.title.lower(), job.location.lower()))
        write_csv(ranked, daily_dir / f"{day}_jobs.csv")
        write_json(ranked, daily_dir / f"{day}_jobs.json")
    return daily_dir


EXCEL_COLUMNS = [
    ("posted_day", "Posted Day"),
    ("title_rank", "Title Rank"),
    ("title", "Title"),
    ("location", "Location"),
    ("employment_type", "Employment"),
    ("pay_range", "Pay"),
    ("contact_info", "Contact Info"),
    ("search_term", "Search Term"),
    ("job_id", "Job ID"),
    ("job_url", "Job URL"),
    ("title_rank_reasons", "Rank Reasons"),
]


def excel_safe_sheet_name(value: str) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", "-", value or "Sheet")
    return cleaned[:31] or "Sheet"


def excel_row(job: AkkodisJob) -> Dict[str, Any]:
    pay_range = ""
    if job.min_pay_rate or job.max_pay_rate:
        if job.min_pay_rate and job.max_pay_rate and job.min_pay_rate != job.max_pay_rate:
            pay_range = f"${job.min_pay_rate}-${job.max_pay_rate}"
        else:
            pay_range = f"${job.max_pay_rate or job.min_pay_rate}"
        if job.pay_rate_unit:
            pay_range = f"{pay_range} / {job.pay_rate_unit}"
    row = asdict(job)
    row["posted_day"] = job_posted_day(job)
    row["pay_range"] = pay_range
    return row


def style_job_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="0B2D3D")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = {
        "A": 13,
        "B": 11,
        "C": 34,
        "D": 24,
        "E": 16,
        "F": 16,
        "G": 30,
        "H": 24,
        "I": 24,
        "J": 14,
        "K": 45,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        rank = row[1].value or 0
        if isinstance(rank, int) and rank >= 45:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="E8F5E9")
        if row[6].value:
            row[6].fill = PatternFill("solid", fgColor="FFF3CD")
    for row_index in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = 42


def append_jobs_sheet(workbook: Workbook, sheet_name: str, jobs: List[AkkodisJob]) -> None:
    sheet = workbook.create_sheet(excel_safe_sheet_name(sheet_name))
    headers = [label for _, label in EXCEL_COLUMNS]
    sheet.append(headers)
    for job in jobs:
        source = excel_row(job)
        values = [source.get(key, "") for key, _ in EXCEL_COLUMNS]
        sheet.append(values)
        link_cell = sheet.cell(row=sheet.max_row, column=10)
        if job.job_url:
            link_cell.value = "Open Job"
            link_cell.hyperlink = job.job_url
            link_cell.style = "Hyperlink"
    style_job_sheet(sheet)


def write_excel(jobs: List[AkkodisJob], path: Path, posted_within_days: Optional[int], min_hourly_rate: Optional[float]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary["A1"] = "Akkodis Daily Jobs"
    summary["A1"].font = Font(size=18, bold=True, color="0B2D3D")
    summary["A3"] = "Generated"
    summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary["A4"] = "Posting Window"
    summary["B4"] = "All dates" if not posted_within_days else f"Last {posted_within_days} days"
    summary["A5"] = "Minimum Hourly Pay"
    summary["B5"] = "Disabled" if not min_hourly_rate else f"${min_hourly_rate:g}/hour"
    summary["A6"] = "Total Jobs"
    summary["B6"] = len(jobs)
    summary["A8"] = "Day"
    summary["B8"] = "Jobs"
    summary["C8"] = "With Contact"
    summary["D8"] = "Top Rank"
    for cell in summary[8]:
        cell.fill = PatternFill("solid", fgColor="0B2D3D")
        cell.font = Font(color="FFFFFF", bold=True)

    grouped: Dict[str, List[AkkodisJob]] = {}
    for job in jobs:
        grouped.setdefault(job_posted_day(job), []).append(job)
    for row_index, (day, day_jobs) in enumerate(sorted(grouped.items(), reverse=True), start=9):
        summary.cell(row=row_index, column=1, value=day)
        summary.cell(row=row_index, column=2, value=len(day_jobs))
        summary.cell(row=row_index, column=3, value=sum(1 for job in day_jobs if job.contact_info))
        summary.cell(row=row_index, column=4, value=max((job.title_rank for job in day_jobs), default=0))

    for column in range(1, 5):
        summary.column_dimensions[get_column_letter(column)].width = 22
    summary.freeze_panes = "A8"

    append_jobs_sheet(workbook, "All Jobs", jobs)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        ranked = sorted(day_jobs, key=lambda job: (not bool(job.contact_info), -job.title_rank, job.title.lower(), job.location.lower()))
        append_jobs_sheet(workbook, day, ranked)

    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape Akkodis jobs into CSV/JSON.")
    parser.add_argument("--term", action="append", dest="terms", help="Search term. Repeat this flag for multiple terms.")
    parser.add_argument("--terms-file", type=Path, help="Optional text file with one search term per line.")
    parser.add_argument("--max-detail-pages", type=int, default=30, help="How many job detail pages to fetch.")
    parser.add_argument(
        "--posted-within-days",
        type=int,
        default=4,
        help="Only keep jobs posted in the last N days. Use 0 to disable this filter.",
    )
    parser.add_argument(
        "--keep-w2-f2f-onsite-interview",
        action="store_true",
        help="Disable the default filter that removes F2F, face-to-face, onsite-interview, full-time, permanent, and direct-hire jobs.",
    )
    parser.add_argument(
        "--min-hourly-rate",
        type=float,
        default=55,
        help="Remove hourly jobs whose max pay rate is below this amount. Use 0 to disable.",
    )
    parser.add_argument("--timeout", type=int, default=20, help="HTTP timeout in seconds.")
    parser.add_argument("--sleep", type=float, default=0.5, help="Seconds to sleep between requests.")
    parser.add_argument("--out-dir", type=Path, default=Path(__file__).resolve().parent / "output", help="Output directory.")
    parser.add_argument("--no-excel", action="store_true", help="Skip writing the Excel workbook.")
    return parser.parse_args()


def load_terms(args: argparse.Namespace) -> List[str]:
    terms = list(args.terms or [])
    if args.terms_file:
        terms.extend(line.strip() for line in args.terms_file.read_text(encoding="utf-8").splitlines())
    return [term for term in terms if term.strip()] or DEFAULT_SEARCH_TERMS


def main() -> int:
    args = parse_args()
    terms = load_terms(args)
    jobs = scrape_akkodis(
        terms,
        args.max_detail_pages,
        args.timeout,
        args.sleep,
        args.posted_within_days,
        not args.keep_w2_f2f_onsite_interview,
        args.min_hourly_rate,
    )
    jobs = sort_jobs(jobs)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = args.out_dir / f"akkodis_jobs_{timestamp}.csv"
    json_path = args.out_dir / f"akkodis_jobs_{timestamp}.json"
    excel_path = args.out_dir / f"akkodis_jobs_{timestamp}.xlsx"
    write_csv(jobs, csv_path)
    write_json(jobs, json_path)
    if not args.no_excel:
        write_excel(jobs, excel_path, args.posted_within_days, args.min_hourly_rate)
    daily_dir = write_daily_outputs(jobs, args.out_dir, timestamp)
    print(f"Saved {len(jobs)} jobs")
    print(f"CSV: {csv_path}")
    print(f"JSON: {json_path}")
    if not args.no_excel:
        print(f"Excel: {excel_path}")
    print(f"Daily ranked files: {daily_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
