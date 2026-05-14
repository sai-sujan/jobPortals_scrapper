#!/usr/bin/env python3
"""Randstad USA job scraper — parses window.__ROUTE_DATA__ from search result pages."""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable, Optional

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = "https://www.randstadusa.com"
# Only slugs that Randstad serves (others return HTTP 410)
DEFAULT_SEARCH_TERMS = [
    "machine learning engineer",   # /jobs/q-machine-learning-engineer/
    "data scientist",              # /jobs/q-data-scientist/
    "python developer",            # /jobs/q-python-developer/
    "software engineer",           # /jobs/q-software-engineer/
    "software developer",          # /jobs/q-software-developer/
    "full stack developer",        # /jobs/q-full-stack-developer/
    "cloud engineer",              # /jobs/q-cloud-engineer/
    "devops engineer",             # /jobs/q-devops-engineer/
]

TITLE_RANKING_WEIGHTS = [
    ("machine learning engineer", 56),
    ("ml engineer", 54),
    ("ai engineer", 54),
    ("ai/ml", 54),
    ("aiml", 54),
    ("generative ai", 52),
    ("gen ai", 52),
    ("llm", 50),
    ("rag", 48),
    ("senior data scientist", 52),
    ("data scientist", 48),
    ("python developer", 46),
    ("python engineer", 46),
    ("senior python", 50),
    ("django", 44),
    ("fastapi", 44),
    ("backend", 38),
    ("back end", 38),
    ("api developer", 36),
    ("software engineer python", 46),
    ("full stack software engineer", 42),
    ("full stack engineer", 40),
    ("full stack developer", 40),
    ("data engineer", 38),
    ("etl", 32),
    ("cloud engineer", 30),
    ("aws", 28),
    ("azure", 28),
    ("senior software engineer", 32),
    ("software engineer", 24),
    ("software developer", 24),
    ("developer", 12),
    ("engineer", 10),
]
TITLE_EXCLUSION_WEIGHTS = [
    ("data engineer", -100),
    ("frontend", -30),
    ("front end", -30),
    ("ui developer", -20),
    ("ux/ui", -12),
    ("java", -12),
    (".net", -35),
    ("qa", -30),
    ("quality assurance", -30),
    ("desktop support", -40),
    ("help desk", -40),
    ("project manager", -35),
    ("business analyst", -30),
    ("junior", -100),
    ("jr ", -100),
    ("entry level", -100),
]
TITLE_EXCLUSION_PATTERNS = [
    (re.compile(r"\bjava\b.*\bfull\s*stack\b|\bfull\s*stack\b.*\bjava\b|\bjava\b.*\bfullstack\b|\bfullstack\b.*\bjava\b", re.I), "Java full stack title"),
    (re.compile(r"\bjava\b.*\b(?:developer|engineer|architect|backend|software)\b|\b(?:developer|engineer|architect|backend|software)\b.*\bjava\b", re.I), "Java developer/engineer title"),
    (re.compile(r"\bdata\s+engineer\b", re.I), "Data Engineer title"),
    (re.compile(r"\bjunior\b|\bjr\.?\s", re.I), "Junior title"),
    (re.compile(r"\bentry[\s-]level\b", re.I), "Entry-level title"),
]
DISALLOWED_WORK_PATTERNS = [
    (re.compile(r"\bno\s+c2c\b", re.I), "No C2C"),
    (re.compile(r"\bno\s+corp(?:oration)?\s*[- ]?\s*to\s*[- ]?\s*corp(?:oration)?\b", re.I), "No corp-to-corp"),
    (re.compile(r"\bf\s*2\s*f\b", re.I), "F2F"),
    (re.compile(r"\bface\s*[- ]?\s*to\s*[- ]?\s*face\b", re.I), "Face-to-face"),
    (re.compile(r"\bin\s*[- ]?\s*person\s+interview\b", re.I), "In-person interview"),
    (re.compile(r"\bon\s*[- ]?\s*site\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\bonsite\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\bfull\s*[- ]?\s*time\b", re.I), "Full-time"),
    (re.compile(r"\bpermanent\b", re.I), "Permanent"),
    (re.compile(r"\bdirect\s*[- ]?\s*hire\b", re.I), "Direct hire"),
    (re.compile(r"\bcontract\s*[- ]?\s*to\s*[- ]?\s*hire\b", re.I), "Contract-to-Hire"),
    (re.compile(r"\bcth\b", re.I), "Contract-to-Hire (CTH)"),
]
EMAIL_RE = re.compile(r"(?<![A-Za-z0-9._%+-])([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})(?![A-Za-z0-9._%+-])")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?1[\s.-]?)?(?:\(?\d{3}\)?[\s.-]?)\d{3}[\s.-]?\d{4}(?!\d)")
MIN_TITLE_RANK = 20


@dataclass
class RandstadJob:
    source_company: str
    search_term: str
    title_rank: int
    title_rank_reasons: str
    title: str
    city: str
    state: str
    location: str
    employment_type: str
    is_remote: str
    lob_name: str
    min_pay_rate: str
    max_pay_rate: str
    pay_rate_unit: str
    posted_date: str
    job_id: str
    job_url: str
    apply_url: str
    contact_info: str
    description_snippet: str
    raw_text: str


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", html.unescape(text or "")).strip()


def normalized_title(text: str) -> str:
    return re.sub(r"[^a-z0-9+#./-]+", " ", (text or "").lower()).strip()


def title_contains(title: str, phrase: str) -> bool:
    phrase = normalized_title(phrase)
    if len(phrase) <= 3:
        return re.search(rf"(?<![a-z0-9]){re.escape(phrase)}(?![a-z0-9])", title) is not None
    return phrase in title


def score_title(title: str) -> tuple[int, str]:
    normalized = normalized_title(title)
    score = 0
    reasons: list[str] = []
    for phrase, weight in TITLE_RANKING_WEIGHTS:
        if title_contains(normalized, phrase):
            score += weight
            reasons.append(f"{phrase}+{weight}")
    for phrase, weight in TITLE_EXCLUSION_WEIGHTS:
        if title_contains(normalized, phrase):
            score += weight
            reasons.append(f"{phrase}{weight}")
    return max(score, 0), "; ".join(reasons)


def title_exclusion_reasons(title: str) -> list[str]:
    return [reason for pattern, reason in TITLE_EXCLUSION_PATTERNS if pattern.search(title or "")]


def disallowed_work_reasons(text: str) -> list[str]:
    reasons: list[str] = []
    cleaned = clean_text(text)
    for pattern, reason in DISALLOWED_WORK_PATTERNS:
        if pattern.search(cleaned) and reason not in reasons:
            reasons.append(reason)
    return reasons


def extract_contact_info(text: str) -> str:
    emails = []
    for email in EMAIL_RE.findall(clean_text(text)):
        lowered = email.lower()
        if lowered not in emails:
            emails.append(lowered)
    phones = []
    for phone in PHONE_RE.findall(text or ""):
        digits = re.sub(r"\D", "", phone)
        if len(digits) == 11 and digits.startswith("1"):
            digits = digits[1:]
        if len(digits) == 10:
            display = f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
            if display not in phones:
                phones.append(display)
    return ", ".join(emails + phones)


def ms_epoch_to_iso(ms: Any) -> str:
    if not ms:
        return ""
    try:
        return datetime.fromtimestamp(int(ms) / 1000, tz=timezone.utc).date().isoformat()
    except (ValueError, OSError):
        return ""


def is_within_posted_days(posted_iso: str, days: Optional[int]) -> bool:
    if not days or days <= 0 or not posted_iso:
        return True
    try:
        posted_dt = datetime.fromisoformat(posted_iso).replace(tzinfo=timezone.utc)
    except ValueError:
        return True
    now = datetime.now(timezone.utc)
    return 0 <= (now - posted_dt).total_seconds() <= days * 86400


def term_to_slug(term: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", term.lower().strip()).strip("-")


def search_url(term: str) -> str:
    return f"{BASE_URL}/jobs/q-{term_to_slug(term)}/"


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
    })
    return session


def extract_route_data(html_text: str) -> Optional[dict]:
    start_marker = "window.__ROUTE_DATA__ = "
    start_idx = html_text.find(start_marker)
    if start_idx == -1:
        return None
    json_start = start_idx + len(start_marker)
    depth = 0
    in_string = False
    escape_next = False
    i = json_start
    while i < len(html_text):
        c = html_text[i]
        if escape_next:
            escape_next = False
        elif c == "\\" and in_string:
            escape_next = True
        elif c == '"' and not escape_next:
            in_string = not in_string
        elif not in_string:
            if c == "{":
                depth += 1
            elif c == "}":
                depth -= 1
                if depth == 0:
                    json_end = i + 1
                    break
        i += 1
    else:
        return None
    try:
        return json.loads(html_text[json_start:json_end])
    except json.JSONDecodeError:
        return None


def parse_hits(data: dict, term: str) -> list[dict]:
    hits = data.get("searchResults", {}).get("hits", [])
    results = []
    for hit in hits:
        loc = hit.get("jobLocation") or {}
        salary = hit.get("salary") or {}
        city = clean_text(loc.get("city") or "")
        state = clean_text(loc.get("state") or "")
        state_abbr = clean_text(loc.get("stateAbbreviation") or "")
        location = f"{city}, {state_abbr}" if city and state_abbr else city or state
        posted = ms_epoch_to_iso(hit.get("createdDate"))
        emp_type = clean_text(hit.get("type") or hit.get("employmentType") or "")
        pay_min = salary.get("min")
        pay_max = salary.get("max")
        pay_type = clean_text(salary.get("type") or "")
        pay_unit = ""
        if "hour" in pay_type.lower():
            pay_unit = "hour"
        elif "year" in pay_type.lower():
            pay_unit = "year"
        description = clean_text(hit.get("description") or "")
        results.append({
            "search_term": term,
            "job_id": str(hit.get("id") or hit.get("atsReference") or ""),
            "title": clean_text(hit.get("title") or ""),
            "city": city,
            "state": state,
            "location": location,
            "employment_type": emp_type,
            "is_remote": "Yes" if hit.get("isRemote") else "No",
            "lob_name": clean_text(hit.get("lobName") or hit.get("opCoName") or ""),
            "min_pay_rate": "" if pay_min is None else f"{pay_min:g}",
            "max_pay_rate": "" if pay_max is None else f"{pay_max:g}",
            "pay_rate_unit": pay_unit,
            "posted_date": posted,
            "job_url": clean_text(hit.get("detailsUrl") or ""),
            "apply_url": clean_text(hit.get("applyUrl") or ""),
            "raw_text": description,
            "description_snippet": description[:900],
        })
    return results


def normalize_job(row: dict) -> RandstadJob:
    title_rank, title_rank_reasons = score_title(row["title"])
    return RandstadJob(
        source_company="Randstad USA",
        search_term=row["search_term"],
        title_rank=title_rank,
        title_rank_reasons=title_rank_reasons,
        title=row["title"],
        city=row["city"],
        state=row["state"],
        location=row["location"],
        employment_type=row["employment_type"],
        is_remote=row["is_remote"],
        lob_name=row["lob_name"],
        min_pay_rate=row["min_pay_rate"],
        max_pay_rate=row["max_pay_rate"],
        pay_rate_unit=row["pay_rate_unit"],
        posted_date=row["posted_date"],
        job_id=row["job_id"],
        job_url=row["job_url"],
        apply_url=row["apply_url"],
        contact_info=extract_contact_info(row["raw_text"]),
        description_snippet=row["description_snippet"],
        raw_text=row["raw_text"],
    )


def below_min_hourly(job: RandstadJob, min_hourly_rate: Optional[float]) -> bool:
    if not min_hourly_rate or min_hourly_rate <= 0 or job.pay_rate_unit != "hour":
        return False
    rates = [float(r) for r in (job.min_pay_rate, job.max_pay_rate) if r]
    return bool(rates) and max(rates) < min_hourly_rate


def sort_jobs(jobs: Iterable[RandstadJob]) -> list[RandstadJob]:
    return sorted(
        jobs,
        key=lambda j: (j.posted_date, bool(j.contact_info), j.title_rank, j.title.lower()),
        reverse=True,
    )


def scrape_randstad(
    search_terms: Iterable[str],
    posted_within_days: Optional[int],
    exclude_disallowed_work: bool,
    min_hourly_rate: Optional[float],
    timeout: int,
    sleep_seconds: float,
) -> list[RandstadJob]:
    session = make_session()
    seen: set[str] = set()
    jobs: list[RandstadJob] = []

    for term in search_terms:
        term = term.strip()
        if not term:
            continue
        url = search_url(term)
        print(f"Searching Randstad: {term} -> {url}")
        try:
            response = session.get(url, timeout=timeout)
            response.raise_for_status()
        except requests.HTTPError as exc:
            if exc.response is not None and exc.response.status_code == 410:
                print(f"  Skipped (slug not served by Randstad): {url}")
            else:
                print(f"  Request failed: {exc}")
        except requests.RequestException as exc:
            print(f"  Request failed: {exc}")
            time.sleep(sleep_seconds)
            continue

        data = extract_route_data(response.text)
        if not data:
            print("  No ROUTE_DATA found")
            time.sleep(sleep_seconds)
            continue

        total = data.get("searchResults", {}).get("totalSize", 0)
        rows = parse_hits(data, term)
        print(f"  Found {total} total, {len(rows)} on page")

        for row in rows:
            key = row["job_id"] or row["job_url"]
            if key in seen:
                continue
            seen.add(key)

            if not is_within_posted_days(row["posted_date"], posted_within_days):
                continue

            job = normalize_job(row)

            if job.title_rank < MIN_TITLE_RANK:
                print(f"  Skipped low-rank ({job.title_rank}): {job.title}")
                continue

            excl = title_exclusion_reasons(job.title)
            if excl:
                print(f"  Excluded ({', '.join(excl)}): {job.title}")
                continue

            if exclude_disallowed_work:
                reasons = disallowed_work_reasons(" ".join([job.title, job.employment_type, job.raw_text]))
                if reasons:
                    print(f"  Excluded ({', '.join(reasons)}): {job.title}")
                    continue

            if below_min_hourly(job, min_hourly_rate):
                print(f"  Excluded (below ${min_hourly_rate:g}/hr): {job.title}")
                continue

            if job.pay_rate_unit == "year":
                print(f"  Excluded (annual salary = full-time signal): {job.title}")
                continue

            jobs.append(job)

        time.sleep(sleep_seconds)

    return sort_jobs(jobs)


# ── output helpers ────────────────────────────────────────────────────────────

def write_csv(jobs: list[RandstadJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(RandstadJob.__dataclass_fields__.keys())
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        for job in jobs:
            writer.writerow(asdict(job))


def write_json(jobs: list[RandstadJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        json.dump([asdict(job) for job in jobs], fh, indent=2)


EXCEL_COLUMNS = [
    ("posted_date", "Posted"),
    ("title_rank", "Rank"),
    ("title", "Title"),
    ("location", "Location"),
    ("employment_type", "Type"),
    ("is_remote", "Remote"),
    ("lob_name", "LOB"),
    ("pay_range", "Pay"),
    ("contact_info", "Contact"),
    ("search_term", "Search Term"),
    ("job_id", "Job ID"),
    ("job_url", "Job URL"),
    ("apply_url", "Apply URL"),
    ("title_rank_reasons", "Rank Reasons"),
]


def excel_row(job: RandstadJob) -> dict[str, Any]:
    row = asdict(job)
    row["pay_range"] = ""
    if job.min_pay_rate or job.max_pay_rate:
        lo = job.min_pay_rate or job.max_pay_rate
        hi = job.max_pay_rate or job.min_pay_rate
        unit = f" / {job.pay_rate_unit}" if job.pay_rate_unit else ""
        row["pay_range"] = f"${lo}-${hi}{unit}"
    return row


def style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="1F3864")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [13, 8, 36, 22, 14, 8, 18, 18, 30, 24, 16, 14, 14, 45]
    for idx, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = w
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        rank = row[1].value or 0
        if isinstance(rank, int) and rank >= 45:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="E8F5E9")
        if row[8].value:
            row[8].fill = PatternFill("solid", fgColor="FFF3CD")


def append_jobs_sheet(workbook: Workbook, sheet_name: str, jobs: list[RandstadJob]) -> None:
    safe_name = re.sub(r"[\[\]:*?/\\]", "-", sheet_name)[:31] or "Sheet"
    sheet = workbook.create_sheet(safe_name)
    sheet.append([label for _, label in EXCEL_COLUMNS])
    for job in jobs:
        source = excel_row(job)
        sheet.append([source.get(key, "") for key, _ in EXCEL_COLUMNS])
        for col_idx, (key, _) in enumerate(EXCEL_COLUMNS, start=1):
            if key in ("job_url", "apply_url"):
                cell = sheet.cell(row=sheet.max_row, column=col_idx)
                url = job.job_url if key == "job_url" else job.apply_url
                if url:
                    cell.value = "Open Job" if key == "job_url" else "Apply"
                    cell.hyperlink = url
                    cell.style = "Hyperlink"
    style_sheet(sheet)


def write_excel(jobs: list[RandstadJob], path: Path, posted_within_days: Optional[int], min_hourly_rate: Optional[float]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Randstad USA Daily Jobs"
    summary["A1"].font = Font(size=18, bold=True, color="1F3864")
    summary["A3"] = "Generated"
    summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary["A4"] = "Posting Window"
    summary["B4"] = "All dates" if not posted_within_days else f"Last {posted_within_days} days"
    summary["A5"] = "Min Hourly Pay"
    summary["B5"] = "Disabled" if not min_hourly_rate else f"${min_hourly_rate:g}/hour"
    summary["A6"] = "Total Jobs"
    summary["B6"] = len(jobs)
    summary.append([])
    summary.append(["Day", "Jobs", "With Contact", "Top Rank"])
    for cell in summary[8]:
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.font = Font(color="FFFFFF", bold=True)
    grouped: dict[str, list[RandstadJob]] = {}
    for job in jobs:
        grouped.setdefault(job.posted_date or "unknown", []).append(job)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        summary.append([
            day, len(day_jobs),
            sum(1 for j in day_jobs if j.contact_info),
            max((j.title_rank for j in day_jobs), default=0),
        ])
    for col in range(1, 5):
        summary.column_dimensions[get_column_letter(col)].width = 22
    append_jobs_sheet(wb, "All Jobs", jobs)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        append_jobs_sheet(wb, day, sorted(day_jobs, key=lambda j: (not bool(j.contact_info), -j.title_rank, j.title.lower())))
    wb.save(path)


# ── CLI ───────────────────────────────────────────────────────────────────────

def load_terms(args: argparse.Namespace) -> list[str]:
    terms = list(args.terms or [])
    if args.terms_file:
        terms.extend(line.strip() for line in args.terms_file.read_text(encoding="utf-8").splitlines())
    return [t for t in terms if t.strip()] or DEFAULT_SEARCH_TERMS


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape Randstad USA jobs into CSV/JSON/Excel.")
    parser.add_argument("--term", action="append", dest="terms", help="Search term. Repeat for multiple.")
    parser.add_argument("--terms-file", type=Path)
    parser.add_argument("--posted-within-days", type=int, default=4)
    parser.add_argument("--keep-w2-f2f-onsite-interview", action="store_true")
    parser.add_argument("--min-hourly-rate", type=float, default=0, help="Min hourly rate filter (0=disabled).")
    parser.add_argument("--timeout", type=int, default=20)
    parser.add_argument("--sleep", type=float, default=0.3)
    parser.add_argument("--out-dir", type=Path, default=Path(__file__).resolve().parent / "output")
    parser.add_argument("--no-excel", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    jobs = scrape_randstad(
        load_terms(args),
        args.posted_within_days,
        not args.keep_w2_f2f_onsite_interview,
        args.min_hourly_rate or None,
        args.timeout,
        args.sleep,
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = args.out_dir / f"randstad_jobs_{timestamp}.csv"
    json_path = args.out_dir / f"randstad_jobs_{timestamp}.json"
    excel_path = args.out_dir / f"randstad_jobs_{timestamp}.xlsx"
    write_csv(jobs, csv_path)
    write_json(jobs, json_path)
    if not args.no_excel:
        write_excel(jobs, excel_path, args.posted_within_days, args.min_hourly_rate or None)
    print(f"\nSaved {len(jobs)} jobs")
    print(f"CSV:   {csv_path}")
    print(f"JSON:  {json_path}")
    if not args.no_excel:
        print(f"Excel: {excel_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
