#!/usr/bin/env python3
"""Brooksource job scraper — uses WP REST API to fetch all live job listings."""

from __future__ import annotations

import argparse
import csv
import html
import json
import re
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_URL = "https://jobs.brooksource.com/jobs"
REST_ENDPOINT = f"{BASE_URL}/wp-json/wp/v2/job-listings"
REST_FIELDS = "id,title,date,link,content,meta"

TITLE_RANKING_WEIGHTS = [
    ("machine learning engineer", 56),
    ("ml engineer", 54),
    ("ai engineer", 54),
    ("ai/ml", 54),
    ("aiml", 54),
    ("generative ai", 52),
    ("gen ai", 52),
    ("llm", 50),
    ("rag", 48),
    ("senior data scientist", 52),
    ("data scientist", 48),
    ("python developer", 46),
    ("python engineer", 46),
    ("senior python", 50),
    ("django", 44),
    ("fastapi", 44),
    ("backend", 38),
    ("back end", 38),
    ("api developer", 36),
    ("software engineer python", 46),
    ("full stack software engineer", 42),
    ("full stack engineer", 40),
    ("full stack developer", 40),
    ("data engineer", 38),
    ("etl", 32),
    ("cloud engineer", 30),
    ("aws", 28),
    ("azure", 28),
    ("senior software engineer", 32),
    ("software engineer", 24),
    ("software developer", 24),
    ("developer", 12),
    ("engineer", 10),
]
TITLE_EXCLUSION_WEIGHTS = [
    ("data engineer", -100),
    ("frontend", -30),
    ("front end", -30),
    ("ui developer", -20),
    ("ux/ui", -12),
    ("java", -12),
    (".net", -35),
    ("qa", -30),
    ("quality assurance", -30),
    ("desktop support", -40),
    ("help desk", -40),
    ("project manager", -35),
    ("business analyst", -30),
    ("coordinator", -35),
    ("technician", -35),
    ("field technician", -40),
    ("junior", -100),
    ("jr ", -100),
    ("entry level", -100),
]
TITLE_EXCLUSION_PATTERNS = [
    (re.compile(r"\bjava\b.*\bfull\s*stack\b|\bfull\s*stack\b.*\bjava\b|\bjava\b.*\bfullstack\b|\bfullstack\b.*\bjava\b", re.I), "Java full stack title"),
    (re.compile(r"\bjava\b.*\b(?:developer|engineer|architect|backend|software)\b|\b(?:developer|engineer|architect|backend|software)\b.*\bjava\b", re.I), "Java developer/engineer title"),
    (re.compile(r"\bdata\s+engineer\b", re.I), "Data Engineer title"),
    (re.compile(r"\bjunior\b|\bjr\.?\s", re.I), "Junior title"),
    (re.compile(r"\bentry[\s-]level\b", re.I), "Entry-level title"),
]
DISALLOWED_WORK_PATTERNS = [
    (re.compile(r"\bno\s+c2c\b", re.I), "No C2C"),
    (re.compile(r"\bno\s+corp(?:oration)?\s*[- ]?\s*to\s*[- ]?\s*corp(?:oration)?\b", re.I), "No corp-to-corp"),
    (re.compile(r"\bf\s*2\s*f\b", re.I), "F2F"),
    (re.compile(r"\bface\s*[- ]?\s*to\s*[- ]?\s*face\b", re.I), "Face-to-face"),
    (re.compile(r"\bin\s*[- ]?\s*person\s+interview\b", re.I), "In-person interview"),
    (re.compile(r"\bon\s*[- ]?\s*site\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\bonsite\s+interview\b", re.I), "Onsite interview"),
    (re.compile(r"\bfull\s*[- ]?\s*time\b", re.I), "Full-time"),
    (re.compile(r"\bpermanent\b", re.I), "Permanent"),
    (re.compile(r"\bdirect\s*[- ]?\s*hire\b", re.I), "Direct hire"),
    (re.compile(r"\bcontract\s*[- ]?\s*to\s*[- ]?\s*hire\b", re.I), "Contract-to-Hire"),
    (re.compile(r"\bcth\b", re.I), "Contract-to-Hire (CTH)"),
]
EMAIL_RE = re.compile(r"(?<![A-Za-z0-9._%+-])([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})(?![A-Za-z0-9._%+-])")
PHONE_RE = re.compile(r"(?<!\d)(?:\+?1[\s.-]?)?(?:\(?\d{3}\)?[\s.-]?)\d{3}[\s.-]?\d{4}(?!\d)")
MIN_TITLE_RANK = 20

# Patterns to extract salary/compensation from content
PAY_RE = re.compile(
    r"(?:compensation|salary|pay|rate)[:\s]*"
    r"\$?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]+)?)\s*(?:–|-|to)\s*\$?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]+)?)\s*(?:per\s+)?([a-z]+)?",
    re.I,
)
PAY_SINGLE_RE = re.compile(
    r"\$\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]+)?)\s*(?:(?:–|-)\$?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]+)?))?\s*(?:per\s+)?([a-z]+)?",
    re.I,
)
WORKSITE_RE = re.compile(r"\b(remote|hybrid|on.?site)\b", re.I)
EMP_TYPE_RE = re.compile(r"\b(contract.to.hire|cth|contract|temp.to.perm|permanent|direct.hire|full.time)\b", re.I)


@dataclass
class BrooksourceJob:
    source_company: str
    title_rank: int
    title_rank_reasons: str
    title: str
    city: str
    state: str
    location: str
    employment_type: str
    worksite: str
    is_remote: str
    min_pay_rate: str
    max_pay_rate: str
    pay_rate_unit: str
    posted_date: str
    job_id: str
    job_url: str
    contact_info: str
    description_snippet: str
    raw_text: str


def clean_html(html_text: str) -> str:
    text = html.unescape(html_text or "")
    text = re.sub(r"<[^>]+>", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def clean_text(text: str) -> str:
    return re.sub(r"\s+", " ", html.unescape(text or "")).strip()


def normalized_title(text: str) -> str:
    return re.sub(r"[^a-z0-9+#./-]+", " ", (text or "").lower()).strip()


def title_contains(title: str, phrase: str) -> bool:
    phrase = normalized_title(phrase)
    if len(phrase) <= 3:
        return re.search(rf"(?<![a-z0-9]){re.escape(phrase)}(?![a-z0-9])", title) is not None
    return phrase in title


def score_title(title: str) -> tuple[int, str]:
    normalized = normalized_title(title)
    score = 0
    reasons: list[str] = []
    for phrase, weight in TITLE_RANKING_WEIGHTS:
        if title_contains(normalized, phrase):
            score += weight
            reasons.append(f"{phrase}+{weight}")
    for phrase, weight in TITLE_EXCLUSION_WEIGHTS:
        if title_contains(normalized, phrase):
            score += weight
            reasons.append(f"{phrase}{weight}")
    return max(score, 0), "; ".join(reasons)


def title_exclusion_reasons(title: str) -> list[str]:
    return [reason for pattern, reason in TITLE_EXCLUSION_PATTERNS if pattern.search(title or "")]


def disallowed_work_reasons(text: str) -> list[str]:
    reasons: list[str] = []
    cleaned = clean_text(text)
    for pattern, reason in DISALLOWED_WORK_PATTERNS:
        if pattern.search(cleaned) and reason not in reasons:
            reasons.append(reason)
    return reasons


def extract_contact_info(text: str) -> str:
    emails = []
    for email in EMAIL_RE.findall(text):
        lowered = email.lower()
        if "staffingfuture" not in lowered and lowered not in emails:
            emails.append(lowered)
    phones = []
    for phone in PHONE_RE.findall(text or ""):
        digits = re.sub(r"\D", "", phone)
        if len(digits) == 11 and digits.startswith("1"):
            digits = digits[1:]
        if len(digits) == 10:
            display = f"({digits[:3]}) {digits[3:6]}-{digits[6:]}"
            if display not in phones:
                phones.append(display)
    return ", ".join(emails + phones)


def _infer_unit(raw_num: str, explicit_unit: str) -> str:
    unit = explicit_unit.lower()
    if "year" in unit or "annual" in unit or "salary" in unit:
        return "year"
    if unit in ("hr", "hour", "hourly"):
        return "hour"
    # No explicit unit — infer from magnitude
    try:
        val = float(raw_num.replace(",", ""))
        return "year" if val >= 10000 else "hour"
    except ValueError:
        return "hour"


def parse_pay(text: str) -> tuple[str, str, str]:
    m = PAY_RE.search(text)
    if m:
        lo = m.group(1).replace(",", "")
        hi = m.group(2).replace(",", "")
        unit = _infer_unit(lo, m.group(3) or "")
        return lo, hi, unit
    m = PAY_SINGLE_RE.search(text)
    if m:
        lo = m.group(1).replace(",", "")
        hi = (m.group(2) or m.group(1)).replace(",", "")
        unit = _infer_unit(lo, m.group(3) or "")
        return lo, hi, unit
    return "", "", ""


def parse_worksite(text: str) -> str:
    m = WORKSITE_RE.search(text)
    if not m:
        return ""
    val = m.group(1).lower()
    if "remote" in val:
        return "Remote"
    if "hybrid" in val:
        return "Hybrid"
    return "On-Site"


def parse_emp_type(text: str) -> str:
    m = EMP_TYPE_RE.search(text)
    if not m:
        return ""
    val = m.group(1).lower()
    if "contract-to-hire" in val or "cth" in val or "contract to hire" in val:
        return "Contract-to-Hire"
    if "contract" in val:
        return "Contract"
    if "temp" in val:
        return "Temp to Perm"
    if "permanent" in val or "direct hire" in val:
        return "Permanent"
    if "full-time" in val or "full time" in val:
        return "Full-time"
    return m.group(1).title()


def parse_location(raw: str) -> tuple[str, str]:
    raw = clean_text(raw)
    parts = [p.strip() for p in raw.split(",")]
    if len(parts) >= 2:
        return parts[0], parts[-1]
    return raw, ""


def is_within_posted_days(posted_iso: str, days: Optional[int]) -> bool:
    if not days or days <= 0 or not posted_iso:
        return True
    try:
        posted_dt = datetime.fromisoformat(posted_iso.replace("Z", "+00:00"))
        if posted_dt.tzinfo is None:
            posted_dt = posted_dt.replace(tzinfo=timezone.utc)
    except ValueError:
        return True
    now = datetime.now(timezone.utc)
    return 0 <= (now - posted_dt).total_seconds() <= days * 86400


def make_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
        "Accept": "application/json",
    })
    return session


def fetch_all_jobs(session: requests.Session, timeout: int) -> list[dict]:
    all_jobs: list[dict] = []
    page = 1
    while True:
        url = f"{REST_ENDPOINT}?per_page=100&_fields={REST_FIELDS}&orderby=date&order=desc&page={page}"
        response = session.get(url, timeout=timeout)
        response.raise_for_status()
        batch = response.json()
        if not batch:
            break
        all_jobs.extend(batch)
        total_pages = int(response.headers.get("X-WP-TotalPages", 1))
        if page >= total_pages:
            break
        page += 1
    return all_jobs


def normalize_job(raw: dict) -> BrooksourceJob:
    title = clean_html(raw["title"]["rendered"])
    title_rank, title_rank_reasons = score_title(title)
    content_html = raw.get("content", {}).get("rendered", "")
    text = clean_html(content_html)
    meta = raw.get("meta", {})
    location_raw = clean_text(meta.get("_job_location") or "")
    city, state = parse_location(location_raw)
    is_remote_meta = bool(meta.get("_remote_position"))
    worksite = parse_worksite(text)
    if not worksite and is_remote_meta:
        worksite = "Remote"
    is_remote = "Yes" if worksite == "Remote" or is_remote_meta else ("Hybrid" if worksite == "Hybrid" else "No")
    emp_type = parse_emp_type(text)
    min_pay, max_pay, pay_unit = parse_pay(text)
    posted = raw.get("date", "")[:10]
    return BrooksourceJob(
        source_company="Brooksource",
        title_rank=title_rank,
        title_rank_reasons=title_rank_reasons,
        title=title,
        city=city,
        state=state,
        location=location_raw,
        employment_type=emp_type,
        worksite=worksite,
        is_remote=is_remote,
        min_pay_rate=min_pay,
        max_pay_rate=max_pay,
        pay_rate_unit=pay_unit,
        posted_date=posted,
        job_id=str(raw.get("id", "")),
        job_url=clean_text(raw.get("link", "")),
        contact_info=extract_contact_info(text),
        description_snippet=text[:900],
        raw_text=text,
    )


def below_min_hourly(job: BrooksourceJob, min_hourly_rate: Optional[float]) -> bool:
    if not min_hourly_rate or min_hourly_rate <= 0 or job.pay_rate_unit != "hour":
        return False
    rates = [float(r) for r in (job.min_pay_rate, job.max_pay_rate) if r]
    return bool(rates) and max(rates) < min_hourly_rate


def sort_jobs(jobs: list[BrooksourceJob]) -> list[BrooksourceJob]:
    return sorted(
        jobs,
        key=lambda j: (j.posted_date, bool(j.contact_info), j.title_rank, j.title.lower()),
        reverse=True,
    )


def scrape_brooksource(
    posted_within_days: Optional[int],
    exclude_disallowed_work: bool,
    min_hourly_rate: Optional[float],
    timeout: int,
) -> list[BrooksourceJob]:
    session = make_session()
    print(f"Fetching all Brooksource jobs from REST API...")
    raw_jobs = fetch_all_jobs(session, timeout)
    print(f"  Found {len(raw_jobs)} total jobs")

    jobs: list[BrooksourceJob] = []
    for raw in raw_jobs:
        if not is_within_posted_days(raw.get("date", ""), posted_within_days):
            continue
        job = normalize_job(raw)
        if job.title_rank < MIN_TITLE_RANK:
            print(f"  Skipped low-rank ({job.title_rank}): {job.title}")
            continue
        excl = title_exclusion_reasons(job.title)
        if excl:
            print(f"  Excluded ({', '.join(excl)}): {job.title}")
            continue
        if exclude_disallowed_work:
            reasons = disallowed_work_reasons(" ".join([job.title, job.employment_type, job.raw_text]))
            if reasons:
                print(f"  Excluded ({', '.join(reasons)}): {job.title}")
                continue
        if below_min_hourly(job, min_hourly_rate):
            print(f"  Excluded (below ${min_hourly_rate:g}/hr): {job.title}")
            continue
        if job.pay_rate_unit == "year":
            print(f"  Excluded (annual salary = full-time signal): {job.title}")
            continue
        jobs.append(job)

    return sort_jobs(jobs)


# ── output helpers ────────────────────────────────────────────────────────────

def write_csv(jobs: list[BrooksourceJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = list(BrooksourceJob.__dataclass_fields__.keys())
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        for job in jobs:
            writer.writerow(asdict(job))


def write_json(jobs: list[BrooksourceJob], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        json.dump([asdict(job) for job in jobs], fh, indent=2)


EXCEL_COLUMNS = [
    ("posted_date", "Posted"),
    ("title_rank", "Rank"),
    ("title", "Title"),
    ("location", "Location"),
    ("employment_type", "Type"),
    ("worksite", "Worksite"),
    ("pay_range", "Pay"),
    ("contact_info", "Contact"),
    ("job_id", "Job ID"),
    ("job_url", "Job URL"),
    ("title_rank_reasons", "Rank Reasons"),
]


def excel_row(job: BrooksourceJob) -> dict[str, Any]:
    row = asdict(job)
    row["pay_range"] = ""
    if job.min_pay_rate or job.max_pay_rate:
        lo = job.min_pay_rate or job.max_pay_rate
        hi = job.max_pay_rate or job.min_pay_rate
        unit = f" / {job.pay_rate_unit}" if job.pay_rate_unit else ""
        row["pay_range"] = f"${lo}-${hi}{unit}"
    return row


def style_sheet(sheet) -> None:
    fill = PatternFill("solid", fgColor="1A3A5C")
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [13, 8, 36, 22, 16, 10, 18, 30, 14, 14, 45]
    for idx, w in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = w
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        rank = row[1].value or 0
        if isinstance(rank, int) and rank >= 45:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="E8F5E9")
        if row[7].value:
            row[7].fill = PatternFill("solid", fgColor="FFF3CD")


def append_jobs_sheet(wb: Workbook, sheet_name: str, jobs: list[BrooksourceJob]) -> None:
    safe_name = re.sub(r"[\[\]:*?/\\]", "-", sheet_name)[:31] or "Sheet"
    sheet = wb.create_sheet(safe_name)
    sheet.append([label for _, label in EXCEL_COLUMNS])
    for job in jobs:
        source = excel_row(job)
        sheet.append([source.get(key, "") for key, _ in EXCEL_COLUMNS])
        for col_idx, (key, _) in enumerate(EXCEL_COLUMNS, start=1):
            if key == "job_url":
                cell = sheet.cell(row=sheet.max_row, column=col_idx)
                if job.job_url:
                    cell.value = "Open Job"
                    cell.hyperlink = job.job_url
                    cell.style = "Hyperlink"
    style_sheet(sheet)


def write_excel(jobs: list[BrooksourceJob], path: Path, posted_within_days: Optional[int], min_hourly_rate: Optional[float]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary["A1"] = "Brooksource Daily Jobs"
    summary["A1"].font = Font(size=18, bold=True, color="1A3A5C")
    summary["A3"] = "Generated"
    summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    summary["A4"] = "Posting Window"
    summary["B4"] = "All dates" if not posted_within_days else f"Last {posted_within_days} days"
    summary["A5"] = "Min Hourly Pay"
    summary["B5"] = "Disabled" if not min_hourly_rate else f"${min_hourly_rate:g}/hour"
    summary["A6"] = "Total Jobs"
    summary["B6"] = len(jobs)
    summary.append([])
    summary.append(["Day", "Jobs", "With Contact", "Top Rank"])
    for cell in summary[8]:
        cell.fill = PatternFill("solid", fgColor="1A3A5C")
        cell.font = Font(color="FFFFFF", bold=True)
    grouped: dict[str, list[BrooksourceJob]] = {}
    for job in jobs:
        grouped.setdefault(job.posted_date or "unknown", []).append(job)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        summary.append([
            day, len(day_jobs),
            sum(1 for j in day_jobs if j.contact_info),
            max((j.title_rank for j in day_jobs), default=0),
        ])
    for col in range(1, 5):
        summary.column_dimensions[get_column_letter(col)].width = 22
    append_jobs_sheet(wb, "All Jobs", jobs)
    for day, day_jobs in sorted(grouped.items(), reverse=True):
        append_jobs_sheet(wb, day, sorted(day_jobs, key=lambda j: (not bool(j.contact_info), -j.title_rank, j.title.lower())))
    wb.save(path)


# ── CLI ───────────────────────────────────────────────────────────────────────

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Scrape Brooksource jobs into CSV/JSON/Excel.")
    parser.add_argument("--posted-within-days", type=int, default=7)
    parser.add_argument("--keep-w2-f2f-onsite-interview", action="store_true")
    parser.add_argument("--min-hourly-rate", type=float, default=0, help="Min hourly rate (0=disabled).")
    parser.add_argument("--timeout", type=int, default=20)
    parser.add_argument("--out-dir", type=Path, default=Path(__file__).resolve().parent / "output")
    parser.add_argument("--no-excel", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    jobs = scrape_brooksource(
        args.posted_within_days,
        not args.keep_w2_f2f_onsite_interview,
        args.min_hourly_rate or None,
        args.timeout,
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_path = args.out_dir / f"brooksource_jobs_{timestamp}.csv"
    json_path = args.out_dir / f"brooksource_jobs_{timestamp}.json"
    excel_path = args.out_dir / f"brooksource_jobs_{timestamp}.xlsx"
    write_csv(jobs, csv_path)
    write_json(jobs, json_path)
    if not args.no_excel:
        write_excel(jobs, excel_path, args.posted_within_days, args.min_hourly_rate or None)
    print(f"\nSaved {len(jobs)} jobs")
    print(f"CSV:   {csv_path}")
    print(f"JSON:  {json_path}")
    if not args.no_excel:
        print(f"Excel: {excel_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
