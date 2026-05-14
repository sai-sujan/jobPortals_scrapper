import json
import re
from datetime import datetime
from html import escape
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .storage import JobStore


SLOW_SITE_THRESHOLD_SECONDS = 15 * 60


JOB_HEADERS = [
    "Company",
    "Title",
    "Score",
    "Bucket",
    "Employment Type",
    "Location",
    "Remote",
    "Posted",
    "Contact Info",
    "First Seen",
    "Last Seen",
    "Matched Terms",
    "Link",
    "Snippet",
]


def row_to_job_values(row):
    return [
        row["source_company"],
        row["title"],
        row["score"],
        row["match_bucket"],
        row["employment_type"],
        row["location"],
        row["remote_status"],
        row["posted_date"],
        row["contact_info"] or "",
        row["first_seen_at"],
        row["last_seen_at"],
        row["matched_terms"],
        row["job_url"],
        row["description_snippet"],
    ]


def safe_table_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "", name)
    if not cleaned or cleaned[0].isdigit():
        cleaned = "T" + cleaned
    return cleaned[:240]


def safe_sheet_name(name: str, existing: Sequence[str]) -> str:
    cleaned = re.sub(r"[\[\]:*?/\\]", " ", name).strip() or "Sheet"
    cleaned = re.sub(r"\s+", " ", cleaned)[:31]
    candidate = cleaned
    index = 2
    while candidate in existing:
        suffix = f" {index}"
        candidate = cleaned[: 31 - len(suffix)] + suffix
        index += 1
    return candidate


def add_table(ws, name: str) -> None:
    if ws.max_row < 2:
        return
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    table = Table(displayName=safe_table_name(name), ref=ref)
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)


def format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    widths = {
        "A": 24,
        "B": 44,
        "C": 10,
        "D": 16,
        "E": 18,
        "F": 24,
        "G": 12,
        "H": 16,
        "I": 20,
        "J": 20,
        "K": 34,
        "L": 48,
        "M": 60,
        "N": 90,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        if len(row) >= 13 and row[12].value:
            row[12].hyperlink = row[12].value
            row[12].style = "Hyperlink"
        if len(row) >= 3 and isinstance(row[2].value, int) and row[2].value >= 60:
            row[2].fill = PatternFill("solid", fgColor="C6EFCE")


def write_jobs_sheet(wb: Workbook, title: str, rows: Iterable, table_name: str) -> None:
    ws = wb.create_sheet(safe_sheet_name(title, wb.sheetnames))
    ws.append(JOB_HEADERS)
    for row in rows:
        ws.append(row_to_job_values(row))
    format_sheet(ws)
    add_table(ws, table_name)


def placeholders(values: Sequence[str]) -> str:
    return ", ".join("?" for _ in values)


def query_priority_jobs(store: JobStore, priority_companies: Sequence[str], where: str = "1=1"):
    if not priority_companies:
        return []
    params = tuple(priority_companies)
    return store.query_jobs(f"source_company IN ({placeholders(priority_companies)}) AND {where}", params)


def display_duration(seconds) -> str:
    try:
        total = int(round(float(seconds or 0)))
    except (TypeError, ValueError):
        total = 0
    minutes, secs = divmod(total, 60)
    hours, minutes = divmod(minutes, 60)
    if hours:
        return f"{hours}h {minutes}m {secs}s"
    if minutes:
        return f"{minutes}m {secs}s"
    return f"{secs}s"


def export_excel(store: JobStore, run_id: int, profile: Dict, reports_dir: Path, priority_companies: Sequence[str] = ()) -> Path:
    reports_dir.mkdir(parents=True, exist_ok=True)
    now = datetime.now()
    dated_path = reports_dir / f"{now:%Y-%m-%d}_jobs.xlsx"
    latest_path = reports_dir / "jobs_latest.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    today_prefix = f"{now:%Y-%m-%d}"
    new_matches = store.query_jobs("status='match' AND first_seen_at LIKE ?", (today_prefix + "%",))
    best = store.query_jobs("status='match' AND match_bucket='best'")
    borderline = store.query_jobs("status='match' AND match_bucket='borderline'")
    confirmed_c2c = store.query_jobs("status='match' AND employment_type='C2C Contract'")
    c2c_review = store.query_jobs("status='match' AND employment_type LIKE 'C2C Review%'")
    all_links = store.query_jobs("1=1")

    priority_matches = query_priority_jobs(store, priority_companies, "status='match'")
    priority_confirmed = query_priority_jobs(store, priority_companies, "status='match' AND employment_type='C2C Contract'")
    priority_review = query_priority_jobs(store, priority_companies, "status='match' AND employment_type LIKE 'C2C Review%'")

    if priority_companies:
        write_priority_dashboard(wb, store, run_id, priority_companies, priority_matches, priority_confirmed, priority_review)
        write_jobs_sheet(wb, "Priority Matches", priority_matches, "PriorityMatches")
        write_jobs_sheet(wb, "Priority Confirmed C2C", priority_confirmed, "PriorityConfirmedC2C")
        write_jobs_sheet(wb, "Priority Needs Confirm", priority_review, "PriorityNeedsConfirm")

    write_jobs_sheet(wb, "New Matches", new_matches, "NewMatches")
    write_jobs_sheet(wb, "Best Matches", best, "BestMatches")
    write_jobs_sheet(wb, "Borderline Matches", borderline, "BorderlineMatches")
    write_jobs_sheet(wb, "Confirmed C2C", confirmed_c2c, "ConfirmedC2C")
    write_jobs_sheet(wb, "Needs C2C Confirm", c2c_review, "NeedsC2CConfirm")
    write_jobs_sheet(wb, "All Stored Links", all_links, "AllStoredLinks")

    daily = wb.create_sheet("Daily Summary")
    daily.append(["First Seen Date", "Day", "New Matches", "Best", "Borderline", "Stored Links", "Excluded"])
    for row in store.daily_summary():
        daily.append([row["first_seen_date"], row["day_name"], row["new_matches"], row["best"], row["borderline"], row["stored"], row["excluded"]])
    format_basic(daily, "DailySummary")

    company_daily = wb.create_sheet("Company By Day")
    company_daily.append(["First Seen Date", "Day", "Company", "New Matches", "Best", "Borderline", "Stored Links", "Excluded"])
    for row in store.company_daily_summary():
        company_daily.append(
            [
                row["first_seen_date"],
                row["day_name"],
                row["source_company"],
                row["new_matches"],
                row["best"],
                row["borderline"],
                row["stored"],
                row["excluded"],
            ]
        )
    format_basic(company_daily, "CompanyByDay")

    for company in store.companies():
        company_rows = store.query_jobs("source_company=? AND status='match'", (company,))
        if company_rows:
            write_jobs_sheet(wb, company, company_rows, f"{company}Links")

    summary = wb.create_sheet("Company Summary")
    summary.append(["Company", "Checked At", "Duration", "Discovered", "Matched", "Status", "Error"])
    problem_rows = []
    for row in store.latest_site_runs(run_id):
        summary.append(
            [
                row["source_company"],
                row["checked_at"],
                display_duration(row["duration_seconds"]),
                row["discovered_count"],
                row["matched_count"],
                row["status"],
                row["error"],
            ]
        )
        if row["status"] != "ok" or float(row["duration_seconds"] or 0) >= SLOW_SITE_THRESHOLD_SECONDS:
            problem_rows.append(row)
    format_basic(summary, "CompanySummary")

    problems = wb.create_sheet("Problem Sites")
    problems.append(["Company", "Checked At", "Duration", "Discovered", "Matched", "Status", "Error"])
    for row in problem_rows:
        problems.append(
            [
                row["source_company"],
                row["checked_at"],
                display_duration(row["duration_seconds"]),
                row["discovered_count"],
                row["matched_count"],
                row["status"],
                row["error"],
            ]
        )
    format_basic(problems, "ProblemSites")

    rules = wb.create_sheet("Profile Rules")
    rules.append(["Rule Group", "Weight", "Terms"])
    rules.append(["Must Have Any", "", ", ".join(profile["matching"]["must_have_any"])])
    for group_name, group in profile["matching"]["strong_terms"].items():
        rules.append([group_name, group["weight"], ", ".join(group["terms"])])
    rules.append(["Seniority Boost", profile["matching"]["seniority_boost"]["weight"], ", ".join(profile["matching"]["seniority_boost"]["terms"])])
    rules.append(["Contract Boost", profile["matching"]["contract_boost"]["weight"], ", ".join(profile["matching"]["contract_boost"]["terms"])])
    rules.append(["Exclusions", "", ", ".join(profile["matching"]["exclusions"])])
    format_basic(rules, "ProfileRules")

    wb.save(dated_path)
    wb.save(latest_path)
    export_website(store, reports_dir / "site", run_id, priority_companies)
    return latest_path


def write_priority_dashboard(
    wb: Workbook,
    store: JobStore,
    run_id: int,
    priority_companies: Sequence[str],
    priority_matches: Sequence,
    priority_confirmed: Sequence,
    priority_review: Sequence,
) -> None:
    ws = wb.create_sheet("Priority Dashboard")
    ws.append(["Metric", "Value"])
    priority_runs = [row for row in store.latest_site_runs(run_id) if row["source_company"] in priority_companies]
    ws.append(["Priority portals configured", len(priority_companies)])
    ws.append(["Priority portals checked this run", len(priority_runs)])
    ws.append(["Priority portals OK", sum(1 for row in priority_runs if row["status"] == "ok")])
    ws.append(["Priority portals timed out/failed", sum(1 for row in priority_runs if row["status"] != "ok")])
    ws.append(["Priority discovered this run", sum(int(row["discovered_count"] or 0) for row in priority_runs)])
    ws.append(["Priority matched this run", sum(int(row["matched_count"] or 0) for row in priority_runs)])
    ws.append(["All stored priority matches", len(priority_matches)])
    ws.append(["Confirmed C2C priority matches", len(priority_confirmed)])
    ws.append(["Needs C2C confirmation priority matches", len(priority_review)])
    format_basic(ws, "PriorityDashboard")

    status = wb.create_sheet("Priority Site Status")
    status.append(["Company", "Checked At", "Duration", "Discovered", "Matched", "Status", "Error"])
    by_company = {row["source_company"]: row for row in priority_runs}
    for company in priority_companies:
        row = by_company.get(company)
        if row:
            status.append(
                [
                    row["source_company"],
                    row["checked_at"],
                    display_duration(row["duration_seconds"]),
                    row["discovered_count"],
                    row["matched_count"],
                    row["status"],
                    row["error"],
                ]
            )
        else:
            status.append([company, "", "", 0, 0, "not_checked", ""])
    format_basic(status, "PrioritySiteStatus")


def format_basic(ws, table_name: str) -> None:
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for column in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(column)].width = min(80, max(14, len(str(ws.cell(1, column).value)) + 8))
    add_table(ws, table_name)


def export_website(store: JobStore, site_dir: Path, run_id: int = None, priority_companies: Sequence[str] = ()) -> Path:
    site_dir.mkdir(parents=True, exist_ok=True)
    jobs = [job_to_dict(row) for row in store.query_jobs("1=1")]
    daily = [dict(row) for row in store.daily_summary()]
    company_daily = [dict(row) for row in store.company_daily_summary()]
    latest_run_id = run_id if run_id is not None else store.latest_run_id()
    site_runs = [dict(row) for row in store.latest_site_runs(latest_run_id)] if latest_run_id else []
    data = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "jobs": jobs,
        "priorityCompanies": list(priority_companies),
        "daily": daily,
        "companyDaily": company_daily,
        "siteRuns": site_runs,
    }
    html = build_site_html(data)
    path = site_dir / "index.html"
    path.write_text(html, encoding="utf-8")
    return path


def job_to_dict(row) -> Dict:
    first_seen = row["first_seen_at"] or ""
    return {
        "company": row["source_company"],
        "title": row["title"],
        "score": row["score"],
        "bucket": row["match_bucket"],
        "status": row["status"],
        "employmentType": row["employment_type"] or "",
        "contactInfo": row["contact_info"] or "",
        "location": row["location"] or "",
        "remote": row["remote_status"] or "",
        "posted": row["posted_date"] or "",
        "firstSeen": first_seen,
        "firstSeenDate": first_seen[:10],
        "lastSeen": row["last_seen_at"] or "",
        "matchedTerms": row["matched_terms"] or "",
        "url": row["job_url"],
        "snippet": row["description_snippet"] or "",
    }


def build_site_html(data: Dict) -> str:
    payload = json.dumps(data, ensure_ascii=False)
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Venkata Dora Job Matches</title>
  <style>
    :root {{ color-scheme: light; --ink:#17202a; --muted:#667085; --line:#d0d5dd; --blue:#175cd3; --green:#067647; --amber:#b54708; --bg:#f6f8fb; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; font-family: Arial, Helvetica, sans-serif; color:var(--ink); background:var(--bg); }}
    header {{ padding:22px 28px; background:#fff; border-bottom:1px solid var(--line); position:sticky; top:0; z-index:2; }}
    h1 {{ margin:0 0 6px; font-size:24px; }}
    .sub {{ color:var(--muted); font-size:14px; }}
    main {{ padding:20px 28px 40px; }}
    .controls {{ display:grid; grid-template-columns: repeat(5, minmax(150px, 1fr)); gap:12px; margin-bottom:16px; }}
    label {{ display:grid; gap:5px; font-size:12px; font-weight:bold; color:#344054; }}
    select, input {{ height:38px; border:1px solid var(--line); border-radius:6px; padding:0 10px; background:#fff; color:var(--ink); }}
    .cards {{ display:grid; grid-template-columns: repeat(5, minmax(120px, 1fr)); gap:12px; margin:0 0 16px; }}
    .metric {{ background:#fff; border:1px solid var(--line); border-radius:8px; padding:14px; }}
    .metric b {{ display:block; font-size:24px; margin-bottom:3px; }}
    .metric span {{ color:var(--muted); font-size:12px; }}
    .layout {{ display:grid; grid-template-columns: 320px 1fr; gap:16px; align-items:start; }}
    .panel {{ background:#fff; border:1px solid var(--line); border-radius:8px; overflow:hidden; }}
    .panel h2 {{ font-size:16px; margin:0; padding:14px 16px; border-bottom:1px solid var(--line); }}
    table {{ width:100%; border-collapse:collapse; }}
    th, td {{ padding:10px 12px; border-bottom:1px solid #eaecf0; text-align:left; vertical-align:top; font-size:13px; }}
    th {{ background:#f9fafb; font-size:12px; color:#475467; position:sticky; top:93px; z-index:1; }}
    a {{ color:var(--blue); text-decoration:none; }}
    a:hover {{ text-decoration:underline; }}
    .badge {{ display:inline-block; border-radius:999px; padding:3px 8px; font-size:12px; font-weight:bold; background:#eef4ff; color:var(--blue); }}
    .best {{ background:#ecfdf3; color:var(--green); }}
    .borderline {{ background:#fffaeb; color:var(--amber); }}
    .excluded {{ background:#f2f4f7; color:#475467; }}
    .snippet {{ color:var(--muted); max-width:560px; }}
    .empty {{ padding:30px; color:var(--muted); text-align:center; }}
    @media (max-width: 1050px) {{ .controls,.cards,.layout {{ grid-template-columns:1fr; }} th {{ position:static; }} }}
  </style>
</head>
<body>
  <header>
    <h1>Venkata Dora Job Matches</h1>
    <div class="sub">Generated <span id="generated"></span>. Links are shown by first-seen date, so the same link is not repeated as new on later days.</div>
  </header>
  <main>
    <section class="controls">
      <label>Date <select id="dateFilter"></select></label>
      <label>Company <select id="companyFilter"></select></label>
      <label>Bucket <select id="bucketFilter"><option value="match">Matches only</option><option value="best">Best</option><option value="borderline">Borderline</option><option value="stored">Stored non-matches</option><option value="excluded">Excluded</option><option value="all">All links</option></select></label>
      <label>Search <input id="textFilter" placeholder="title, skill, location"></label>
      <label>View <select id="viewFilter"><option value="firstSeen">New on selected date</option><option value="lastSeen">Seen/checked on selected date</option><option value="allDates">All dates</option></select></label>
    </section>
    <section class="cards">
      <div class="metric"><b id="visibleCount">0</b><span>Visible links</span></div>
      <div class="metric"><b id="bestCount">0</b><span>Best matches</span></div>
      <div class="metric"><b id="borderlineCount">0</b><span>Borderline</span></div>
      <div class="metric"><b id="companyCount">0</b><span>Companies</span></div>
      <div class="metric"><b id="newDateCount">0</b><span>New selected day</span></div>
    </section>
    <section class="layout">
      <aside class="panel">
        <h2>Day By Day</h2>
        <table><thead><tr><th>Date</th><th>Day</th><th>Matches</th></tr></thead><tbody id="dailyRows"></tbody></table>
      </aside>
      <section class="panel">
        <h2>Links</h2>
        <div id="jobTable"></div>
      </section>
    </section>
  </main>
  <script>
    const DATA = {payload};
    const jobs = DATA.jobs;
    const $ = id => document.getElementById(id);
    function uniq(values) {{ return [...new Set(values.filter(Boolean))].sort(); }}
    function option(value, label=value) {{ return `<option value="${{escapeHtml(value)}}">${{escapeHtml(label)}}</option>`; }}
    function escapeHtml(value) {{ return String(value ?? '').replace(/[&<>"']/g, c => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[c])); }}
    function init() {{
      $('generated').textContent = DATA.generatedAt;
      const dates = uniq(jobs.map(j => j.firstSeenDate)).reverse();
      $('dateFilter').innerHTML = option('latest', 'Latest day') + option('all', 'All days') + dates.map(d => option(d)).join('');
      $('companyFilter').innerHTML = option('all', 'All companies') + uniq(jobs.map(j => j.company)).map(c => option(c)).join('');
      $('dailyRows').innerHTML = DATA.daily.map(d => `<tr><td><a href="#" data-date="${{escapeHtml(d.first_seen_date)}}">${{escapeHtml(d.first_seen_date)}}</a></td><td>${{escapeHtml(d.day_name)}}</td><td>${{d.new_matches}}</td></tr>`).join('');
      document.querySelectorAll('[data-date]').forEach(a => a.addEventListener('click', event => {{ event.preventDefault(); $('dateFilter').value = a.dataset.date; render(); }}));
      ['dateFilter','companyFilter','bucketFilter','textFilter','viewFilter'].forEach(id => $(id).addEventListener('input', render));
      render();
    }}
    function selectedDate() {{
      const value = $('dateFilter').value;
      if (value === 'latest') return uniq(jobs.map(j => j.firstSeenDate)).sort().pop() || '';
      return value;
    }}
    function render() {{
      const date = selectedDate();
      const company = $('companyFilter').value;
      const bucket = $('bucketFilter').value;
      const text = $('textFilter').value.toLowerCase().trim();
      const view = $('viewFilter').value;
      let visible = jobs.filter(j => {{
        if (date && date !== 'all' && view === 'firstSeen' && j.firstSeenDate !== date) return false;
        if (date && date !== 'all' && view === 'lastSeen' && !j.lastSeen.startsWith(date)) return false;
        if (company !== 'all' && j.company !== company) return false;
        if (bucket === 'match' && j.status !== 'match') return false;
        if (bucket !== 'all' && bucket !== 'match' && j.bucket !== bucket) return false;
        if (text) {{
          const haystack = `${{j.company}} ${{j.title}} ${{j.location}} ${{j.remote}} ${{j.employmentType}} ${{j.contactInfo}} ${{j.matchedTerms}} ${{j.snippet}}`.toLowerCase();
          if (!haystack.includes(text)) return false;
        }}
        return true;
      }});
      visible.sort((a,b) => b.score - a.score || String(b.firstSeen).localeCompare(a.firstSeen));
      $('visibleCount').textContent = visible.length;
      $('bestCount').textContent = visible.filter(j => j.bucket === 'best').length;
      $('borderlineCount').textContent = visible.filter(j => j.bucket === 'borderline').length;
      $('companyCount').textContent = uniq(visible.map(j => j.company)).length;
      $('newDateCount').textContent = jobs.filter(j => j.status === 'match' && j.firstSeenDate === date).length;
      $('jobTable').innerHTML = tableHtml(visible);
    }}
    function tableHtml(rows) {{
      if (!rows.length) return '<div class="empty">No links for these filters.</div>';
      return `<table><thead><tr><th>Company</th><th>Title</th><th>Score</th><th>First Seen</th><th>Details</th></tr></thead><tbody>${{rows.map(j => `
        <tr>
          <td>${{escapeHtml(j.company)}}<br><span class="badge ${{escapeHtml(j.bucket)}}">${{escapeHtml(j.bucket)}}</span></td>
          <td><a href="${{escapeHtml(j.url)}}" target="_blank" rel="noopener noreferrer">${{escapeHtml(j.title)}}</a><br><span class="snippet">${{escapeHtml(j.snippet).slice(0, 260)}}</span></td>
          <td>${{j.score}}</td>
          <td>${{escapeHtml(j.firstSeenDate)}}<br><span class="snippet">${{escapeHtml(j.lastSeen)}}</span></td>
          <td>${{escapeHtml([j.employmentType, j.remote, j.location].filter(Boolean).join(' | '))}}<br><span class="snippet">${{escapeHtml(j.contactInfo || j.matchedTerms)}}</span></td>
        </tr>`).join('')}}</tbody></table>`;
    }}
    init();
  </script>
</body>
</html>
"""
